#!/usr/bin/env python3
# migrate_bl_excel.py
# 把公司現行的 Excel 流水帳(得盛/BL 庫存表)轉成庫存管理系統的匯入檔。
#
# 用法:
#     python migrate_bl_excel.py "庫存表.xlsx" [工作表名稱] [輸出資料夾]
#     (工作表名稱預設 2026,輸出資料夾預設 migrate_out)
#
# 產出三個檔案:
#     products_import.csv   → 直接上傳到系統的「CSV 匯入 → 商品匯入」
#     suppliers.csv         → 供應商清單(僅供核對;匯入商品時會自動建立)
#     migration_report.csv  → 逐項對照報告,含每一項的處理方式與需人工確認的標記
#
# 設計決策(經使用者確認):
#   1. 沒有正式料號的物料 → 自動編 TMP-xxxxx 臨時料號,品名照舊保留,一項都不漏
#   2. 只搬「目前結存」當期初庫存,13 年歷史留在原 Excel 供查詢(批次帳才會乾淨)
#   3. 分類沿用原本的 Type 欄(電阻-0603、Cable、PCBA…),同事不必重新學
#
# 轉檔規則:
#   - 以「得盛料號」分組;沒有料號者以「品名規格」分組
#   - 目前庫存 = 該物料最後一列的「得盛」(結存)欄
#   - 櫃位、廠商、品名、Type 皆取最後一次出現的值(以最新狀態為準)
#   - 結存為「10米」「1捲」這類文字時,取出開頭數字;取不出來記 0 並標記待確認
#   - 結存為負數時匯入 0 並標記(系統不允許負庫存,須實地盤點確認)

import csv
import os
import re
import sys
from collections import OrderedDict

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("需要 openpyxl:請先執行 pip install -r requirements.txt")

# 來源表的欄位位置(依實際檔案確認,若日後表頭調整只需改這裡)
COL = {
    "date": 0, "type": 1, "kind": 2, "sku": 3, "name": 4, "project": 5,
    "balance": 6, "location": 7, "in": 8, "out": 9, "in2": 10, "out3": 11,
    "remark": 12, "supplier": 13, "maker": 14, "datecode": 15, "order_no": 16,
}
TMP_PREFIX = "TMP-"

# 櫃位欄裡這幾個值不是存放位置,而是「目前沒有庫存」的狀態標記,不可當儲位帶進系統。
# (其餘像「桌上」「防潮箱N層」「Joyce保管」「倉庫-進料暫存」都是真的位置,保留。)
NON_LOCATION = {"無庫存", "無在庫", "無", "-", "－", "x", "X", "0"}


def cell(row, idx):
    v = row[idx] if idx < len(row) else None
    if v is None:
        return ""
    return v.strip() if isinstance(v, str) else v


def to_qty(value):
    """把結存欄轉成整數;回傳 (數量, 註記)。文字如「10米」取開頭數字。"""
    if value in ("", None):
        return None, ""
    if isinstance(value, (int, float)):
        n = int(value)
        if n < 0:
            return 0, f"原結存為負數({n}),已匯入 0,請實地盤點確認"
        return n, ""
    text = str(value).strip()
    m = re.match(r"^-?\d+(?:\.\d+)?", text)
    if not m:
        return 0, f"結存欄為文字「{text}」無法判讀,已匯入 0,請人工確認"
    n = int(float(m.group()))
    if n < 0:
        return 0, f"原結存為負數({text}),已匯入 0,請實地盤點確認"
    return n, f"原結存為文字「{text}」,取數字 {n},請確認單位"


def read_ledger(path, sheet_name):
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        sys.exit(f"找不到工作表「{sheet_name}」;此檔案有:{wb.sheetnames}")
    rows = [r for r in wb[sheet_name].iter_rows(values_only=True)]
    wb.close()
    return [r for r in rows[1:] if any(c not in (None, "") for c in r)]


def build_materials(data):
    """把流水帳彙整成一項一筆的物料清單(取每項最後狀態)。"""
    materials = OrderedDict()
    skipped = 0
    for row in data:
        sku = str(cell(row, COL["sku"])).strip()
        name = str(cell(row, COL["name"])).strip()
        if not sku and not name:
            skipped += 1              # 料號與品名都空白的列,無從辨識,略過
            continue
        key = sku if sku else "NAME::" + name
        m = materials.setdefault(key, {
            "sku": sku, "name": name, "rows": 0, "names": set(),
            "type": "", "kind": "", "location": "", "supplier": "",
            "balance": None, "locations": set(), "suppliers": set(),
        })
        m["rows"] += 1
        if name:
            m["name"] = name          # 取最後一次出現的品名(最新)
            m["names"].add(name)
        for field, col in (("type", "type"), ("kind", "kind"),
                           ("location", "location"), ("supplier", "supplier")):
            v = str(cell(row, COL[col])).strip()
            if not v:
                continue
            if field == "location":
                if v in NON_LOCATION:   # 「無庫存」不是儲位,略過不覆蓋既有位置
                    continue
                m["locations"].add(v)
            if field == "supplier":
                m["suppliers"].add(v)
            m[field] = v
        bal = cell(row, COL["balance"])
        if bal not in ("", None):
            m["balance"] = bal        # 取最後一次有值的結存
    return materials, skipped


def assign_skus(materials):
    """沒有正式料號者編臨時料號;確保全體唯一。"""
    used = {m["sku"] for m in materials.values() if m["sku"]}
    seq = 0
    for m in materials.values():
        if m["sku"]:
            m["final_sku"] = m["sku"]
            m["sku_note"] = ""
            continue
        while True:
            seq += 1
            candidate = f"{TMP_PREFIX}{seq:05d}"
            if candidate not in used:
                break
        used.add(candidate)
        m["final_sku"] = candidate
        m["sku_note"] = "原始資料無料號,系統自動編臨時料號,請日後補正式料號"
    return materials


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    src = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else "2026"
    outdir = sys.argv[3] if len(sys.argv) > 3 else "migrate_out"
    os.makedirs(outdir, exist_ok=True)

    print(f"讀取:{src} [{sheet}]")
    data = read_ledger(src, sheet)
    print(f"  流水帳有效列數:{len(data)}")

    materials, skipped = build_materials(data)
    materials = assign_skus(materials)
    print(f"  彙整後不重複物料:{len(materials)} 項(略過無法辨識的列 {skipped} 列)")

    products, report, suppliers = [], [], OrderedDict()
    stats = {"有料號": 0, "臨時料號": 0, "有庫存": 0, "零庫存": 0, "需人工確認": 0}

    for m in materials.values():
        qty, qty_note = to_qty(m["balance"])
        if qty is None:
            qty, qty_note = 0, "原始資料無結存欄位,已匯入 0"
        notes = [n for n in (m["sku_note"], qty_note) if n]
        if len(m["names"]) > 1:
            notes.append(f"此料號有 {len(m['names'])} 種品名,已取最新;其餘:"
                         + " / ".join(sorted(x for x in m["names"] if x != m["name"])[:2]))
        if len(m["locations"]) > 1:
            notes.append(f"曾出現 {len(m['locations'])} 個櫃位,已取最新({m['location']})")
        if len(m["suppliers"]) > 1:
            notes.append(f"曾有 {len(m['suppliers'])} 家廠商,已取最新({m['supplier']})")

        if m["supplier"]:
            suppliers.setdefault(m["supplier"], 0)
            suppliers[m["supplier"]] += 1

        products.append([
            m["final_sku"],                       # SKU
            m["name"] or m["final_sku"],          # 名稱(品名空白時用料號墊)
            m["type"],                            # 分類 = 原本的 Type
            "個",                                  # 單位(原始資料無單位欄,統一預設)
            "",                                   # 單價(流水帳無價格欄)
            "",                                   # 低庫存門檻(日後用「存貨規劃」推導)
            m["supplier"],                        # 供應商(不存在時系統自動建立)
            str(qty),                             # 初始庫存
            m["location"],                        # 儲位
        ])
        report.append([
            m["final_sku"], m["sku"] or "(無)", m["name"], m["type"], m["kind"],
            m["location"], m["supplier"], str(qty), str(m["rows"]),
            ";".join(notes),
        ])
        stats["有料號" if m["sku"] else "臨時料號"] += 1
        stats["有庫存" if qty > 0 else "零庫存"] += 1
        if notes:
            stats["需人工確認"] += 1

    def write(name, header, rows):
        path = os.path.join(outdir, name)
        with open(path, "w", encoding="utf-8-sig", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(header)
            w.writerows(rows)
        print(f"  產出:{path}({len(rows)} 列)")

    write("products_import.csv",
          ["SKU", "名稱", "分類", "單位", "單價", "低庫存門檻", "供應商", "初始庫存", "儲位"],
          products)
    write("suppliers.csv", ["廠商", "物料數"],
          [[k, str(v)] for k, v in sorted(suppliers.items(), key=lambda x: -x[1])])
    write("migration_report.csv",
          ["系統料號", "原始料號", "品名", "分類(Type)", "物料種", "櫃位", "廠商",
           "期初庫存", "原始異動列數", "備註"],
          report)

    print("\n===== 轉檔摘要 =====")
    print(f"  物料總數      {len(materials)}")
    print(f"    有正式料號  {stats['有料號']}")
    print(f"    臨時料號    {stats['臨時料號']}(TMP- 開頭,日後補正式料號)")
    print(f"  期初有庫存    {stats['有庫存']}")
    print(f"  期初零庫存    {stats['零庫存']}")
    print(f"  供應商        {len(suppliers)}")
    print(f"  需人工確認    {stats['需人工確認']}(詳見 migration_report.csv 備註欄)")
    print("\n下一步:用管理員帳號登入系統 →「系統管理 → CSV 匯入」→ 選「商品匯入」")
    print(f"        上傳 {os.path.join(outdir, 'products_import.csv')}")


if __name__ == "__main__":
    main()

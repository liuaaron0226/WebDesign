# inventory_app.py
# 庫存管理系統(完整版)— 單檔 Flask 應用
# 功能:多使用者帳號、商品管理、供應商管理、入庫/出庫、庫存查詢與搜尋、
#       低庫存警示、異動歷史、進出統計報表、CSV 匯出、
#       跨公司料號對照(別名)、物料照片、以圖搜圖(dHash)、CSV 匯入、內網 IP 白名單
# 與 patrick_method_solver.py 完全獨立,互不 import。

import csv
import io
import math
import os
import secrets
import shutil
import socket
import sqlite3
import sys
import threading
import time
from datetime import datetime, timedelta, timezone
from functools import wraps

from flask import (Flask, Response, g, redirect, render_template_string,
                   request, send_from_directory, session, url_for)
from markupsafe import Markup, escape
from werkzeug.security import check_password_hash, generate_password_hash

# Pillow 供以圖搜圖使用;缺席時 app 仍可啟動,僅該功能停用
try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# qrcode 供料架標籤使用;同樣採優雅停用
try:
    import qrcode
    HAS_QRCODE = True
except ImportError:
    HAS_QRCODE = False

# openpyxl 供 Excel(.xlsx)匯入使用;缺席時 CSV 匯入照常運作
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 版本標示:此系統以「下載 ZIP 覆蓋」的方式更新,畫面上看不出跑的是哪一版時,
# 使用者會誤以為舊版是新版(實際發生過:舊版匯入器只讀 8 欄,靜默丟掉儲位欄)。
# 每次發版時更新此字串,頁尾與啟動訊息都會顯示。
APP_VERSION = "2026.09.05d"

app = Flask(__name__)

# 資料庫路徑可用環境變數覆蓋,驗收測試用 /tmp 下的乾淨 DB
DB_PATH = os.environ.get("INVENTORY_DB", "inventory.db")
DATA_DIR = os.path.dirname(os.path.abspath(DB_PATH))

# 物料照片存放目錄(gitignored),同樣可用環境變數覆蓋
IMAGE_DIR = os.path.abspath(os.environ.get("INVENTORY_IMAGES", "inventory_images"))

# 備份目錄:本機固定備到 DB 旁的 backups/;另可設 BACKUP_DIR 同步複製到共用槽/NAS
BACKUP_DIR = os.path.join(DATA_DIR, "backups")
EXTRA_BACKUP_DIR = os.environ.get("BACKUP_DIR", "").strip()
BACKUP_KEEP = 14              # 本機保留份數
BACKUP_INTERVAL_SEC = 24 * 3600

ALLOWED_IMAGE_EXTS = {"jpg", "jpeg", "png", "gif", "webp", "bmp"}

# 上傳大小上限 10MB:照片與 CSV 皆足夠,可擋下把記憶體吃滿的超大檔
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024

# 數值合理上限:避免超大整數在寫入 SQLite 時丟 OverflowError 造成 500
MAX_QUANTITY = 1_000_000_000

# 顯示時區:資料庫一律存 UTC(正確做法),顯示時轉台灣時間 UTC+8
LOCAL_TZ = timezone(timedelta(hours=8))

# 內網白名單:設定 ALLOWED_IPS(逗號分隔)後,只有名單內來源可存取。
# 未設定 = 功能關閉(內網部署靠網路隔離本身,不需要此機制)。
ALLOWED_IPS = {ip.strip() for ip in os.environ.get("ALLOWED_IPS", "").split(",") if ip.strip()}
# 只有確實部署在可信反向代理後方(如 Render)才可開啟;直連部署務必維持關閉,
# 否則任何人都能自行送出 X-Forwarded-For 標頭假冒白名單 IP。
TRUST_PROXY = os.environ.get("TRUST_PROXY", "").strip().lower() in ("1", "true", "yes", "on")


def load_secret_key():
    # 絕不使用可預測的硬編碼金鑰(否則任何人都能離線偽造 session)。
    # 優先讀環境變數;否則在資料目錄產生一把持久隨機金鑰,達成零設定又不可預測。
    env_key = os.environ.get("SECRET_KEY", "").strip()
    if env_key:
        return env_key
    key_path = os.path.join(DATA_DIR, "secret_key.txt")
    try:
        with open(key_path, "r", encoding="ascii") as fh:
            key = fh.read().strip()
            if key:
                return key
    except OSError:
        pass
    key = secrets.token_hex(32)
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(key_path, "w", encoding="ascii") as fh:
            fh.write(key)
        os.chmod(key_path, 0o600)
    except OSError:
        pass  # 無法寫檔時仍以本次隨機金鑰運行(重啟後 session 失效,但不可偽造)
    return key


app.secret_key = load_secret_key()


@app.before_request
def restrict_to_allowed_ips():
    # 來源 IP 預設取實際連線位址;唯有明確設定 TRUST_PROXY 才採信
    # X-Forwarded-For(且取最後一段,即最靠近本機的代理所填入的值)。
    if not ALLOWED_IPS:
        return None
    client_ip = request.remote_addr or ""
    if TRUST_PROXY:
        xff = request.headers.get("X-Forwarded-For", "")
        if xff:
            client_ip = xff.split(",")[-1].strip()
    if client_ip not in ALLOWED_IPS:
        return Response(
            "<h1>403 禁止存取</h1><p>此系統僅限公司內部網路使用。</p>",
            status=403, mimetype="text/html")
    return None


@app.errorhandler(413)
def handle_too_large(err):
    return Response(
        "<h1>413 檔案過大</h1><p>上傳檔案不可超過 10MB,請縮小檔案後再試。</p>",
        status=413, mimetype="text/html")


# ---------------------------------------------------------------------------
# 資料庫連線與初始化
# ---------------------------------------------------------------------------

def get_db():
    # 每個 request 一條連線,存在 flask.g,teardown 時關閉
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH, timeout=15)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
        # WAL:讀取不再阻塞寫入。修正前「有人開歷史頁時他人入庫噴 500 掉資料」
        # 的並發問題;busy_timeout 讓寫入排隊而非立刻報錯。
        g.db.execute("PRAGMA journal_mode = WAL")
        g.db.execute("PRAGMA busy_timeout = 15000")
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    # CREATE TABLE IF NOT EXISTS 為冪等操作,啟動時自動建表
    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.execute("PRAGMA journal_mode = WAL")  # WAL 是 DB 持久屬性,設一次即可
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            username      TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            is_admin      INTEGER NOT NULL DEFAULT 0,
            created_at    TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS suppliers (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT NOT NULL,
            contact    TEXT DEFAULT '',
            phone      TEXT DEFAULT '',
            note       TEXT DEFAULT '',
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS products (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            name                TEXT NOT NULL,
            sku                 TEXT NOT NULL UNIQUE,
            category            TEXT DEFAULT '',
            unit                TEXT DEFAULT '個',
            unit_price          REAL NOT NULL DEFAULT 0,
            low_stock_threshold INTEGER NOT NULL DEFAULT 0,
            quantity            INTEGER NOT NULL DEFAULT 0 CHECK (quantity >= 0),
            supplier_id         INTEGER REFERENCES suppliers(id) ON DELETE SET NULL,
            created_at          TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS transactions (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL REFERENCES products(id),
            user_id    INTEGER NOT NULL REFERENCES users(id),
            type       TEXT NOT NULL CHECK (type IN ('in','out')),
            quantity   INTEGER NOT NULL CHECK (quantity > 0),
            note       TEXT DEFAULT '',
            created_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_tx_product ON transactions(product_id);
        CREATE INDEX IF NOT EXISTS idx_tx_created ON transactions(created_at);
        CREATE TABLE IF NOT EXISTS part_aliases (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL REFERENCES products(id) ON DELETE CASCADE,
            company    TEXT NOT NULL,
            alias_sku  TEXT NOT NULL,
            note       TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            UNIQUE(company, alias_sku)
        );
        CREATE INDEX IF NOT EXISTS idx_alias_sku ON part_aliases(alias_sku);
        CREATE INDEX IF NOT EXISTS idx_alias_product ON part_aliases(product_id);
        CREATE TABLE IF NOT EXISTS product_images (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id    INTEGER NOT NULL REFERENCES products(id) ON DELETE CASCADE,
            filename      TEXT NOT NULL,
            original_name TEXT DEFAULT '',
            phash         TEXT NOT NULL DEFAULT '',
            created_at    TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_img_product ON product_images(product_id);
        -- 批次管理(Lot Tracking):每筆入庫一批,出庫依 FIFO 消耗並記錄追溯
        CREATE TABLE IF NOT EXISTS lots (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id     INTEGER NOT NULL REFERENCES products(id) ON DELETE CASCADE,
            transaction_id INTEGER REFERENCES transactions(id),
            lot_no         TEXT NOT NULL,
            qty_received   INTEGER NOT NULL CHECK (qty_received > 0),
            qty_remaining  INTEGER NOT NULL CHECK (qty_remaining >= 0),
            unit_cost      REAL,
            note           TEXT DEFAULT '',
            received_at    TEXT NOT NULL,
            UNIQUE(product_id, lot_no)
        );
        CREATE INDEX IF NOT EXISTS idx_lots_product ON lots(product_id);
        CREATE TABLE IF NOT EXISTS lot_consumptions (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            transaction_id INTEGER NOT NULL REFERENCES transactions(id),
            lot_id         INTEGER NOT NULL REFERENCES lots(id),
            quantity       INTEGER NOT NULL CHECK (quantity > 0)
        );
        CREATE INDEX IF NOT EXISTS idx_lc_tx ON lot_consumptions(transaction_id);
        -- 異動歷史每筆 in 都要回查建立的批號,無此索引在兩萬筆時會全表掃描(實測 6 秒)
        CREATE INDEX IF NOT EXISTS idx_lots_tx ON lots(transaction_id);
        -- 稽核軌跡:誰在什麼時候改了什麼(建檔維護與管理操作皆留痕)
        CREATE TABLE IF NOT EXISTS audit_log (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER REFERENCES users(id),
            username    TEXT DEFAULT '',
            action      TEXT NOT NULL,
            target_type TEXT DEFAULT '',
            target_id   TEXT DEFAULT '',
            detail      TEXT DEFAULT '',
            created_at  TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_audit_created ON audit_log(created_at);
        -- 預留:現貨量之外的第二個數字,可用量 = 現貨 − 有效預留
        CREATE TABLE IF NOT EXISTS reservations (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id  INTEGER NOT NULL REFERENCES products(id) ON DELETE CASCADE,
            quantity    INTEGER NOT NULL CHECK (quantity > 0),
            purpose     TEXT DEFAULT '',
            username    TEXT DEFAULT '',
            status      TEXT NOT NULL DEFAULT 'active' CHECK (status IN ('active','released')),
            created_at  TEXT NOT NULL,
            released_at TEXT DEFAULT ''
        );
        CREATE INDEX IF NOT EXISTS idx_resv_product ON reservations(product_id, status);
        -- 循環盤點:盤點單 + 逐項實盤數,過帳後產生調整異動修正帳
        CREATE TABLE IF NOT EXISTS stock_counts (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT NOT NULL,
            scope      TEXT DEFAULT '',
            status     TEXT NOT NULL DEFAULT 'open' CHECK (status IN ('open','posted')),
            username   TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            posted_at  TEXT DEFAULT '',
            accuracy   REAL
        );
        CREATE TABLE IF NOT EXISTS stock_count_items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            count_id    INTEGER NOT NULL REFERENCES stock_counts(id) ON DELETE CASCADE,
            product_id  INTEGER NOT NULL REFERENCES products(id) ON DELETE CASCADE,
            system_qty  INTEGER NOT NULL DEFAULT 0,
            counted_qty INTEGER,
            note        TEXT DEFAULT '',
            counted_at  TEXT DEFAULT '',
            UNIQUE(count_id, product_id)
        );
        CREATE INDEX IF NOT EXISTS idx_sci_count ON stock_count_items(count_id);
        -- 收貨單(ASN,預先到貨通知):供應商檔案先進系統成為待核對單據,
        -- 放行前完全不動庫存;放行時才產生 in 異動與批次
        CREATE TABLE IF NOT EXISTS receipts (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            ref_no        TEXT DEFAULT '',
            supplier_id   INTEGER REFERENCES suppliers(id) ON DELETE SET NULL,
            supplier_name TEXT DEFAULT '',
            source        TEXT DEFAULT '',
            status        TEXT NOT NULL DEFAULT 'open'
                          CHECK (status IN ('open','posted','cancelled')),
            note          TEXT DEFAULT '',
            username      TEXT DEFAULT '',
            created_at    TEXT NOT NULL,
            posted_at     TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS receipt_items (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            receipt_id   INTEGER NOT NULL REFERENCES receipts(id) ON DELETE CASCADE,
            line_no      INTEGER NOT NULL DEFAULT 0,
            raw_sku      TEXT DEFAULT '',
            raw_name     TEXT DEFAULT '',
            product_id   INTEGER REFERENCES products(id) ON DELETE SET NULL,
            match_type   TEXT DEFAULT 'none',
            match_note   TEXT DEFAULT '',
            expected_qty INTEGER NOT NULL DEFAULT 0,
            received_qty INTEGER,
            lot_no       TEXT DEFAULT '',
            expiry_date  TEXT DEFAULT '',
            unit_cost    REAL,
            note         TEXT DEFAULT '',
            checked_at   TEXT DEFAULT ''
        );
        CREATE INDEX IF NOT EXISTS idx_ritems_receipt ON receipt_items(receipt_id);
        -- 採購訂單:已下訂 → 已出貨 → 已到貨待驗 → 已入庫(結案)。
        -- 訂了但還沒入庫的量叫「在途」,它讓人在下單前看得到「其實已經在路上了」
        CREATE TABLE IF NOT EXISTS purchase_orders (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            po_no         TEXT DEFAULT '',
            supplier_id   INTEGER REFERENCES suppliers(id) ON DELETE SET NULL,
            supplier_name TEXT DEFAULT '',
            status        TEXT NOT NULL DEFAULT 'ordered'
                          CHECK (status IN ('ordered','shipped','arrived','closed','cancelled')),
            eta           TEXT DEFAULT '',
            shipped_at    TEXT DEFAULT '',
            tracking_no   TEXT DEFAULT '',
            arrived_at    TEXT DEFAULT '',
            closed_at     TEXT DEFAULT '',
            note          TEXT DEFAULT '',
            username      TEXT DEFAULT '',
            created_at    TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS purchase_order_items (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            po_id        INTEGER NOT NULL REFERENCES purchase_orders(id) ON DELETE CASCADE,
            line_no      INTEGER NOT NULL DEFAULT 0,
            raw_sku      TEXT DEFAULT '',
            raw_name     TEXT DEFAULT '',
            product_id   INTEGER REFERENCES products(id) ON DELETE SET NULL,
            match_type   TEXT DEFAULT 'none',
            match_note   TEXT DEFAULT '',
            ordered_qty  INTEGER NOT NULL DEFAULT 0,
            received_qty INTEGER NOT NULL DEFAULT 0,
            unit_cost    REAL,
            eta          TEXT DEFAULT '',
            note         TEXT DEFAULT ''
        );
        CREATE INDEX IF NOT EXISTS idx_poitems_po ON purchase_order_items(po_id);
        CREATE INDEX IF NOT EXISTS idx_poitems_product ON purchase_order_items(product_id);
    """)
    # 收貨單回指採購單:到貨時自動生成的收貨單要知道自己來自哪一張訂單
    ensure_column(conn, "receipts", "po_id", "INTEGER")
    # 既有資料庫的欄位擴充(SQLite 的 ADD COLUMN 重複執行會報錯,故先檢查)
    ensure_column(conn, "transactions", "reverses", "INTEGER")
    ensure_column(conn, "products", "location", "TEXT DEFAULT ''")
    ensure_column(conn, "products", "purchase_unit", "TEXT DEFAULT ''")
    ensure_column(conn, "products", "units_per_purchase", "INTEGER DEFAULT 1")
    ensure_column(conn, "products", "lead_time_days", "INTEGER DEFAULT 7")
    ensure_column(conn, "products", "service_level", "REAL DEFAULT 95")
    ensure_column(conn, "products", "issue_strategy", "TEXT DEFAULT 'FIFO'")
    ensure_column(conn, "lots", "expiry_date", "TEXT DEFAULT ''")
    ensure_column(conn, "transactions", "purpose", "TEXT DEFAULT ''")
    # 待辦帶的四個彙總各自吃得到索引;不做快取,因為快取會讓使用者放行完
    # 回頭看到數字沒變、以為沒存到而再按一次——把不存在的效能問題換成真的正確性問題
    conn.execute("CREATE INDEX IF NOT EXISTS idx_receipts_status ON receipts(status)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_poi_open ON purchase_order_items(po_id, ordered_qty, received_qty)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_products_threshold ON products(low_stock_threshold)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_products_location ON products(location)")
    conn.commit()
    conn.close()
    os.makedirs(IMAGE_DIR, exist_ok=True)
    reconcile_lots()


def ensure_column(conn, table, column, decl):
    # 冪等地新增欄位:舊資料庫升級時自動補上,已存在則略過
    cols = {r[1] for r in conn.execute(f"PRAGMA table_info({table})")}
    if column not in cols:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {decl}")


def audit(action, target_type="", target_id="", detail=""):
    # 寫入稽核軌跡;必須在呼叫端的交易內(由呼叫端 commit),失敗不可影響主流程
    try:
        get_db().execute("""
            INSERT INTO audit_log (user_id, username, action, target_type, target_id, detail, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (session.get("user_id"), session.get("username", ""), action,
              target_type, str(target_id), detail, now_str()))
    except Exception:
        pass


def backup_db():
    # 用 sqlite3 的 backup API 取得一致性快照(WAL 模式下直接複製檔案並不安全)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    name = f"inventory_{stamp}.db"
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        target = os.path.join(BACKUP_DIR, name)
        src = sqlite3.connect(DB_PATH, timeout=15)
        dst = sqlite3.connect(target)
        with dst:
            src.backup(dst)
        dst.close()
        src.close()
        # 只保留最近 BACKUP_KEEP 份,避免磁碟被無限佔用
        backups = sorted(f for f in os.listdir(BACKUP_DIR)
                         if f.startswith("inventory_") and f.endswith(".db"))
        for old in backups[:-BACKUP_KEEP]:
            try:
                os.remove(os.path.join(BACKUP_DIR, old))
            except OSError:
                pass
        # 另備一份到共用槽/NAS(設了 BACKUP_DIR 才做);失敗不影響本機備份
        if EXTRA_BACKUP_DIR:
            try:
                os.makedirs(EXTRA_BACKUP_DIR, exist_ok=True)
                shutil.copy2(target, os.path.join(EXTRA_BACKUP_DIR, name))
            except OSError as exc:
                print(f"[備份] 共用槽備份失敗({EXTRA_BACKUP_DIR}):{exc}")
        return target
    except Exception as exc:
        print(f"[備份] 失敗:{exc}")
        return None


def start_backup_thread():
    # 啟動時先備一份,之後每 24 小時一次(daemon thread,關閉服務即結束)
    def loop():
        while True:
            time.sleep(BACKUP_INTERVAL_SEC)
            backup_db()
    backup_db()
    threading.Thread(target=loop, daemon=True).start()


def reconcile_lots():
    # 期初自動補批:舊資料庫升級時,若商品現時庫存大於批次剩餘總和,
    # 以「期初批」補齊缺口,確保批次帳與總帳恆一致(批次剩餘總和 = quantity)
    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT p.id, p.quantity,
               COALESCE((SELECT SUM(l.qty_remaining) FROM lots l WHERE l.product_id = p.id), 0) AS lot_sum
        FROM products p
    """).fetchall()
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    for r in rows:
        gap = r["quantity"] - r["lot_sum"]
        if gap > 0:
            conn.execute("""
                INSERT OR IGNORE INTO lots (product_id, lot_no, qty_received, qty_remaining, note, received_at)
                VALUES (?, ?, ?, ?, '期初庫存自動補批', ?)
            """, (r["id"], f"INIT-{r['id']}", gap, gap, stamp))
    conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# 以圖搜圖:dHash 感知雜湊(9x8 灰階縮圖 → 64-bit 指紋)+ Hamming 距離
# ---------------------------------------------------------------------------

def compute_dhash(stream):
    # 回傳 16 位 hex 字串;讀不出圖片時丟出例外由呼叫端處理
    img = Image.open(stream).convert("L").resize((9, 8), Image.LANCZOS)
    px = list(img.getdata())
    bits = 0
    for row in range(8):
        for col in range(8):
            bits = (bits << 1) | (1 if px[row * 9 + col] > px[row * 9 + col + 1] else 0)
    return f"{bits:016x}"


def hamming_distance(h1, h2):
    return bin(int(h1, 16) ^ int(h2, 16)).count("1")


# 虛擬網卡的位址範圍。公司電腦常裝了 Docker、VirtualBox、WSL 或 VPN,
# 它們各自會多一張網卡;拿到那種位址貼給同事,同事永遠連不上而且完全看不出原因。
VIRTUAL_IP_PREFIXES = tuple(
    ["127.", "169.254.", "192.168.56.", "192.168.99.", "198.18."]
    + [f"172.{n}." for n in range(17, 32)]      # Docker 預設橋接網段
)


def is_usable_lan_ip(ip):
    return bool(ip) and not ip.startswith(VIRTUAL_IP_PREFIXES)


def local_ip():
    """取本機在區域網路上的 IP(供同事連線用);取不到回空字串。
    連 UDP socket 不會真的送出封包,只是讓作業系統選出對外的那張網卡。"""
    for probe in ("192.168.1.1", "10.0.0.1", "8.8.8.8"):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            try:
                s.connect((probe, 1))
                ip = s.getsockname()[0]
            finally:
                s.close()
        except OSError:
            continue
        if is_usable_lan_ip(ip):
            return ip
    try:   # 離線機器上上一段會失敗,退而用主機名稱解析
        for info in socket.getaddrinfo(socket.gethostname(), None, socket.AF_INET):
            if is_usable_lan_ip(info[4][0]):
                return info[4][0]
    except OSError:
        pass
    return ""


def lan_url(port=None):
    """同事要輸入的那一行網址。抓不到內網位址時回空字串,呼叫端自己決定怎麼講。"""
    port = port or int(os.environ.get("PORT", 5000))
    ip = local_ip()
    return f"http://{ip}:{port}" if ip else ""


def has_any_user():
    """資料庫裡有沒有任何帳號。沒有的話 /register 是開著的,
    而在公司內網「開著」等於任何同事都能搶到管理員。"""
    con = sqlite3.connect(DB_PATH)
    try:
        row = con.execute("SELECT COUNT(*) FROM users").fetchone()
        return bool(row and row[0])
    except sqlite3.Error:
        return False
    finally:
        con.close()


def public_base_url():
    """料架標籤的 QR 要給手機掃,所以網址不能是 localhost。
    管理員多半是在伺服器那台電腦上開 http://localhost:5000 按列印
    (start_inventory.bat 就是自動開這個位址),而 localhost 在手機上指向手機自己,
    整批印出來的標籤會一張都掃不開。本機位址一律換成內網 IP;
    要固定成別的位址(例如有 DNS 名稱)就設 PUBLIC_BASE_URL。"""
    fixed = os.environ.get("PUBLIC_BASE_URL", "").strip().rstrip("/")
    if fixed:
        return fixed
    host, _, port = request.host.partition(":")
    if host in ("localhost", "127.0.0.1", "::1", "0.0.0.0"):
        ip = local_ip()
        if ip:
            return f"http://{ip}" + (f":{port}" if port else "")
    return request.host_url.rstrip("/")


def now_str():
    # 資料庫一律存 UTC 字串(跨時區正確、可直接字串比較);顯示時才轉台灣時間
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def parse_utc(s):
    # 把 DB 內的 UTC 字串轉成有時區資訊的 datetime
    return datetime.strptime(s, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)


def fmt_local(s):
    # UTC 字串 → 台灣時間顯示字串(給頁面與 CSV 用)
    if not s:
        return ""
    try:
        return parse_utc(s).astimezone(LOCAL_TZ).strftime("%Y-%m-%d %H:%M")
    except (ValueError, TypeError):
        return s


def local_date_to_utc_range(start_date, end_date):
    # 使用者輸入的是台灣日期,DB 存 UTC:把 [start 00:00, end 23:59:59] 台灣時間
    # 換算成對應的 UTC 字串區間,日期篩選才不會整體偏移 8 小時
    start_utc = end_utc = None
    try:
        if start_date:
            local_start = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=LOCAL_TZ)
            start_utc = local_start.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        if end_date:
            local_end = (datetime.strptime(end_date, "%Y-%m-%d")
                         .replace(hour=23, minute=59, second=59, tzinfo=LOCAL_TZ))
            end_utc = local_end.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    except (ValueError, TypeError):
        return None, None
    return start_utc, end_utc


def safe_int(value, default=None):
    # 取代「先 isdigit() 再 int()」:isdigit 對上標²等 Unicode 數字為真但 int() 會爆
    try:
        return int(str(value).strip())
    except (ValueError, TypeError):
        return default


def today_local():
    return datetime.now(LOCAL_TZ).date()


# ---------------------------------------------------------------------------
# 存貨規劃:安全庫存推導、XYZ 分級
# 業界系統不讓使用者憑印象填低庫存門檻,而是由用量變異、前置期與目標服務水準推導。
# ---------------------------------------------------------------------------

# 服務水準對應的常態分位數 Z(業界常用檔位)
Z_TABLE = {80: 0.84, 85: 1.04, 90: 1.28, 95: 1.65, 97: 1.88, 98: 2.05, 99: 2.33}

USAGE_WINDOW_DAYS = 90   # 統計用量的回溯天數


def z_for_service_level(level):
    # 取最接近的檔位,避免使用者填了非標準值就算不出來
    try:
        lv = float(level)
    except (TypeError, ValueError):
        lv = 95.0
    return Z_TABLE[min(Z_TABLE, key=lambda k: abs(k - lv))]


def usage_stats(product_id, window_days=USAGE_WINDOW_DAYS):
    """回傳期間內的日用量平均與標準差(只看出庫,那才是真正的消耗)。"""
    since = (datetime.now(timezone.utc) - timedelta(days=window_days)).strftime("%Y-%m-%d %H:%M:%S")
    rows = get_db().execute("""
        SELECT created_at, quantity FROM transactions
        WHERE product_id = ? AND type = 'out' AND created_at >= ?
    """, (product_id, since)).fetchall()
    if not rows:
        return None
    # 先彙總成每日用量,沒有出庫的日子算 0(否則會嚴重高估日均)
    per_day = {}
    for r in rows:
        day = r["created_at"][:10]
        per_day[day] = per_day.get(day, 0) + r["quantity"]
    series = [per_day.get(
        (datetime.now(timezone.utc) - timedelta(days=i)).strftime("%Y-%m-%d"), 0)
        for i in range(window_days)]
    n = len(series)
    mean = sum(series) / n
    var = sum((x - mean) ** 2 for x in series) / n
    return {"mean": mean, "sd": math.sqrt(var), "total": sum(series), "days": n}


def suggest_safety_stock(product, stats):
    """SS = Z × σ_d × √L(僅需求變異版本;前置期變異未知時的標準做法)。
    回傳 (建議安全庫存, 再訂購點) — 皆無條件進位為整數,寧多勿少。"""
    if not stats or stats["mean"] <= 0:
        return None, None
    lead = max(1, product["lead_time_days"] or 7)
    z = z_for_service_level(product["service_level"])
    ss = z * stats["sd"] * math.sqrt(lead)
    rop = stats["mean"] * lead + ss
    return math.ceil(ss), math.ceil(rop)


def xyz_class(stats):
    """依變異係數 CV 分級:X 穩定、Y 中等、Z 高度波動(業界常用門檻 0.5 / 1.0)。"""
    if not stats or stats["mean"] <= 0:
        return "—", None
    cv = stats["sd"] / stats["mean"]
    return ("X" if cv < 0.5 else "Y" if cv < 1.0 else "Z"), cv


# ---------------------------------------------------------------------------
# 預留:可用量 = 現貨 − 有效預留(業界的 on-hand vs available 區分)
# ---------------------------------------------------------------------------

                                                                        # noqa: E302
# 在途:已下訂但尚未入庫的量。只算還活著的訂單(結案與作廢都不算)
ONORDER_STATUSES = ("ordered", "shipped", "arrived")


def onorder_map():
    """回傳 {product_id: 在途數量};在途 = Σ(訂購量 − 已收量),負數視為 0。"""
    rows = get_db().execute(f"""
        SELECT i.product_id AS pid,
               SUM(CASE WHEN i.ordered_qty > i.received_qty
                        THEN i.ordered_qty - i.received_qty ELSE 0 END) AS qty
        FROM purchase_order_items i
        JOIN purchase_orders o ON i.po_id = o.id
        WHERE i.product_id IS NOT NULL
          AND o.status IN ({','.join('?' * len(ONORDER_STATUSES))})
        GROUP BY i.product_id
    """, ONORDER_STATUSES).fetchall()
    return {r["pid"]: r["qty"] or 0 for r in rows}


def po_status_summary():
    """首頁用的採購狀態總覽:每個階段有幾張單、幾個品項、多少在途量。"""
    out = {}
    for st in ONORDER_STATUSES:
        r = get_db().execute("""
            SELECT COUNT(DISTINCT o.id) AS orders,
                   COUNT(i.id) AS items,
                   COALESCE(SUM(CASE WHEN i.ordered_qty > i.received_qty
                                     THEN i.ordered_qty - i.received_qty ELSE 0 END), 0) AS qty
            FROM purchase_orders o LEFT JOIN purchase_order_items i ON i.po_id = o.id
            WHERE o.status = ?
        """, (st,)).fetchone()
        out[st] = {"orders": r["orders"] or 0, "items": r["items"] or 0, "qty": r["qty"] or 0}
    return out


def reserved_map():
    return {r["product_id"]: r["qty"] for r in get_db().execute("""
        SELECT product_id, SUM(quantity) AS qty FROM reservations
        WHERE status = 'active' GROUP BY product_id
    """).fetchall()}


def reserved_qty(product_id):
    row = get_db().execute("""
        SELECT COALESCE(SUM(quantity), 0) AS qty FROM reservations
        WHERE product_id = ? AND status = 'active'
    """, (product_id,)).fetchone()
    return row["qty"]


# ---------------------------------------------------------------------------
# 批次消耗:FIFO / FEFO(出庫與盤虧調整共用,確保批次帳恆等式在兩條路徑都成立)
# ---------------------------------------------------------------------------

def consume_lots(db, pid, qty, tx_id, ts):
    """依商品的出庫策略消耗批次並記錄追溯明細。
    FIFO = 先進先出(依入庫時間);FEFO = 先到期先出(依有效期,無效期者排最後)。"""
    strategy = (db.execute("SELECT issue_strategy FROM products WHERE id = ?",
                           (pid,)).fetchone() or {"issue_strategy": "FIFO"})["issue_strategy"]
    if strategy == "FEFO":
        # 有效期空字串排最後,同效期再依入庫順序
        order = "CASE WHEN expiry_date = '' THEN 1 ELSE 0 END, expiry_date, received_at, id"
    else:
        order = "received_at, id"
    remaining = qty
    lots = db.execute(f"""
        SELECT id, qty_remaining FROM lots
        WHERE product_id = ? AND qty_remaining > 0
        ORDER BY {order}
    """, (pid,)).fetchall()
    for lot in lots:
        if remaining <= 0:
            break
        take = min(lot["qty_remaining"], remaining)
        db.execute("UPDATE lots SET qty_remaining = qty_remaining - ? WHERE id = ?",
                   (take, lot["id"]))
        db.execute("""
            INSERT INTO lot_consumptions (transaction_id, lot_id, quantity)
            VALUES (?, ?, ?)
        """, (tx_id, lot["id"], take))
        remaining -= take
    if remaining > 0:
        # 防禦:批次帳有缺口時建調整批補齊(正常流程不會發生)
        cur = db.execute("""
            INSERT INTO lots (product_id, transaction_id, lot_no, qty_received,
                              qty_remaining, note, received_at)
            VALUES (?, ?, ?, ?, 0, '缺口調整批', ?)
        """, (pid, tx_id, f"ADJ-{tx_id}", remaining, ts))
        db.execute("""
            INSERT INTO lot_consumptions (transaction_id, lot_id, quantity)
            VALUES (?, ?, ?)
        """, (tx_id, cur.lastrowid, remaining))
        # 正常流程不該走到這裡:留下明確告警供人工稽核,不靜默吸收
        audit("批次帳異常", "商品", pid,
              f"消耗 {qty} 時批次剩餘不足 {remaining},已建立缺口調整批 ADJ-{tx_id}")
        print(f"[警告] 商品 {pid} 批次帳出現缺口 {remaining},已記錄稽核軌跡")


def fmt_num(value):
    # 數字(無千分位):整數不帶小數點,小數去尾零(125.0 → 125、25.5 → 25.5)
    # 用於表單值與 CSV(千分位會讓 float() 與 Excel 數值欄位解析失敗)
    return f"{value:.2f}".rstrip("0").rstrip(".")


def fmt_money(value):
    # 金額顯示(含千分位,僅供頁面顯示):12000 → 12,000
    return f"{value:,.2f}".rstrip("0").rstrip(".")


# ---------------------------------------------------------------------------
# 登入保護
# ---------------------------------------------------------------------------

def forbidden(msg="您沒有執行此操作的權限,請洽管理員。"):
    return Response(f"<h1>403 權限不足</h1><p>{msg}</p>"
                    f"<p><a href='/'>返回庫存總覽</a></p>",
                    status=403, mimetype="text/html")


def admin_required(view):
    # 破壞性操作(刪除、整批匯入、帳號管理)限管理員
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        if not session.get("is_admin"):
            return forbidden()
        return view(*args, **kwargs)
    return wrapped


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped


# ---------------------------------------------------------------------------
# 版面模板:共用 LAYOUT + 各頁 body 片段,render_page() 組裝
# ---------------------------------------------------------------------------

LAYOUT = """
<!doctype html>
<html lang="zh-Hant">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{% if page_title %}{{ page_title }} — {% endif %}庫存管理系統</title>
    <style>
        /* ============================================================
           設計代幣(第九階段第 1 批)
           規則一:一色一義。琥珀=等你動手;靛藍=在路上/純資訊;綠=已定案;
                   紫=有貨但不能動;紅=壞消息與破壞性動作。不得兼差。
           規則二:字重只有 400 與 700。Windows 中文落到微軟正黑體,
                   只有這兩個真字重,600/800 是瀏覽器合成的假粗體。
           規則三:所有中性色都由 --hull 深藏青混白而來,不借用框架預設灰。
           ============================================================ */
        :root {
            color-scheme: light dark;
            /* 骨架與中性階 */
            --hull: #0f1b2e; --hull-2: #16243c; --hull-3: #22334f;
            --on-hull: #f2f5fa; --on-hull-2: #aebbce;
            --text: #121b2b; --mute: #495467; --line-str: #7f8d9f;
            --line: #d5dae1; --line-2: #c1c8d2; --wash: #edeff3;
            --paper: #dfe3e9; --card: #ffffff;
            /* 語意色:每個只有一個意思 */
            --amber: #f2a20c; --amber-on: #2a1902; --amber-text: #8a5106; --amber-soft: #fdf0d8;
            --transit: #2e5aac; --transit-text: #24488a; --transit-soft: #e4ecf9;
            --onhand: #12674a; --onhand-text: #0e5239; --onhand-soft: #e0f0e8;
            --held: #5a4aa0;   --held-text: #4a3c8c;   --held-soft: #ebe8f7;
            --fault: #b0202e;  --fault-text: #96131f;  --fault-soft: #fbe6e7;
            /* 過期/呆滯的第二個編碼:色盲、黑白列印、投影失真都分得出來 */
            --hatch: repeating-linear-gradient(45deg, #fbe6e7 0 7px, #f4d2d5 7px 14px);
            /* 每個語意色的邊框階與互動階,避免在規則裡散落一次性 hex */
            --amber-edge: #e9c787; --amber-hi: #ffb020; --amber-line: #d68b04;
            --transit-edge: #b7cbec; --onhand-edge: #a9cebe;
            --held-edge: #c3b9ea; --fault-edge: #e7aeb3;
            --white: #ffffff;
            --search-icon: #8496ae; --search-ph: #8fa0b8; --search-focus: #1b2c48;
            --print-ink: #121b2b; --print-mute: #495467;
            /* 字體:拉丁排前面,數字與英文才走得到有真等寬數字的字體 */
            --font-ui: "Segoe UI", -apple-system, BlinkMacSystemFont, Roboto,
                       "Noto Sans TC", "Microsoft JhengHei", "PingFang TC", sans-serif;
            --font-code: "Cascadia Mono", Consolas, "SF Mono", ui-monospace, Menlo, monospace;
            /* 字級:六級全整數,中文下限 13px */
            --t1: 12px; --t2: 13px; --t3: 15px; --t4: 18px; --t5: 24px; --t6: 34px;
            /* 間距:基數 4,只允許七個值 */
            --s1: 4px; --s2: 8px; --s3: 12px; --s4: 16px; --s5: 24px; --s6: 32px; --s7: 48px;
            --r-ctl: 4px; --r-pane: 10px;
            --sh-1: 0 1px 2px rgba(15,27,46,.06), 0 1px 3px rgba(15,27,46,.05);
            --sh-2: 0 2px 6px rgba(15,27,46,.08), 0 8px 20px rgba(15,27,46,.10);
            --sh-3: 0 10px 24px rgba(15,27,46,.16), 0 20px 48px rgba(15,27,46,.14);
            --ease: cubic-bezier(.2,.7,.3,1);
        }
        /* 深色主題:整套代幣重新定義。舊版只覆寫六條規則,
           造成淺底上出現深色區塊、儲位欄變成 1.48:1 幾乎看不見 */
        @media (prefers-color-scheme: dark) {
            :root {
                --hull: #060b14; --hull-2: #0e1727; --hull-3: #26344b;
                --on-hull: #f2f5fa; --on-hull-2: #a9b6c9;
                --text: #e8ecf3; --mute: #a3afc0; --line-str: #6c7a8d;
                --line: #2a3548; --line-2: #3a4759; --wash: #182234;
                --paper: #0b1220; --card: #131d2e;
                --amber: #f2a20c; --amber-on: #201302; --amber-text: #f0b44a; --amber-soft: #3a2a0c;
                --transit: #5b8ce0; --transit-text: #8fb3ee; --transit-soft: #16243f;
                --onhand: #2fa277; --onhand-text: #63c39c; --onhand-soft: #102a22;
                --held: #9184d8;   --held-text: #a79be6;   --held-soft: #221e38;
                --fault: #e05563;  --fault-text: #f1919a;  --fault-soft: #331419;
                --hatch: repeating-linear-gradient(45deg, #331419 0 7px, #43191f 7px 14px);
                --search-icon: #7f8ea6; --search-ph: #8494ad; --search-focus: #16243c;
                /* 白紙色與列印色不隨主題改變:QR 與貨架標籤在深色主題下仍印在白紙上 */
                --white: #ffffff; --print-ink: #121b2b; --print-mute: #495467;
                --amber-edge: #6b4d10; --amber-hi: #ffb020; --amber-line: #8a6408;
                --transit-edge: #2c4270; --onhand-edge: #1c4a38;
                --held-edge: #3d3468; --fault-edge: #5c2830;
                --sh-1: 0 1px 2px rgba(0,0,0,.40);
                --sh-2: 0 2px 6px rgba(0,0,0,.44), 0 8px 20px rgba(0,0,0,.40);
                --sh-3: 0 10px 24px rgba(0,0,0,.50);
            }
        }
        @media (prefers-reduced-motion: reduce) {
            * { transition: none !important; animation: none !important; }
        }
        * { box-sizing: border-box; }
        body { margin: 0; background: var(--paper); color: var(--text);
               font-size: var(--t3); line-height: 1.75; font-family: var(--font-ui);
               -webkit-font-smoothing: antialiased; }
        .mono { font-family: var(--font-code); font-variant-numeric: tabular-nums; }
        .unit { font-family: var(--font-ui); font-size: var(--t2); color: var(--mute);
                font-weight: 400; margin-left: 2px; }
        .sr-only { position: absolute; width: 1px; height: 1px; overflow: hidden;
                   clip: rect(0 0 0 0); white-space: nowrap; }

        /* ── 頂列:品牌 + 全站搜尋 + 使用者。搜尋在每一頁的同一個位置 ── */
        .topbar { position: sticky; top: 0; z-index: 40; background: var(--hull);
                  display: flex; align-items: center; gap: var(--s4);
                  padding: 0 var(--s4); height: 48px; border-bottom: 1px solid var(--hull-3); }
        .topbar .brand { display: flex; align-items: center; gap: var(--s2); color: #fff;
                         text-decoration: none; font-size: var(--t3); font-weight: 700; white-space: nowrap; }
        .topbar .brand-mark { width: 9px; height: 20px; background: var(--amber);
                              display: block; border-radius: 1px;
                              box-shadow: 5px 0 0 -2px rgba(242,162,12,.45); }
        .gsearch { flex: 1 1 auto; max-width: 520px; display: flex; position: relative; }
        .gsearch svg { position: absolute; left: 11px; top: 50%; transform: translateY(-50%);
                       width: 17px; height: 17px; color: var(--search-icon); pointer-events: none; }
        .topbar .gsearch input[type=search] {
            width: 100%; height: 32px; margin: 0; padding: 0 var(--s3) 0 34px;
            font: inherit; font-size: var(--t2); color: var(--on-hull);
            background: var(--hull-2); border: 1px solid var(--hull-3);
            border-radius: var(--r-ctl); -webkit-appearance: none; }
        .topbar .gsearch input[type=search]::placeholder { color: var(--search-ph); }
        .topbar .gsearch input[type=search]:focus { outline: 3px solid var(--amber);
            outline-offset: 1px; background: var(--search-focus); border-color: var(--amber); }
        .user-info { margin-left: auto; color: var(--on-hull-2); font-size: var(--t2); white-space: nowrap; }
        .user-info a { color: var(--on-hull-2); }

        /* ── 導覽:倉庫語彙分組,零 JS 的 details ── */
        nav { position: sticky; top: 48px; z-index: 35; background: var(--hull-2);
              display: flex; padding: 0 var(--s4); border-bottom: 1px solid var(--hull-3); }
        nav > a.top, nav summary { display: flex; align-items: center; padding: 0 var(--s3);
              height: 34px; color: var(--on-hull-2); text-decoration: none;
              font-size: var(--t2); font-weight: 400; border-bottom: 3px solid transparent;
              cursor: pointer; white-space: nowrap; }
        nav > a.top:hover, nav summary:hover { color: #fff; }
        nav > a.top.active { color: #fff; font-weight: 700; border-bottom-color: var(--amber); }
        nav details.menu { position: relative; }
        nav details.menu > summary { list-style: none; gap: 5px; user-select: none; }
        nav details.menu > summary::-webkit-details-marker { display: none; }
        nav details.menu > summary::after { content: ""; width: 0; height: 0;
              border: 4px solid transparent; border-top-color: currentColor; margin-top: 4px; }
        nav details.menu[open] > summary { color: #fff; background: var(--hull); }
        nav details.menu.here > summary { color: #fff; font-weight: 700; border-bottom-color: var(--amber); }
        nav .menu-panel { position: absolute; top: 100%; left: 0; min-width: 186px;
              background: var(--card); border: 1px solid var(--line);
              border-radius: 0 var(--r-pane) var(--r-pane) var(--r-pane);
              padding: var(--s1); box-shadow: var(--sh-2); z-index: 50; }
        nav .menu-panel a { display: block; color: var(--text); text-decoration: none;
              padding: var(--s2) var(--s3); border-radius: var(--r-ctl);
              font-size: var(--t2); white-space: nowrap; }
        nav .menu-panel a:hover { background: var(--wash); }
        nav .menu-panel a.active { background: var(--amber-soft); color: var(--amber-text); font-weight: 700; }


        /* ── 待辦帶:招牌元素,全站常駐。四格全部是可以「數」的件數 ── */
        .rail { background: var(--hull); display: grid; grid-template-columns: repeat(4, 1fr);
                border-bottom: 1px solid var(--hull-3); }
        .rail a { display: block; padding: 4px var(--s4) 5px; text-decoration: none;
                  color: var(--on-hull); border-left: 1px solid var(--hull-3);
                  position: relative; transition: background .15s var(--ease); }
        .rail a:first-child { border-left: none; }
        .rail a::before { content: ""; position: absolute; left: 0; top: 0; bottom: 0;
                          width: 4px; background: var(--hull-3); }
        .rail a:hover { background: var(--hull-2); }
        .rail .k { display: block; font-size: var(--t1); color: var(--on-hull-2); line-height: 1.35; }
        .rail .v { display: block; font-size: 21px; font-weight: 700;
                   line-height: 1.2; font-family: var(--font-code);
                   font-variant-numeric: tabular-nums; }
        .rail .v .unit { color: var(--on-hull-2); }
        .rail .s { display: block; font-size: 11px; line-height: 1.35; color: var(--on-hull-2);
                   overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
        /* 語意色:琥珀=等你動手、靛藍=純資訊、紅=壞消息、灰=沒事 */
        .rail .act::before { background: var(--amber); }
        .rail .act .v { color: var(--amber); }
        .rail .info::before { background: var(--transit); }
        .rail .info .v { color: #89aeec; }
        .rail .bad::before { background: var(--fault); }
        .rail .bad .v { color: #f2909a; }
        .rail .none::before { background: var(--hull-3); }
        .rail .none .v { color: var(--on-hull-2); font-size: var(--t3);
                         font-weight: 400; font-family: var(--font-ui); line-height: 1.5; }
        /* ── 庫齡長條:純 CSS,不引入任何圖表函式庫 ── */
        .bars { display: flex; flex-direction: column; gap: var(--s2); }
        .bar { display: grid; grid-template-columns: 92px 1fr 76px; align-items: center;
               gap: var(--s3); font-size: var(--t2); }
        .bar .track { display: block; height: 22px; background: var(--wash);
                      border: 1px solid var(--line); border-radius: 3px; overflow: hidden; }
        .bar .fill { display: block; height: 100%; border-radius: 2px; min-width: 3px; }
        /* ── 空狀態:一定要有出路,不能只寫「沒有資料」 ── */
        .empty { text-align: center; padding: var(--s6) var(--s4); }
        .empty .h { font-size: var(--t4); font-weight: 700; margin-bottom: var(--s1); }
        .empty .p { font-size: var(--t2); color: var(--mute); max-width: 460px;
                    margin: 0 auto var(--s4); line-height: 1.8; }
        td.chips { white-space: nowrap; }
        td.chips .chip { margin-right: var(--s1); }
        /* ── 版面 ── */
        .container { max-width: 1240px; margin: var(--s5) auto; padding: 0 var(--s4); }
        /* 純表單頁收窄,避免 420px 的輸入框漂在 1200px 的版面裡 */
        .container.narrow { max-width: 820px; }
        .crumb { font-size: var(--t2); color: var(--mute); margin-bottom: var(--s2); }
        .crumb a { color: var(--transit-text); text-decoration: none; }
        .crumb a:hover { text-decoration: underline; }
        h1 { font-size: var(--t5); font-weight: 700; margin: 0 0 var(--s4);
             line-height: 1.35; color: var(--text); }
        /* 標題與該頁的動作放同一排:同層級的東西不另起一排,省一整條帶 */
        .head { display: flex; align-items: baseline; gap: var(--s4);
                flex-wrap: wrap; margin: 0 0 var(--s4); }
        .head h1 { margin: 0; }
        .head .acts { margin-left: auto; display: flex; flex-wrap: wrap;
                      gap: var(--s3); font-size: var(--t2); }
        .head .acts a { color: var(--transit-text); text-decoration: none; }
        .head .acts a:hover { text-decoration: underline; }
        h2 { font-size: var(--t4); font-weight: 700; margin: var(--s6) 0 var(--s3); line-height: 1.45; }
        .pane { background: var(--card); border: 1px solid var(--line);
                border-radius: var(--r-pane); box-shadow: var(--sh-1);
                margin-bottom: var(--s5); overflow: hidden; }
        .pane-h { display: flex; align-items: center; gap: var(--s3); background: var(--hull);
                  color: var(--on-hull); padding: 10px var(--s4); font-size: var(--t2); font-weight: 700; }
        .pane-h .r { margin-left: auto; font-weight: 400; color: var(--on-hull-2);
                     font-size: var(--t2); display: flex; align-items: center; gap: var(--s3); }
        /* 面板標題列內的控件要壓扁,否則一條標題列會被撐成兩倍高 */
        .pane-h .filters { margin: 0; gap: var(--s2); }
        .pane-h .filters select, .pane-h .filters input[type=submit] {
            height: 26px; min-height: 26px; margin: 0; padding: 0 var(--s2);
            font-size: var(--t2); border-radius: 3px; }
        .pane-h .filters select { background: var(--hull-2); color: var(--on-hull);
                                  border-color: var(--hull-3); max-width: 180px; }
        .pane-h .filters input[type=submit] { padding: 0 var(--s3); }
        .pane-h a { color: var(--amber); text-decoration: none; }
        .pane-b { padding: var(--s4); }
        .pane-b.flush { padding: 0; }

        /* ── 資料表:沒有斑馬紋。狀態靠左緣龍骨 + 文字晶片,
              舊版的斑馬紋權重贏過缺料紅底,實測一半的缺料列失去紅色警示 ── */
        table { border-collapse: separate; border-spacing: 0; width: 100%;
                font-size: var(--t2); line-height: 1.45; background: var(--card); }
        .table-scroll { overflow: auto; border-radius: var(--r-pane);
                        border: 1px solid var(--line); box-shadow: var(--sh-1); margin-top: var(--s4); }
        .table-scroll > table { margin: 0; }
        .pane .table-scroll { border: none; border-radius: 0; box-shadow: none; margin: 0; }
        th { position: sticky; top: 0; z-index: 2; background: var(--hull); color: var(--on-hull);
             font-weight: 700; font-size: var(--t1); text-align: left;
             padding: 9px var(--s3); white-space: nowrap; border-bottom: 1px solid var(--hull-3); }
        td { padding: 9px var(--s3); border-bottom: 1px solid var(--line);
             vertical-align: middle; word-break: keep-all; overflow-wrap: anywhere; }
        tr:last-child td { border-bottom: none; }
        tbody tr:hover td { background: var(--wash); }
        th.num, td.num { text-align: right; font-variant-numeric: tabular-nums; }
        td a.plain, td a { color: var(--transit-text); text-decoration: none; }
        td a:hover { text-decoration: underline; }
        /* 左緣龍骨:狀態一律看得到,不會被任何底色蓋掉 */
        td.keel { box-shadow: inset 4px 0 0 var(--line-2); }
        tr.low-stock td:first-child { box-shadow: inset 4px 0 0 var(--fault); }
        tr.low-stock td { background: var(--fault-soft); }
        tr.lot-empty td { color: var(--mute); }
        tr.has-diff td { background: var(--amber-soft); }
        /* 剛儲存的那一列高亮:舊版一列一送出後跳回最頂端,
           現場在第 40 列存完要自己捲回去確認存到沒有 */
        tr:target td { background: var(--transit-soft); box-shadow: inset 0 2px 0 var(--transit),
                       inset 0 -2px 0 var(--transit); }
        tr:target td:first-child { box-shadow: inset 4px 0 0 var(--transit),
                       inset 0 2px 0 var(--transit), inset 0 -2px 0 var(--transit); }
        td[id^="qty-"] { font-family: var(--font-code); font-variant-numeric: tabular-nums;
                         font-weight: 700; font-size: var(--t3); color: var(--text); }
        td[data-label="SKU"], td[data-label="料號"], td[data-label="儲位"],
        td[data-label="批號"], td[data-label="單位"], .loc-cell {
            font-family: var(--font-code); white-space: nowrap; }
        .loc-cell { font-size: var(--t2); color: var(--mute); }
        td[data-label="名稱"] { min-width: 190px; }
        /* 短欄位不換行,列高才穩定;只有品名該吃掉剩餘寬度。
           舊版分類欄會把「AC Adapter」斷成兩行,列高從 40px 變 90px */
        td[data-label="分類"], td[data-label="供應商"], td[data-label="別名料號"],
        td[data-label="狀態"], td[data-label="時間(台灣)"], td[data-label="批號"] {
            white-space: nowrap; text-overflow: ellipsis; overflow: hidden; max-width: 160px; }
        /* 操作欄:橫排不斷字,並釘在表格右緣——橫捲時最常用的那欄不該被推出畫面 */
        td[data-label="操作"], td[data-label="單列"], th:last-child { white-space: nowrap; }
        td[data-label="單列"] .small-btn { white-space: nowrap; }
        td[data-label="操作"] a.plain { margin-right: var(--s2); font-weight: 700; }
        .table-scroll td[data-label="操作"], .table-scroll th:last-child {
            position: sticky; right: 0; z-index: 1;
            box-shadow: -8px 0 10px -8px rgba(15,27,46,.18); }
        .table-scroll td[data-label="操作"] { background: var(--card); }
        .table-scroll tr.low-stock td[data-label="操作"] { background: var(--fault-soft); }
        .table-scroll th:last-child { background: var(--hull); z-index: 3; }
        tfoot td { background: var(--wash); font-weight: 700; border-top: 2px solid var(--hull); }

        /* ── 狀態晶片:方角(工業儀表語彙),永遠帶文字 ── */
        .chip, .badge-low, .chip-open, .chip-posted, .chip-void, .chip-po, .chip-link {
            display: inline-block; font-size: var(--t1); font-weight: 700;
            padding: 2px 7px; border-radius: var(--r-ctl); white-space: nowrap;
            border: 1px solid transparent; }
        .chip.act, .chip-open, .chip-arrived { background: var(--amber-soft);
            color: var(--amber-text); border-color: var(--amber-edge); }
        .chip.info, .chip-ordered, .chip-shipped { background: var(--transit-soft);
            color: var(--transit-text); border-color: var(--transit-edge); }
        .chip.ok, .chip-posted, .chip-closed { background: var(--onhand-soft);
            color: var(--onhand-text); border-color: var(--onhand-edge); }
        .chip.held { background: var(--held-soft); color: var(--held-text); border-color: var(--held-edge); }
        .chip.bad, .badge-low { background: var(--fault-soft); color: var(--fault-text); border-color: var(--fault-edge); }
        .chip.expired { background: var(--hatch); color: var(--fault-text); border-color: var(--fault-edge); }
        .chip.off, .chip-void, .chip-cancelled { background: var(--wash);
            color: var(--mute); border-color: var(--line); }
        .badge-low { margin-left: var(--s1); }
        .chip-link { border-color: var(--line); background: var(--card); color: var(--text);
                     text-decoration: none; padding: 5px var(--s3); font-size: var(--t2); }
        .chip-link:hover { border-color: var(--amber); background: var(--amber-soft); }
        .chip-link.on { background: var(--hull); color: #fff; border-color: var(--hull); }
        .resv-note { font-size: var(--t1); color: var(--held-text); font-weight: 700; white-space: nowrap; }
        .onorder-cell { color: var(--transit-text); font-weight: 700; }
        .alias-cell { font-size: var(--t1); color: var(--mute); }

        /* ── 訊息 ── */
        .msg { padding: 11px var(--s4); border-radius: var(--r-ctl); margin-bottom: var(--s4);
               font-size: var(--t2); border: 1px solid transparent; border-left-width: 4px; }
        .msg.error { background: var(--fault-soft); color: var(--fault-text); border-color: var(--fault); }
        .msg.ok { background: var(--onhand-soft); color: var(--onhand-text); border-color: var(--onhand); }
        .banner { background: var(--amber-soft); color: var(--amber-text);
                  border: 1px solid var(--amber-edge); border-left: 4px solid var(--amber);
                  padding: 11px var(--s4); border-radius: var(--r-ctl);
                  margin-bottom: var(--s4); font-weight: 700; font-size: var(--t2); }
        .banner a { color: var(--amber-text); }
        .note, .sub-links { color: var(--mute); font-size: var(--t2); }
        .pager { margin-top: var(--s4); font-size: var(--t2); }
        .pager .page-no { color: var(--mute); }
        a.plain { color: var(--transit-text); text-decoration: none; }
        a.plain:hover { text-decoration: underline; }

        /* ── 按鈕:實心琥珀 = 這一頁唯一的主要動作 ── */
        input[type=submit], button, .btn {
            display: inline-flex; align-items: center; justify-content: center; gap: var(--s2);
            min-height: 40px; padding: 0 var(--s5); margin-top: var(--s4);
            font: inherit; font-size: var(--t3); font-weight: 700;
            color: var(--amber-on); background: var(--amber); border: 1px solid var(--amber-line);
            border-radius: var(--r-ctl); cursor: pointer; text-decoration: none;
            box-shadow: 0 1px 0 rgba(0,0,0,.10), inset 0 1px 0 rgba(255,255,255,.35);
            transition: background .14s var(--ease), box-shadow .14s var(--ease),
                        transform .08s var(--ease); }
        input[type=submit]:hover, button:hover, .btn:hover { background: var(--amber-hi); }
        input[type=submit]:active, button:active, .btn:active {
            transform: translateY(1px);
            box-shadow: inset 0 2px 3px rgba(90,52,0,.28); }
        /* 焦點框:深藏青對琥珀 8.6:1、對白 16.6:1,遠高於 3:1 門檻(舊版只有 1.44:1) */
        input[type=submit]:focus-visible, button:focus-visible, .btn:focus-visible,
        a:focus-visible, input:focus-visible, select:focus-visible, summary:focus-visible {
            outline: 3px solid var(--hull); outline-offset: 2px; }
        .topbar :focus-visible, nav :focus-visible { outline-color: var(--amber); }
        .btn.ghost { background: var(--card); color: var(--text); border-color: var(--line-str);
                     box-shadow: none; font-weight: 400; }
        .btn.ghost:hover { background: var(--wash); border-color: var(--hull-3); }
        .btn.sm { min-height: 30px; padding: 0 var(--s3); font-size: var(--t2); margin-top: 0; }
        .btnrow { display: flex; gap: var(--s2); flex-wrap: wrap; align-items: center; margin-top: var(--s4); }
        .btnrow .btn, .btnrow input[type=submit], .btnrow button { margin-top: 0; }
        /* 次要動作:表格內的小按鈕。預設中性,危險動作才變紅 */
        .small-btn { display: inline-flex; align-items: center; justify-content: center;
                     min-height: 30px; padding: 0 11px; margin: 0;
                     font: inherit; font-size: var(--t2); font-weight: 700; cursor: pointer;
                     color: var(--fault-text); background: var(--card);
                     border: 1px solid var(--fault-edge); border-radius: var(--r-ctl); box-shadow: none;
                     transition: background .15s var(--ease), border-color .15s var(--ease); }
        .small-btn:hover { background: var(--fault-soft); border-color: var(--fault); }
        .small-btn:active { transform: none; background: var(--fault-soft); }
        .small-btn.ok-btn { color: var(--transit-text); border-color: var(--transit-edge); }
        .small-btn.ok-btn:hover { background: var(--transit-soft); border-color: var(--transit); }
        .small-btn.icon-btn { width: 32px; height: 32px; min-height: 32px; padding: 0; }
        .small-btn.icon-btn svg { width: 16px; height: 16px; display: block; }

        /* ── 表單:欄寬暗示輸入長度 ── */
        label { display: block; margin-top: var(--s4); font-weight: 700;
                font-size: var(--t2); color: var(--text); }
        label .opt { color: var(--mute); font-weight: 400; }
        label .req { color: var(--fault-text); font-weight: 700; }
        .field-help { font-size: var(--t2); color: var(--mute); margin: 2px 0 0; font-weight: 400; }
        input[type=text], input[type=password], input[type=number], input[type=date],
        input[type=search], select {
            width: 100%; max-width: 460px; height: 42px; padding: 0 var(--s3);
            margin-top: var(--s1); font: inherit; font-size: 16px; color: var(--text);
            background: var(--card); border: 1px solid var(--line-str); border-radius: var(--r-ctl); }
        input.w-qty { max-width: 200px; text-align: right; font-family: var(--font-code); }
        input.w-code { max-width: 320px; font-family: var(--font-code); }
        /* 焦點框不是語意色:它的職責是對所在底色最大對比。淺色面用藏青(對白 16.6:1),
           深色面(頂列)才用琥珀。舊版焦點框只有 1.44:1 */
        input:focus, select:focus { outline: 3px solid var(--hull); outline-offset: 0;
                                     border-color: var(--hull); }
        input[type=file] { margin-top: var(--s2); font-size: var(--t2); }
        .filters { display: flex; flex-wrap: wrap; gap: var(--s2); align-items: center; margin: var(--s2) 0; }
        .filters input, .filters select { width: auto; margin-top: 0; height: 38px; }
        .filters input[type=submit] { margin-top: 0; min-height: 38px; padding: 0 var(--s4); }
        label.chk { display: inline-flex; align-items: center; gap: var(--s1); margin-top: 0;
                    font-size: var(--t2); font-weight: 400; color: var(--mute); white-space: nowrap; }
        label.chk input { width: auto; height: auto; margin: 0; }
        .count-form { display: flex; gap: var(--s2); align-items: center; flex-wrap: wrap; }
        .count-input { width: 104px !important; max-width: 104px !important; margin-top: 0 !important;
                       height: 38px; text-align: right; font-family: var(--font-code); }
        .count-note { width: 150px !important; max-width: 150px !important; margin-top: 0 !important;
                      height: 38px; font-size: var(--t2); }

        /* ── 選料元件(第 2 批):取代 2,281 項的下拉選單 ── */
        .picklist { display: flex; flex-direction: column; gap: var(--s2); }
        .pickrow { display: flex; align-items: center; gap: var(--s3); padding: 10px var(--s3);
                   background: var(--card); border: 1px solid var(--line);
                   border-radius: var(--r-ctl); text-decoration: none; color: var(--text);
                   min-height: 52px; }
        .pickrow:hover { border-color: var(--amber); background: var(--amber-soft); }
        .pickrow .sku { font-family: var(--font-code); font-size: var(--t2);
                        color: var(--mute); white-space: nowrap; min-width: 132px; }
        .pickrow .nm { font-weight: 700; font-size: var(--t3); flex: 1 1 auto; min-width: 0; }
        .pickrow .loc { font-family: var(--font-code); font-size: var(--t2); background: var(--wash);
                        padding: 2px var(--s2); border-radius: var(--r-ctl); white-space: nowrap; }
        .pickrow .pq { font-family: var(--font-code); font-weight: 700; font-size: var(--t3);
                       white-space: nowrap; min-width: 96px; text-align: right; }
        /* 料件確認卡:第二步永遠知道自己在登記哪一項 */
        .pickcard { display: flex; gap: var(--s4); align-items: center; flex-wrap: wrap;
                    background: var(--hull); color: var(--on-hull);
                    border-radius: var(--r-pane); padding: var(--s4); margin-bottom: var(--s4); }
        .pickcard .nm { font-size: var(--t4); font-weight: 700; line-height: 1.4; }
        .pickcard .meta { font-size: var(--t2); color: var(--on-hull-2); font-family: var(--font-code); }
        .pickcard .right { margin-left: auto; text-align: right; }
        .pickcard .big { font-family: var(--font-code); font-size: var(--t6);
                         font-weight: 700; line-height: 1.15; }
        .pickcard a { color: var(--amber); }

        /* ── KPI 區塊 ── */
        .stat-row { display: flex; flex-wrap: wrap; gap: var(--s3); margin: var(--s4) 0 var(--s5); }
        /* 預設中性。舊版依 nth-child 輪流換色——顏色由「排第幾個」決定而不是由意思決定,
           等於顏色不傳遞任何資訊。要語意色請由樣板指定 .stat-box.good/.warn/.bad/.info */
        .stat-box { flex: 1 1 150px; background: var(--card); border: 1px solid var(--line);
                    border-top: 4px solid var(--line-2); border-radius: var(--r-pane);
                    padding: var(--s3) var(--s4); box-shadow: var(--sh-1); }
        .stat-box.good { border-top-color: var(--onhand); }
        .stat-box.info { border-top-color: var(--transit); }
        .stat-box.warn { border-top-color: var(--amber); }
        .stat-box.bad  { border-top-color: var(--fault); }
        .stat-box.held { border-top-color: var(--held); }
        .stat-num { font-size: var(--t6); font-weight: 700; font-family: var(--font-code);
                    font-variant-numeric: tabular-nums; line-height: 1.15; color: var(--text); }
        .stat-cap { font-size: var(--t2); color: var(--mute); }

        /* ── 採購狀態軌與入口卡 ── */
        .po-flow { display: grid; grid-template-columns: repeat(auto-fit, minmax(190px, 1fr));
                   gap: var(--s3); margin: var(--s4) 0 var(--s2); }
        .po-step { display: flex; flex-direction: column; gap: 2px; text-decoration: none;
                   background: var(--card); border: 1px solid var(--line);
                   border-left: 4px solid var(--transit); border-radius: var(--r-pane);
                   padding: var(--s3) var(--s4); color: var(--text); box-shadow: var(--sh-1);
                   transition: box-shadow .16s var(--ease), transform .16s var(--ease); }
        .po-step:hover { transform: translateY(-1px); box-shadow: var(--sh-2); }
        .po-step.po-shipped { border-left-color: var(--transit-text); }
        .po-step.po-arrived { border-left-color: var(--amber); }
        .po-step-label { font-size: var(--t2); font-weight: 700; color: var(--mute); }
        .po-step-n { font-size: 28px; font-weight: 700; font-family: var(--font-code);
                     font-variant-numeric: tabular-nums; line-height: 1.2; }
        .po-step-n small { font-size: var(--t2); font-weight: 400; color: var(--mute); margin-left: 2px; }
        .po-step-sub { font-size: var(--t2); color: var(--mute); }
        .po-home { margin-top: var(--s5); }
        .po-home .po-flow { margin: 0; }
        .po-home .po-step { flex-direction: row; align-items: baseline; gap: var(--s2);
                            padding: var(--s2) var(--s4); }
        .po-home .po-step-label { color: var(--text); }
        .po-home .po-step-n { font-size: var(--t5); }
        .po-home .po-step-sub { margin-left: auto; text-align: right; }
        .po-track { display: grid; grid-template-columns: repeat(4, 1fr); gap: 2px; margin: var(--s4) 0; }
        .po-node { padding: var(--s3); background: var(--card); border: 1px solid var(--line);
                   text-align: center; }
        .po-node:first-child { border-radius: var(--r-pane) 0 0 var(--r-pane); }
        .po-node:last-child { border-radius: 0 var(--r-pane) var(--r-pane) 0; }
        .po-node-label { font-size: var(--t2); font-weight: 700; color: var(--mute); }
        .po-node.done { background: var(--hull); color: #fff; border-color: var(--hull); }
        .po-node.done .po-node-label { color: #fff; }
        .po-node.now { background: var(--amber); border-color: var(--amber-line); }
        .po-node.now .po-node-label { color: var(--amber-on); }
        .po-dot { display: none; }

        /* ── 其他 ── */
        .detail-section { margin-top: var(--s6); border-top: 1px solid var(--line); padding-top: var(--s4); }
        .import-help { background: var(--wash); border: 1px solid var(--line);
                       padding: var(--s3) var(--s4); border-radius: var(--r-pane);
                       font-size: var(--t2); color: var(--text); }
        .import-help code { background: var(--card); padding: 1px 6px; border-radius: 3px;
                            font-family: var(--font-code); }
        .hero-search { display: flex; flex-wrap: wrap; gap: var(--s2); margin: var(--s1) 0; }
        .hero-search .search-field { position: relative; flex: 1 1 260px; display: flex; }
        .hero-search .search-field svg { position: absolute; left: 14px; top: 50%;
             transform: translateY(-50%); width: 18px; height: 18px; color: var(--line-str); }
        .hero-search input[type=text] { flex: 1 1 auto; max-width: none; margin-top: 0;
             padding-left: 42px; }
        .hero-search select { width: auto; margin-top: 0; }
        .hero-search input[type=submit] { margin-top: 0; }
        .photo-wall { display: flex; flex-wrap: wrap; gap: var(--s3); margin: var(--s3) 0; }
        .photo-wall .photo-item { text-align: center; }
        .photo-wall img, img.thumb { max-width: 150px; max-height: 150px;
             border: 1px solid var(--line); border-radius: var(--r-pane); display: block; }
        .photo-wall .photo-item form { margin-top: var(--s1); }
        .qr-row { display: flex; gap: var(--s4); align-items: center; flex-wrap: wrap; margin: var(--s4) 0; }
        .qr-img { width: 120px; height: 120px; border: 1px solid var(--line);
                  border-radius: var(--r-pane); background: var(--white); }
        .qr-cap { font-weight: 700; font-size: var(--t2); }
        .label-sheet { display: grid; grid-template-columns: repeat(auto-fill, minmax(230px, 1fr));
                       gap: var(--s2); }
        .label { display: flex; gap: var(--s3); align-items: center; border: 1px solid var(--line-str);
                 border-radius: var(--r-ctl); padding: var(--s3); background: var(--white);
                 break-inside: avoid; color: var(--print-ink); }
        .label img { width: 84px; height: 84px; flex-shrink: 0; }
        .label-text { min-width: 0; }
        .label-sku { font-family: var(--font-code); font-weight: 700; font-size: var(--t2); }
        .label-name { font-size: var(--t2); margin: 1px 0; word-break: break-word; }
        .label-loc { font-size: var(--t1); color: var(--print-mute); }
        .auth-box { max-width: 380px; margin: 0 auto; }
        .container:has(.auth-box) { max-width: 440px; margin-top: 9vh; }
        form.inline { display: inline; }
        /* 頁尾:版本號存在的唯一目的就是讓人辨識版本,舊版對比只有 2.18:1 */
        footer { text-align: center; color: var(--mute); padding: var(--s5) var(--s4);
                 font-size: var(--t2); }
        footer .ver { font-family: var(--font-code); color: var(--text); }

        /* 商品詳細頁的主要動作。第 1 批清掉首頁動作磚時把這一區的樣式一併移除了,
           .action-links / .primary 之後一直沒有樣式,實測手機上只有 17px 高、
           還用 emoji 當圖示——掃完料架 QR 之後最常做的兩件事就長這樣。 */
        .action-links { display: flex; flex-wrap: wrap; gap: var(--s2); margin: var(--s4) 0; }
        .action-links a { display: inline-flex; align-items: center; gap: var(--s2);
                          min-height: 44px; padding: 0 var(--s4); border-radius: var(--r-ctl);
                          border: 1px solid var(--line-str); background: var(--card);
                          color: var(--text); text-decoration: none; font-size: var(--t3);
                          font-weight: 700; box-shadow: 0 1px 0 var(--shadow-1); }
        .action-links a svg { width: 18px; height: 18px; flex-shrink: 0; }
        .action-links a:hover { border-color: var(--hull); }
        .action-links a:active { transform: translateY(1px); box-shadow: none; }
        .action-links a.primary { background: var(--hull); border-color: var(--hull); color: #fff; }
        .action-links a.primary:hover { background: var(--hull-2); }
        .action-links a.primary.out { background: var(--amber); border-color: var(--amber);
                                      color: var(--amber-on); }
        .action-links a.primary.out:hover { filter: brightness(1.06); }

        /* 手機常駐工具列。桌面的 nav 是 sticky,760px 以下卻被覆寫成 static——
           實測 /history 手機版整頁 54,999px,捲到 2,000px 時 nav 已經在 −1,952px,
           要換頁得先捲回最頂端。最需要常駐導覽的裝置反而沒有。
           桌面完全不出現,列印也不出現。 */
        .tabbar { display: none; }
        .tabbar .who { display: block; padding: var(--s3) var(--s4); color: var(--on-hull-2);
                       font-size: var(--t2); border-bottom: 1px solid var(--hull-3); }

        @media print {
            /* 印出來的是「單據」不是「畫面」。舊版這裡有一條
               form { display: none !important; },而盤點單與收貨單的整張明細表格
               就包在 form 裡——實測 /counts/<id> 螢幕上 50 列輸入格,印成 A4 只剩
               1 頁 290 個字,一列料都沒有。倉管拿著那張紙走進倉庫是盤不了點的。
               所以這裡改成:拿掉「操作用」的元件,留下「單據本身」。 */
            .topbar, nav, footer, .rail, .tabbar, .no-print, .filters,
            .savebar, .btnrow, .pager, .crumb, .msg, .action-links, .pane-h .r,
            input[type=submit], input[type=file], button, .btn, .small-btn { display: none !important; }
            form { display: block !important; }
            .container { max-width: none; margin: 0; padding: 0; }
            .pane { box-shadow: none; border: none; }
            .table-scroll { overflow: visible !important; }
            /* 從手機按列印時視窗寬度 < 760px,卡片式表格會跟著套用,
               印出來是一疊卡片而不是一張表。列印一律還原成表格。 */
            table, table.cards { display: table !important; width: 100% !important; }
            table.cards tbody { display: table-row-group !important; }
            table.cards tr, table.cards tr:first-child {
                display: table-row !important; border: none !important;
                margin: 0 !important; padding: 0 !important; }
            table.cards td { display: table-cell !important; text-align: left;
                             border: none; border-bottom: 1px solid #999 !important; }
            table.cards td::before { content: none !important; }
            /* 列印永遠白底黑字:深色模式的色票會讓整張單印成黑底,吃掉整支碳粉匣 */
            body, .pane, .pane-b, .card, table, tr, td, th, .stat-box {
                background: #fff !important; color: #000 !important;
                box-shadow: none !important; }
            .pane-h { background: #fff !important; color: #000 !important;
                      border-bottom: 2px solid #000; }
            th { background: #fff !important; color: #000 !important;
                 border-bottom: 2px solid #000; }
            a { color: #000 !important; text-decoration: none !important; }
            /* 表頭每頁重印、資料列與標籤不被分頁切成兩半 */
            thead { display: table-header-group; }
            tr, .label, .stat-box { page-break-inside: avoid; }
            /* 實盤數/實收數要留成可以用原子筆填的空格,不是藏起來的輸入框 */
            input[type=number], input[type=text], input[type=date], select {
                display: inline-block !important; border: 1px solid #000 !important;
                background: #fff !important; color: #000 !important; min-width: 56px;
                height: 26px !important; min-height: 0 !important; padding: 0 4px !important;
                -webkit-appearance: none; appearance: none; box-shadow: none !important; }
            input::placeholder { color: transparent !important; }
            /* 紙張很貴:一張 120 列的盤點單原本要 8 頁,收緊列高後省一半 */
            td, th { padding: 3px 6px !important; font-size: 11px !important; line-height: 1.3; }
            h1 { font-size: 17px; margin: 0 0 6px; }
            .stat-row { gap: 6px; }
            .stat-num { font-size: 18px !important; }
            .label { border-color: #999; }
            @page { margin: 12mm 10mm; }
        }

        @media (max-width: 760px) {
            .container { margin: var(--s3) auto; padding: 0 var(--s3); }
            .topbar { height: 48px; padding: 0 var(--s3); gap: var(--s2); }
            .topbar .brand span { display: none; }
            nav { position: static; overflow-x: auto; padding: 0 var(--s3);
                  scrollbar-width: none; }
            nav details.menu { position: static; }
            nav details.menu[open] { flex-basis: 100%; }
            nav .menu-panel { position: absolute; left: var(--s3); right: var(--s3); }
            h1 { font-size: 20px; }
            /* 一般表格:橫向捲動 */
            table { display: block; overflow-x: auto; }
            table:not(.cards) tbody { display: table; width: 100%; }
            /* 卡片式表格:名稱當標題、數量緊隨 */
            table.cards { display: block; overflow: visible; }
            table.cards tbody { display: block; width: 100%; }
            table.cards tr:first-child { display: none; }
            table.cards tr { display: flex; flex-direction: column;
                             border: 1px solid var(--line); border-left: 4px solid var(--line-2);
                             border-radius: var(--r-ctl); margin-bottom: var(--s2);
                             padding: var(--s2) var(--s3); background: var(--card); }
            table.cards tr.low-stock { border-left-color: var(--fault); background: var(--fault-soft); }
            table.cards tr.low-stock td { background: transparent; }
            /* 桌面版靠 td:first-child 的內陰影畫出左邊那條紅線,但卡片版把「名稱」
               和「現貨」用 order 提到前面,DOM 的第一格(料號)會落在卡片中間,
               紅線就變成一段浮在字上的紅槓。卡片本身已經有紅色左框了。 */
            table.cards tr.low-stock td:first-child { box-shadow: none; }
            table.cards td { display: flex; justify-content: space-between; align-items: center;
                             gap: var(--s3); border: none; box-shadow: none; padding: 6px 0;
                             text-align: right; border-bottom: 1px dashed var(--line); }
            table.cards td:last-child { border-bottom: none; }
            table.cards td::before { content: attr(data-label); font-weight: 700;
                                     color: var(--mute); font-size: var(--t2);
                                     text-align: left; flex-shrink: 0; }
            table.cards td[data-label="名稱"] { order: -2; font-size: 17px; font-weight: 700;
                                                justify-content: flex-start; text-align: left; }
            table.cards td[data-label="名稱"]::before { content: none; }
            table.cards td[data-label="庫存"], table.cards td[data-label="現貨"],
            table.cards td[data-label="目前庫存"] { order: -1; }
            table.cards td[data-label="操作"] a.plain { display: inline-block; padding: var(--s2) var(--s3); }
            .small-btn { min-height: 44px; padding: 0 var(--s3); }
            .small-btn.icon-btn { width: 44px; height: 44px; min-height: 44px; }
            input[type=submit], button, .btn { min-height: 44px; }
            .filters input[type=submit] { min-height: 38px; }
            .hero-search .search-field { flex-basis: 100%; }
            .hero-search select, .hero-search input[type=submit] { width: 100%; }
            form:not(.filters):not(.hero-search):not(.inline) > input[type=submit] { width: 100%; }
            table.cards td[data-label="實盤數"], table.cards td[data-label="實收數"],
            table.cards td[data-label="對應商品"] { flex-direction: column; align-items: stretch;
                                                     justify-content: flex-start; gap: var(--s1); }
            .count-form { width: 100%; }
            .count-input, .count-note { width: 100% !important; max-width: none !important; flex: 1 1 100%; }
            .count-form .small-btn { width: 100%; margin-top: var(--s1); }
            .rail { grid-template-columns: 1fr 1fr; }
            .rail a { padding: 8px var(--s3) 9px; border-top: 1px solid var(--hull-3); }
            .rail a:nth-child(-n+2) { border-top: none; }
            .rail a:nth-child(odd) { border-left: none; }
            .rail .v { font-size: 23px; }
            .rail .s { font-size: 11px; line-height: 1.4; }
            .bar { grid-template-columns: 68px 1fr 62px; gap: var(--s2); font-size: var(--t1); }
            .pickrow { flex-wrap: wrap; min-height: 60px; }
            .pickrow .sku { min-width: 0; flex-basis: 100%; }
            .pickcard .right { margin-left: 0; text-align: left; }

            /* ---- 手機常駐工具列 ---- */
            /* 分頁列在手機上交給底部工具列,省下 35px 又永遠構得到。
               但功能一項都不能因此消失,所以「更多」面板要涵蓋原本 nav 的全部項目。 */
            nav { display: none; }
            .user-info { display: none; }
            /* top: auto 是必要的:.tabbar 本身是 <nav>,會吃到基底的
               nav { position: sticky; top: 48px },fixed 元素同時有 top 與 bottom
               就會被拉開成整個畫面高(實測 796px)。 */
            .tabbar { display: grid; grid-template-columns: repeat(5, 1fr);
                      position: fixed; left: 0; right: 0; bottom: 0; top: auto; z-index: 60;
                      padding: 0; border-bottom: none;
                      background: var(--hull); border-top: 1px solid var(--hull-3);
                      box-shadow: 0 -2px 10px rgba(6, 14, 28, .28);
                      padding-bottom: env(safe-area-inset-bottom, 0px); }
            .tabbar > a, .tabbar > details > summary {
                      display: flex; flex-direction: column; align-items: center;
                      justify-content: center; gap: 3px; min-height: 54px; padding: 4px 2px;
                      color: var(--on-hull-2); text-decoration: none; font-size: 11px;
                      line-height: 1.2; list-style: none; cursor: pointer;
                      border-top: 3px solid transparent; }
            .tabbar > details > summary::-webkit-details-marker { display: none; }
            .tabbar svg { width: 21px; height: 21px; stroke: currentColor; fill: none;
                          stroke-width: 1.9; stroke-linecap: round; stroke-linejoin: round; }
            .tabbar > a.active { color: var(--amber); border-top-color: var(--amber); }
            .tabbar > a:active, .tabbar > details > summary:active { background: var(--hull-2); }
            .tabbar a:focus-visible, .tabbar summary:focus-visible {
                      outline: 3px solid var(--amber); outline-offset: -3px; }
            .tabbar details[open] > summary { color: var(--amber); background: var(--hull-2); }
            /* 面板往上開。fixed 定位才不會被工具列自己的 overflow 裁掉,
               高度上限留給頁面內容,項目多時面板自己捲。 */
            .tabbar .more-panel { position: fixed; left: 0; right: 0;
                      bottom: calc(54px + env(safe-area-inset-bottom, 0px));
                      max-height: 68vh; overflow-y: auto; background: var(--hull-2);
                      border-top: 1px solid var(--hull-3);
                      box-shadow: 0 -6px 18px rgba(6, 14, 28, .4); }
            .tabbar .more-panel .grp { padding: var(--s3) var(--s4) 2px; color: var(--on-hull-2);
                      font-size: 11px; letter-spacing: .08em; }
            .tabbar .more-panel a { display: flex; align-items: center; min-height: 48px;
                      padding: 0 var(--s5); color: #fff; text-decoration: none;
                      font-size: var(--t3); border-top: 1px solid var(--hull-3); }
            .tabbar .more-panel a.active { color: var(--amber); }
            .tabbar .more-panel a:active { background: var(--hull-3); }
            /* 工具列是浮在內容上的,不留這段內距就會蓋住頁尾與最後一列資料 */
            body { padding-bottom: calc(62px + env(safe-area-inset-bottom, 0px)); }

            /* ---- 觸控目標 ---- */
            /* 目標是 44px(WCAG 2.5.5 / Apple HIG)。實測修改前:盤點與收貨的
               數量輸入格只有 20px、商品頁的入庫出庫連結 17px、頂列登出 15px。
               現場戴著工作手套,20px 的格子按不進去。 */
            input:not([type=hidden]):not([type=checkbox]):not([type=radio]),
            select, textarea, button, .btn, .small-btn, summary {
                      min-height: 44px !important; }
            .count-input, .count-note { height: 44px !important; }
            .topbar .gsearch input[type=search] { height: 44px; }
            /* 純文字連結沒有內距,天生就只有一行字高。表格與面板標題列裡的
               操作連結要撐開成可以按的區塊。 */
            td a.plain, .action-links a, .pane-h .r a, .crumb a,
            .pager a, .filters label, .filters a, .detail-section a.plain, .import-help a.plain,
            .note a.plain, .head .acts a, .pane-b a.plain {
                      display: inline-flex; align-items: center; min-height: 44px; }
            /* 面板標題列在手機上要疊起來。橫排時「明細」兩個字被四個操作連結
               擠成一直行(明/細 各佔一列),看起來像壞掉。 */
            .pane-h { flex-wrap: wrap; align-items: flex-start;
                      padding: 6px var(--s3); gap: 0; }
            .pane-h .r { flex-basis: 100%; margin-left: 0; display: flex; flex-wrap: wrap;
                         gap: 0 var(--s3); align-items: center; }
            .head .acts { margin-left: 0; }
            .pager a { padding: 0 var(--s3); }
            /* 現場是站在貨架前看這一頁的,四格統計不該把清單推到一個半螢幕以外 */
            .stat-row { gap: var(--s2); margin: var(--s3) 0 var(--s3); }
            .stat-box { flex: 1 1 44%; padding: var(--s2) var(--s3); }
            .stat-num { font-size: 22px; }
            .stat-cap { font-size: var(--t1); }
            /* 相鄰目標留 8px 以上,免得誤觸隔壁 */
            .filters { gap: var(--s3); }
            .action-links { gap: var(--s2); }
            .action-links a { flex: 1 1 46%; justify-content: center; }
            /* 篩選列原本刻意壓成 38px 求密度,但那是桌面的權衡;
               手機上按不到就不是密度問題,是功能問題。 */
            .filters input[type=submit] { min-height: 44px !important; }
        }

        /* 橫拿手機(高度矮)時工具列縮一點,免得吃掉整個畫面 */
        @media (max-width: 760px) and (max-height: 460px) {
            .tabbar > a, .tabbar > details > summary { min-height: 44px; font-size: 10px; }
            .tabbar svg { width: 17px; height: 17px; }
            .tabbar .more-panel { bottom: calc(44px + env(safe-area-inset-bottom, 0px));
                                  max-height: 60vh; }
            body { padding-bottom: calc(52px + env(safe-area-inset-bottom, 0px)); }
        }
    </style>
</head>
<body>
    {% if session.get('user_id') %}
    <header class="topbar">
        <a class="brand" href="{{ url_for('index') }}" aria-label="庫存管理系統首頁"><span class="brand-mark"></span><span>庫存管理</span></a>
        <form class="gsearch" role="search" action="{{ url_for('index') }}" method="get">
            <label class="sr-only" for="gq">搜尋料號、品名、儲位或跨公司別名料號</label>
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" aria-hidden="true"><circle cx="11" cy="11" r="7"/><path d="M20 20l-3.5-3.5"/></svg>
            <input id="gq" type="search" name="q" value="{{ request.args.get('q', '') if request.path == '/' else '' }}" placeholder="料號、品名、儲位、別名料號…掃碼槍可直接掃">
        </form>
        <span class="user-info">{{ session.get('username') }}{% if session.get('is_admin') %}(管理員){% endif %}&nbsp;·&nbsp;<a href="{{ url_for('logout') }}">登出</a></span>
    </header>
    <nav>
        <a class="top{% if request.path == '/' %} active{% endif %}" href="{{ url_for('index') }}">總覽</a>
        {% set navp = request.path %}
        <details class="menu{% if navp in ('/stock/in', '/stock/out', '/search/image', '/labels') or navp.startswith('/receipts') or navp.startswith('/pick') %} here{% endif %}">
            <summary>現場收發</summary>
            <div class="menu-panel">
                <a {% if navp == '/stock/in' %}class="active" {% endif %}href="{{ url_for('stock_in') }}">入庫登記</a>
                <a {% if navp == '/stock/out' %}class="active" {% endif %}href="{{ url_for('stock_out') }}">出庫登記</a>
                <a {% if navp.startswith('/receipts') %}class="active" {% endif %}href="{{ url_for('receipts_page') }}">收貨單</a>
                <a {% if navp == '/search/image' %}class="active" {% endif %}href="{{ url_for('image_search') }}">以圖找料</a>
                <a {% if navp == '/labels' %}class="active" {% endif %}href="{{ url_for('labels_page') }}">料架標籤</a>
            </div>
        </details>
        <details class="menu{% if navp == '/alerts' or navp.startswith('/orders') or navp.startswith('/planning') %} here{% endif %}">
            <summary>採購進貨</summary>
            <div class="menu-panel">
                <a {% if navp.startswith('/orders') %}class="active" {% endif %}href="{{ url_for('orders_page') }}">採購單</a>
                <a {% if navp == '/alerts' %}class="active" {% endif %}href="{{ url_for('alerts') }}">短缺與效期</a>
                <a {% if navp.startswith('/planning') %}class="active" {% endif %}href="{{ url_for('planning_page') }}">補貨規劃</a>
            </div>
        </details>
        <details class="menu{% if navp in ('/report', '/history') or navp.startswith('/counts') or navp.startswith('/reservations') %} here{% endif %}">
            <summary>查帳與報表</summary>
            <div class="menu-panel">
                <a {% if navp == '/report' %}class="active" {% endif %}href="{{ url_for('report') }}">庫存報表</a>
                <a {% if navp == '/history' %}class="active" {% endif %}href="{{ url_for('history') }}">異動歷史</a>
                <a {% if navp.startswith('/counts') %}class="active" {% endif %}href="{{ url_for('counts_page') }}">循環盤點</a>
                <a {% if navp.startswith('/reservations') %}class="active" {% endif %}href="{{ url_for('reservations_page') }}">預留</a>
            </div>
        </details>
        <details class="menu{% if navp in ('/products/new', '/import', '/audit') or navp.startswith('/suppliers') or navp.startswith('/users') %} here{% endif %}">
            <summary>{% if session.get('is_admin') %}系統設定{% else %}商品資料{% endif %}</summary>
            <div class="menu-panel">
                <a {% if navp == '/products/new' %}class="active" {% endif %}href="{{ url_for('product_new') }}">新增商品</a>
                <a {% if navp.startswith('/suppliers') %}class="active" {% endif %}href="{{ url_for('suppliers') }}">供應商</a>
                {% if session.get('is_admin') %}
                <a {% if navp == '/import' %}class="active" {% endif %}href="{{ url_for('csv_import') }}">CSV 匯入</a>
                <a {% if navp.startswith('/users') %}class="active" {% endif %}href="{{ url_for('users_page') }}">帳號管理</a>
                <a {% if navp == '/audit' %}class="active" {% endif %}href="{{ url_for('audit_page') }}">稽核紀錄</a>
                {% endif %}
            </div>
        </details>
    </nav>
    <div class="rail">
        {% for c in rail %}
        <a class="{{ c['tone'] }}" href="{{ c['href'] }}">
            <span class="k">{{ c['label'] }}</span>
            <span class="v">{{ c['value'] }}{% if c['unit'] %}<span class="unit">{{ c['unit'] }}</span>{% endif %}</span>
            <span class="s">{{ c['sub'] }}</span>
        </a>
        {% endfor %}
    </div>
    {% endif %}
    <div class="container{% if narrow %} narrow{% endif %}">
        {% if back_url %}<p class="crumb"><a href="{{ back_url }}">← {{ back_label }}</a></p>{% endif %}
        {% if error %}<div class="msg error">{{ error }}</div>{% endif %}
        {% if msg %}<div class="msg ok">{{ msg }}</div>{% endif %}
        __BODY__
    </div>
    <footer>庫存管理系統 &copy; {{ year }}　·　版本 <span class="ver">{{ app_version }}</span></footer>
    {% if session.get('user_id') %}
    {# 手機常駐工具列。四個固定格是「從別的地方走過去」次數最多的作業:
       入庫、出庫是逐件發生的(一天幾十次),收貨是每次來車就要從任何頁面切過去。
       盤點不放進來——盤點是一天開一次盤點單、然後整天待在那一頁,
       放進來會排擠掉真正需要隨時切換的入口。盤點在「更多」第一個位置。 #}
    <nav class="tabbar" aria-label="主要功能">
        <a {% if request.path == '/' %}class="active" {% endif %}href="{{ url_for('index') }}">
            <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M4 10.5 12 4l8 6.5"/><path d="M6 9.8V20h12V9.8"/><path d="M10 20v-5.5h4V20"/></svg>總覽</a>
        <a {% if request.path == '/stock/in' %}class="active" {% endif %}href="{{ url_for('stock_in') }}">
            <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M12 3v10"/><path d="M8 9.5l4 4 4-4"/><path d="M4 16v3.5h16V16"/></svg>入庫</a>
        <a {% if request.path == '/stock/out' %}class="active" {% endif %}href="{{ url_for('stock_out') }}">
            <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M12 14V4"/><path d="M8 7.5l4-4 4 4"/><path d="M4 16v3.5h16V16"/></svg>出庫</a>
        <a {% if request.path.startswith('/receipts') %}class="active" {% endif %}href="{{ url_for('receipts_page') }}">
            <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M4 7.5 12 4l8 3.5v9L12 20l-8-3.5z"/><path d="M4 7.5 12 11l8-3.5"/><path d="M12 11v9"/></svg>收貨</a>
        <details class="more">
            <summary aria-label="更多功能">
                <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M4 7h16"/><path d="M4 12h16"/><path d="M4 17h16"/></svg>更多</summary>
            <div class="more-panel">
                <span class="who">{{ session.get('username') }}{% if session.get('is_admin') %}(管理員){% endif %}</span>
                <span class="grp">現場收發</span>
                <a {% if request.path.startswith('/counts') %}class="active" {% endif %}href="{{ url_for('counts_page') }}">循環盤點</a>
                <a {% if request.path == '/pick' %}class="active" {% endif %}href="{{ url_for('pick', to='history') }}">找料號</a>
                <a {% if request.path == '/search/image' %}class="active" {% endif %}href="{{ url_for('image_search') }}">以圖找料</a>
                <a {% if request.path == '/labels' %}class="active" {% endif %}href="{{ url_for('labels_page') }}">料架標籤</a>
                <span class="grp">採購進貨</span>
                <a {% if request.path.startswith('/orders') %}class="active" {% endif %}href="{{ url_for('orders_page') }}">採購單</a>
                <a {% if request.path == '/alerts' %}class="active" {% endif %}href="{{ url_for('alerts') }}">短缺與效期</a>
                <a {% if request.path.startswith('/planning') %}class="active" {% endif %}href="{{ url_for('planning_page') }}">補貨規劃</a>
                <span class="grp">查帳與報表</span>
                <a {% if request.path == '/report' %}class="active" {% endif %}href="{{ url_for('report') }}">庫存報表</a>
                <a {% if request.path == '/history' %}class="active" {% endif %}href="{{ url_for('history') }}">異動歷史</a>
                <a {% if request.path.startswith('/reservations') %}class="active" {% endif %}href="{{ url_for('reservations_page') }}">預留</a>
                <span class="grp">{% if session.get('is_admin') %}系統設定{% else %}商品資料{% endif %}</span>
                <a {% if request.path == '/products/new' %}class="active" {% endif %}href="{{ url_for('product_new') }}">新增商品</a>
                <a {% if request.path.startswith('/suppliers') %}class="active" {% endif %}href="{{ url_for('suppliers') }}">供應商</a>
                {% if session.get('is_admin') %}
                <a {% if request.path == '/import' %}class="active" {% endif %}href="{{ url_for('csv_import') }}">CSV 匯入</a>
                <a {% if request.path.startswith('/users') %}class="active" {% endif %}href="{{ url_for('users_page') }}">帳號管理</a>
                <a {% if request.path == '/audit' %}class="active" {% endif %}href="{{ url_for('audit_page') }}">稽核紀錄</a>
                {% endif %}
                <a href="{{ url_for('logout') }}">登出</a>
            </div>
        </details>
    </nav>
    {% endif %}
</body>
</html>
"""



def todo_rail():
    """待辦帶:全站常駐的四格。規則有三條——
    (1) 四格全部是可以「數」的件數,沒有一格需要把個、米、捲、箱加起來;
    (2) 即時計算不快取;
    (3) 沒有資料時說實話並給出路,不顯示一個會被當成「一切正常」的 0。"""
    db = get_db()
    one = lambda sql: db.execute(sql).fetchone()

    pending = one("SELECT COUNT(*) AS c FROM receipts WHERE status = 'open'")["c"]
    po = one("""
        SELECT COUNT(DISTINCT o.id) AS orders,
               COALESCE(SUM(i.ordered_qty - i.received_qty), 0) AS qty
        FROM purchase_orders o JOIN purchase_order_items i ON i.po_id = o.id
        WHERE o.status IN ('ordered', 'shipped', 'arrived')
          AND i.ordered_qty > i.received_qty""")
    has_threshold = one("SELECT COUNT(*) AS c FROM products WHERE low_stock_threshold > 0")["c"]
    short = one("""
        SELECT COUNT(*) AS c FROM products p
        WHERE p.low_stock_threshold > 0
          AND p.quantity - COALESCE((SELECT SUM(r.quantity) FROM reservations r
                WHERE r.product_id = p.id AND r.status = 'active'), 0) <= p.low_stock_threshold""")["c"]
    noloc = one("SELECT COUNT(*) AS c FROM products WHERE location IS NULL OR location = ''")["c"]
    tmpsku = one("SELECT COUNT(*) AS c FROM products WHERE sku LIKE 'TMP-%'")["c"]
    # 一項料可能同時無儲位又掛臨時料號,不能把兩個數字相加(會重複計算)
    pend_data = one("""SELECT COUNT(*) AS c FROM products
        WHERE location IS NULL OR location = '' OR sku LIKE 'TMP-%'""")["c"]
    total = one("SELECT COUNT(*) AS c FROM products")["c"]

    cells = [
        {"tone": "act" if pending else "none", "label": "待驗收貨",
         "value": str(pending), "unit": "張" if pending else "",
         "sub": "收貨單已上傳,等人核對" if pending else "沒有等著核對的收貨單",
         "href": url_for("receipts_page")},
        {"tone": "info" if po["orders"] else "none", "label": "在途採購",
         "value": str(po["orders"]), "unit": "張" if po["orders"] else "",
         "sub": f"{po['qty']:,} 件在路上" if po["orders"] else "沒有未到貨的採購單",
         "href": url_for("orders_page")},
    ]
    if has_threshold == 0 and total:
        # 全部商品都沒有補貨門檻時,「短缺 0」是騙人的——那不是「沒有短缺」,是「不知道」
        cells.append({"tone": "none", "label": "短缺", "value": "尚未設定", "unit": "",
                      "sub": f"{total:,} 項都還沒有補貨門檻",
                      "href": url_for("planning_page")})
    else:
        cells.append({"tone": "bad" if short else "none", "label": "短缺",
                      "value": str(short), "unit": "項" if short else "",
                      "sub": "低於補貨門檻,要補貨" if short else f"{has_threshold:,} 項有設門檻,目前都夠",
                      "href": url_for("alerts")})
    parts = []
    if tmpsku: parts.append(f"{tmpsku:,} 無正式料號")
    if noloc: parts.append(f"{noloc:,} 無儲位")
    cells.append({"tone": "bad" if pend_data else "none", "label": "資料待補",
                  "value": f"{pend_data:,}", "unit": "項" if pend_data else "",
                  "sub": " · ".join(parts) if parts else "料號與儲位都齊全",
                  "href": url_for("index", missing=1)})
    return cells


def build_pager(endpoint, page, has_next, **params):
    # 純連結分頁(零 JS);保留目前的搜尋/篩選參數。
    # 回傳 Markup 讓 Jinja2 直接輸出 HTML(否則會被自動跳脫成字面文字)
    parts = []
    if page > 1:
        prev_url = escape(url_for(endpoint, page=page - 1, **params))
        parts.append(f'<a class="plain" href="{prev_url}">← 上一頁</a>')
    parts.append(f'<span class="page-no">第 {page} 頁</span>')
    if has_next:
        next_url = escape(url_for(endpoint, page=page + 1, **params))
        parts.append(f'<a class="plain" href="{next_url}">下一頁 →</a>')
    return Markup('<p class="pager">' + '　'.join(parts) + '</p>')


def render_page(body, **ctx):
    ctx.setdefault("error", None)
    ctx.setdefault("msg", None)
    # 每頁自己的分頁標題(WCAG 2.4.2 最低 A 級);title 是各頁本來就會傳的頁名
    ctx.setdefault("page_title", ctx.get("title"))
    ctx.setdefault("narrow", False)       # 純表單頁收窄容器
    ctx.setdefault("back_url", None)      # 明細頁的具名返回連結
    ctx.setdefault("back_label", "返回")
    if "rail" not in ctx:
        ctx["rail"] = todo_rail() if session.get("user_id") else []
    ctx.setdefault("year", datetime.now().year)
    ctx.setdefault("app_version", APP_VERSION)
    return render_template_string(LAYOUT.replace("__BODY__", body), **ctx)


PAGE_REGISTER = """
<div class="auth-box">
{% if closed %}
<h1>不開放自助註冊</h1>
<p>本系統已完成初始設定。為了資料安全,新帳號一律由管理員建立。</p>
<p><a class="plain" href="{{ url_for('login') }}">前往登入</a></p>
{% else %}
<h1>建立管理員帳號</h1>
<p class="note">這是系統的第一個帳號,將自動成為管理員(可管理帳號、刪除資料、CSV 匯入)。建立後系統即關閉自助註冊。</p>
<form method="post">
    <label>帳號</label><input type="text" name="username" value="{{ username or '' }}">
    <label>密碼(至少 8 碼)</label><input type="password" name="password">
    <input type="submit" value="建立管理員帳號">
</form>
<p>已有帳號?<a class="plain" href="{{ url_for('login') }}">前往登入</a></p>
{% endif %}
</div>
"""

PAGE_LOGIN = """
<div class="auth-box">
<h1>登入庫存管理系統</h1>
<form method="post">
    <label>帳號</label><input type="text" name="username" value="{{ username or '' }}">
    <label>密碼</label><input type="password" name="password">
    <input type="submit" value="登入">
</form>
<p>還沒有帳號?<a class="plain" href="{{ url_for('register') }}">前往註冊</a></p>
</div>
"""

PAGE_INDEX = """
<div class="head">
    <h1>{% if missing %}資料待補的商品{% elif q %}符合「{{ q }}」的商品{% elif category %}分類:{{ category }}{% else %}全部商品{% endif %}</h1>
    <div class="acts">
        {% if q or category or missing %}<a href="{{ url_for('index') }}">看全部商品</a>{% else %}<a href="{{ url_for('index', missing=1) }}">只看資料待補</a>{% endif %}
        {% if cols_full %}
        <a href="{{ url_for('index', q=q, category=category, missing=(1 if missing else None)) }}">只看常用 7 欄</a>
        {% else %}
        <a href="{{ url_for('index', q=q, category=category, cols='full', missing=(1 if missing else None)) }}">顯示全部欄位</a>
        {% endif %}
        <a href="{{ url_for('export_inventory') }}">匯出 CSV</a>
    </div>
</div>
{% if rows %}
<div class="pane">
    <div class="pane-h">庫存清單<span class="r">
        <form method="get" class="filters" style="margin:0">
            {% if q %}<input type="hidden" name="q" value="{{ q }}">{% endif %}
            {% if cols_full %}<input type="hidden" name="cols" value="full">{% endif %}
            {% if missing %}<input type="hidden" name="missing" value="1">{% endif %}
            <label class="sr-only" for="cat">依分類篩選</label>
            <select id="cat" name="category">
                <option value="">全部分類</option>
                {% for c in categories %}
                <option value="{{ c }}" {% if c == category %}selected{% endif %}>{{ c }}</option>
                {% endfor %}
            </select>
            <input type="submit" value="套用">
        </form>
        第 {{ page }} 頁</span></div>
    <div class="table-scroll">
        <table class="cards">
            <tr>
                <th>料號</th><th>名稱</th><th>儲位</th><th class="num">現貨</th><th class="num">可用</th>
                {% if cols_full %}<th class="num">在途</th><th>別名料號</th><th>分類</th><th>單位</th><th class="num">單價</th><th class="num">低庫存門檻</th><th>供應商</th>{% endif %}
                <th>狀態</th><th>操作</th>
            </tr>
            {% for p in rows %}
            <tr{% if p['low'] %} class="low-stock"{% endif %}>
                <td data-label="料號" class="mono">{{ p['sku'] }}</td>
                <td data-label="名稱"><a class="plain" href="{{ url_for('product_detail', pid=p['id']) }}">{{ p['name'] }}</a></td>
                <td data-label="儲位" class="mono">{{ p['location'] or '—' }}</td>
                <td data-label="現貨" class="num" id="qty-{{ p['id'] }}">{{ p['quantity'] }}</td>
                <td data-label="可用" class="num" id="avail-{{ p['id'] }}">{{ p['available'] }}</td>
                {% if cols_full %}
                <td data-label="在途" class="num">{% if p['onorder'] %}<span class="onorder-cell">{{ p['onorder'] }}</span>{% else %}—{% endif %}</td>
                <td data-label="別名料號" class="alias-cell">{{ p['alias_text'] or '—' }}</td>
                <td data-label="分類">{{ p['category'] }}</td>
                <td data-label="單位">{{ p['unit'] }}</td>
                <td data-label="單價" class="num">{{ p['unit_price_str'] }}</td>
                <td data-label="低庫存門檻" class="num">{{ p['low_stock_threshold'] }}</td>
                <td data-label="供應商">{{ p['supplier_name'] or '—' }}</td>
                {% endif %}
                <td data-label="狀態" class="chips">
                    {% if p['onorder'] %}<span class="chip info">在途 {{ p['onorder'] }}</span>{% endif %}
                    {% if p['low'] %}<span class="chip bad">短缺</span>{% endif %}
                    {% if p['reserved'] %}<span class="chip held">已預留 {{ p['reserved'] }}</span>{% endif %}
                    {% if not p['location'] %}<span class="chip off">無儲位</span>{% endif %}
                </td>
                <td data-label="操作">
                    <a class="btn sm ghost" href="{{ url_for('stock_in', product_id=p['id']) }}">入庫</a>
                    <a class="btn sm ghost" href="{{ url_for('stock_out', product_id=p['id']) }}">出庫</a>
                    <a class="plain" href="{{ url_for('product_edit', pid=p['id']) }}">編輯</a>
                    {% if session.get('is_admin') %}
                    <form class="inline" method="post" action="{{ url_for('product_delete', pid=p['id']) }}"
                          onsubmit="return confirm('確定刪除「{{ p['name'] }}」?此操作無法復原。');">
                        <button class="small-btn icon-btn" type="submit" title="刪除商品" aria-label="刪除商品"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M4 7h16"/><path d="M9.5 7V5.5a1 1 0 0 1 1-1h3a1 1 0 0 1 1 1V7"/><path d="M6.5 7l.8 12a1.5 1.5 0 0 0 1.5 1.4h6.4a1.5 1.5 0 0 0 1.5-1.4l.8-12"/><path d="M10.5 11v6M13.5 11v6"/></svg></button>
                    </form>
                    {% endif %}
                </td>
            </tr>
            {% endfor %}
        </table>
    </div>
</div>
{{ pager }}
{% else %}
<div class="pane"><div class="pane-b"><div class="empty">
    <div class="h">{% if missing %}沒有資料待補的商品{% else %}查無商品{% endif %}</div>
    <div class="p">{% if missing %}所有商品的料號與儲位都齊全。{% else %}換個關鍵字試試,或用對方公司的料號、儲位再找一次。{% endif %}</div>
    <div class="btnrow" style="justify-content:center">
        <a class="btn ghost" href="{{ url_for('index') }}">看全部商品</a>
        <a class="btn ghost" href="{{ url_for('product_new') }}">新增商品</a>
    </div>
</div></div></div>
{% endif %}
"""


PAGE_ALERTS = """
<h1>短缺與效期</h1>
{% if products %}
<div class="pane">
    <div class="pane-h">要補貨的料<span class="r">{{ products|length }} 項 · 依可用量由少到多</span></div>
    <div class="table-scroll">
        <table class="cards">
            <tr><th>料號</th><th>名稱</th><th>儲位</th><th class="num">現貨</th><th class="num">可用</th>
                <th class="num">門檻</th><th class="num">建議補量</th><th>狀態</th><th>供應商</th><th>操作</th></tr>
            {% for p in products %}
            <tr class="low-stock">
                <td data-label="料號" class="mono">{{ p['sku'] }}</td>
                <td data-label="名稱"><a class="plain" href="{{ url_for('product_detail', pid=p['id']) }}">{{ p['name'] }}</a></td>
                <td data-label="儲位" class="mono">{{ p['location'] or '—' }}</td>
                <td data-label="現貨" class="num">{{ p['quantity'] }}</td>
                <td data-label="可用" class="num">{{ p['available'] }}</td>
                <td data-label="門檻" class="num">{{ p['low_stock_threshold'] }}</td>
                <td data-label="建議補量" class="num"><b>{{ p['suggest'] }}</b></td>
                <td data-label="狀態" class="chips">
                    {% if p['onorder'] %}<span class="chip info">已下訂 {{ p['onorder'] }}</span>
                    {% else %}<span class="chip bad">尚未下訂</span>{% endif %}
                </td>
                <td data-label="供應商">{{ p['supplier_name'] or '—' }}</td>
                <td data-label="操作">
                    <a class="btn sm ghost" href="{{ url_for('order_new', product_id=p['id'], qty=p['suggest']) }}">建立採購單</a>
                    <a class="btn sm ghost" href="{{ url_for('stock_in', product_id=p['id']) }}">入庫</a>
                </td>
            </tr>
            {% endfor %}
        </table>
    </div>
</div>
{% else %}
<div class="pane"><div class="pane-b"><div class="empty">
    <div class="h">{% if has_threshold %}目前沒有低於門檻的料{% else %}還沒有設定任何補貨門檻{% endif %}</div>
    <div class="p">{% if has_threshold %}有設門檻的 {{ has_threshold }} 項料目前都夠用。{% else %}沒有門檻就算不出短缺——顯示「短缺 0」會讓人以為一切正常,其實是「不知道」。補貨規劃會用最近 90 天的實際出庫量幫你推導建議門檻,可以一次套用。{% endif %}</div>
    <div class="btnrow" style="justify-content:center">
        <a class="btn{% if not has_threshold %}{% else %} ghost{% endif %}" href="{{ url_for('planning_page') }}">前往補貨規劃</a>
        <a class="btn ghost" href="{{ url_for('index') }}">回庫存總覽</a>
    </div>
</div></div></div>
{% endif %}

<h2>效期警示</h2>
{% if expiring %}
<div class="pane">
    <div class="pane-h">30 天內到期或已過期的批次<span class="r">{{ expiring|length }} 批</span></div>
    <div class="table-scroll">
        <table class="cards">
            <tr><th>批號</th><th>名稱</th><th class="num">剩餘</th><th>有效期</th><th>狀態</th><th>操作</th></tr>
            {% for l in expiring %}
            <tr{% if l['expired'] %} class="low-stock"{% endif %}>
                <td data-label="批號" class="mono">{{ l['lot_no'] }}</td>
                <td data-label="名稱"><a class="plain" href="{{ url_for('product_detail', pid=l['product_id']) }}">{{ l['name'] }}</a></td>
                <td data-label="剩餘" class="num">{{ l['qty_remaining'] }}{{ l['unit'] }}</td>
                <td data-label="有效期" class="mono">{{ l['expiry_date'] }}</td>
                <td data-label="狀態" class="chips">
                    {% if l['expired'] %}<span class="chip expired">已過期</span>
                    {% else %}<span class="chip act">剩 {{ l['days_left'] }} 天</span>{% endif %}
                </td>
                <td data-label="操作"><a class="btn sm ghost" href="{{ url_for('stock_out', product_id=l['product_id']) }}">出庫</a></td>
            </tr>
            {% endfor %}
        </table>
    </div>
</div>
{% else %}
<p class="note">沒有即將到期的批次(僅統計有設定有效期的批次)。</p>
{% endif %}
"""


PAGE_PRODUCT_FORM = """
<h1>{{ title }}</h1>
<form method="post">
    <label>商品名稱 *</label><input type="text" name="name" value="{{ f['name'] }}">
    <label>SKU(商品編號)*</label><input type="text" name="sku" value="{{ f['sku'] }}">
    <label>分類</label><input type="text" name="category" value="{{ f['category'] }}">
    <label>單位</label><input type="text" name="unit" value="{{ f['unit'] }}">
    <label>單價</label><input type="text" name="unit_price" value="{{ f['unit_price'] }}">
    <label>低庫存門檻(0 = 不警示)</label><input type="number" name="low_stock_threshold" min="0" value="{{ f['low_stock_threshold'] }}">
    <label>儲位(料放在哪一架/格)</label><input type="text" name="location" value="{{ f['location'] }}" placeholder="例:A-03-2">
    <label>採購單位(選填,如「箱」)</label><input type="text" name="purchase_unit" value="{{ f['purchase_unit'] }}">
    <label>每採購單位含幾個庫存單位</label><input type="number" name="units_per_purchase" min="1" value="{{ f['units_per_purchase'] }}">
    <label>採購前置期(天,供安全庫存推導)</label><input type="number" name="lead_time_days" min="1" value="{{ f['lead_time_days'] }}">
    <label>目標服務水準(%)</label>
    <select name="service_level">
        {% for lv in [80, 85, 90, 95, 97, 98, 99] %}
        <option value="{{ lv }}" {% if f['service_level']|string == lv|string %}selected{% endif %}>{{ lv }}%</option>
        {% endfor %}
    </select>
    <label>出庫策略</label>
    <select name="issue_strategy">
        <option value="FIFO" {% if f['issue_strategy'] == 'FIFO' %}selected{% endif %}>FIFO 先進先出(依入庫時間)</option>
        <option value="FEFO" {% if f['issue_strategy'] == 'FEFO' %}selected{% endif %}>FEFO 先到期先出(有保存期限的料件)</option>
    </select>
    <label>供應商</label>
    <select name="supplier_id">
        <option value="">(不指定)</option>
        {% for s in supplier_list %}
        <option value="{{ s['id'] }}" {% if f['supplier_id']|string == s['id']|string %}selected{% endif %}>{{ s['name'] }}</option>
        {% endfor %}
    </select>
    <input type="submit" value="儲存">
</form>
<p class="note">庫存數量不在此處修改,請透過「入庫 / 出庫」登記異動。</p>
"""

PAGE_PICK = """
<h1>{{ title }}</h1>
<p class="note">掃碼槍可直接掃。只找到一筆時系統會自動帶你進下一步。</p>
<form method="get" class="hero-search">
    <input type="hidden" name="to" value="{{ to }}">
    {% for k, v in extra.items() %}<input type="hidden" name="{{ k }}" value="{{ v }}">{% endfor %}
    <span class="search-field">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" aria-hidden="true"><circle cx="11" cy="11" r="7"/><path d="m20 20-3.6-3.6"/></svg>
        <label class="sr-only" for="pq">輸入料號、品名或儲位</label>
        <input id="pq" type="text" name="q" value="{{ q }}" placeholder="料號、品名、儲位、跨公司別名料號" autofocus>
    </span>
    <input type="submit" value="搜尋">
</form>
{% if q and not rows %}
<p class="note">找不到符合「{{ q }}」的料。可以改用品名的一部分、儲位,或對方公司的料號再試一次。</p>
{% endif %}
{% if rows %}
<div class="pane">
    <div class="pane-h">{% if q %}符合「{{ q }}」的料{% else %}最近登記過的料{% endif %}<span class="r">{{ rows|length }} 筆{% if truncated %}(只顯示前 {{ rows|length }} 筆,請輸入更完整的料號){% endif %}</span></div>
    <div class="pane-b">
    <div class="picklist">
        {% for p in rows %}
        {% if post_to %}
        <form method="post" action="{{ post_to }}">
            <input type="hidden" name="product_id" value="{{ p['id'] }}">
            {% if remember %}<input type="hidden" name="remember" value="1">{% endif %}
            <button class="pickrow" type="submit" style="width:100%">
                <span class="sku">{{ p['sku'] }}</span>
                <span class="nm">{{ p['name'] }}</span>
                <span class="loc">{{ p['location'] or '無儲位' }}</span>
                <span class="pq">{{ p['quantity'] }}<span class="unit">{{ p['unit'] }}</span></span>
            </button>
        </form>
        {% else %}
        <a class="pickrow" href="{{ link_base }}{{ p['id'] }}">
            <span class="sku">{{ p['sku'] }}</span>
            <span class="nm">{{ p['name'] }}</span>
            <span class="loc">{{ p['location'] or '無儲位' }}</span>
            <span class="pq">{{ p['quantity'] }}<span class="unit">{{ p['unit'] }}</span></span>
        </a>
        {% endif %}
        {% endfor %}
    </div>
    </div>
</div>
{% endif %}
"""

PAGE_STOCK_FORM = """
<h1>{{ title }}</h1>
<div class="pickcard">
    <div>
        <div class="nm">{{ prod['name'] }}</div>
        <div class="meta">{{ prod['sku'] }}　·　儲位 {{ prod['location'] or '未設定' }}　·　{{ prod['supplier_name'] or '無供應商' }}</div>
    </div>
    <div class="right">
        <div class="big">{{ prod['quantity'] }}</div>
        <div class="meta">目前現貨({{ prod['unit'] }}){% if prod['reserved'] %}　·　已預留 {{ prod['reserved'] }}{% endif %}　·　<a href="{{ url_for('pick', to=('in' if is_in else 'out')) }}">換一項料</a></div>
    </div>
</div>
<form method="post">
    <input type="hidden" name="product_id" value="{{ prod['id'] }}">
    <label for="qty">數量 <span class="req">必填</span></label>
    <input id="qty" class="w-qty" type="number" name="quantity" min="1" inputmode="numeric" value="{{ f['quantity'] }}">
    {% if is_in %}
    <label for="qu">數量單位</label>
    <select id="qu" name="qty_unit">
        <option value="stock" {% if f['qty_unit'] != 'purchase' %}selected{% endif %}>庫存單位({{ prod['unit'] }})——直接輸入實際數量</option>
        <option value="purchase" {% if f['qty_unit'] == 'purchase' %}selected{% endif %}>採購單位({{ prod['purchase_unit'] or '未設定' }})——自動換算</option>
    </select>
    <label for="lot">批號 <span class="opt">選填</span></label>
    <p class="field-help">不填會自動編號。填了以後可以追這批貨從哪來、什麼時候到。</p>
    <input id="lot" class="w-code" type="text" name="lot_no" value="{{ f['lot_no'] }}" placeholder="例:供應商批號">
    <label for="exp">有效期 <span class="opt">選填</span></label>
    <p class="field-help">設定後可做 FEFO 先到期先出與到期警示。</p>
    <input id="exp" class="w-code" type="date" name="expiry_date" value="{{ f['expiry_date'] }}">
    <label for="cost">成本單價 <span class="opt">選填</span></label>
    <p class="field-help">供加權平均成本報表。</p>
    <input id="cost" class="w-qty" type="text" name="unit_cost" value="{{ f['unit_cost'] }}" inputmode="decimal">
    {% endif %}
    <label for="pp">用途 / 工單 <span class="opt">選填</span></label>
    <input id="pp" class="w-code" type="text" name="purpose" value="{{ f['purpose'] }}" placeholder="例:WO-1001、產線A、維修">
    <label for="nt">備註 <span class="opt">選填</span></label>
    <input id="nt" type="text" name="note" value="{{ f['note'] }}">
    <input type="submit" value="{{ title }}">
    {% if not is_in %}
    <p class="note">出庫依 FIFO 先進先出原則,自動從最早入庫的批次扣減,消耗明細記錄於異動歷史。可出庫數量為<b>可用量</b>(現貨扣掉已預留)。</p>
    {% endif %}
</form>
{% if recent %}
<div class="pane" style="margin-top:24px">
    <div class="pane-h">本次已登記<span class="r">最近 {{ recent|length }} 筆 · 打錯可以直接沖銷</span></div>
    <div class="table-scroll">
        <table>
            <tr><th>時間(台灣)</th><th>類型</th><th class="num">數量</th><th>批號</th><th>操作人</th><th>操作</th></tr>
            {% for t in recent %}
            <tr>
                <td class="mono">{{ t['local_time'] }}</td>
                <td>{% if t['type'] == 'in' %}<span class="chip ok">入庫</span>{% else %}<span class="chip info">出庫</span>{% endif %}</td>
                <td class="num mono">{{ t['quantity'] }}</td>
                <td class="mono">{{ t['lot_info'] or '—' }}</td>
                <td>{{ t['username'] }}</td>
                <td data-label="操作">
                    {% if t['reversed'] %}<span class="chip off">已沖銷</span>
                    {% elif t['is_reversal'] %}<span class="chip off">沖銷紀錄</span>
                    {% else %}
                    <form class="inline" method="post" action="{{ url_for('transaction_reverse', tid=t['id']) }}"
                          onsubmit="return confirm('確定沖銷這筆{% if t['type'] == 'in' %}入庫{% else %}出庫{% endif %} {{ t['quantity'] }}?庫存與批次都會回到登記前的狀態。');">
                        <button class="small-btn" type="submit">沖銷這筆</button>
                    </form>
                    {% endif %}
                </td>
            </tr>
            {% endfor %}
        </table>
    </div>
</div>
{% endif %}
"""


PAGE_HISTORY = """
<h1>異動歷史</h1>
<form method="get" class="filters">
    {% if filters['product_id'] and picked_product %}
    <input type="hidden" name="product_id" value="{{ filters['product_id'] }}">
    <span class="chip info">只看 {{ picked_product['sku'] }} {{ picked_product['name'] }}</span>
    <a class="chip-link" href="{{ url_for('history') }}">看全部商品</a>
    {% else %}
    <label class="sr-only" for="hq">依料號、品名或儲位篩選商品</label>
    <input id="hq" type="text" name="pq" value="{{ filters['pq'] }}" placeholder="料號 / 品名 / 儲位(留白=全部商品)">
    {% endif %}
    <select name="type">
        <option value="">入庫+出庫</option>
        <option value="in" {% if filters['type'] == 'in' %}selected{% endif %}>只看入庫</option>
        <option value="out" {% if filters['type'] == 'out' %}selected{% endif %}>只看出庫</option>
    </select>
    起 <input type="date" name="start" value="{{ filters['start'] }}">
    迄 <input type="date" name="end" value="{{ filters['end'] }}">
    <input type="text" name="purpose" placeholder="用途／工單" value="{{ filters['purpose'] }}">
    <input type="submit" value="篩選">
    <a class="plain" href="{{ url_for('history') }}">清除</a>
    &nbsp;|&nbsp;
    <a class="plain" href="{{ url_for('export_transactions', **filters) }}">匯出異動 CSV</a>
</form>
{% if rows %}
<div class="table-scroll">
    <table>
        <tr><th>時間(台灣)</th><th>類型</th><th>商品</th><th>SKU</th><th class="num">數量</th><th>批次</th><th>用途／工單</th><th>備註</th><th>操作人員</th></tr>
        {% for r in rows %}
        <tr>
            <td>{{ r['created_local'] }}</td>
            <td>{% if r['type'] == 'in' %}入庫{% else %}出庫{% endif %}</td>
            <td>{{ r['product_name'] }}</td>
            <td>{{ r['sku'] }}</td>
            <td class="num">{{ r['quantity'] }}</td>
            <td class="alias-cell">{{ r['lot_info'] or '—' }}</td>
            <td>{{ r['purpose'] or '—' }}</td>
            <td>{{ r['note'] }}</td>
            <td>{{ r['username'] }}</td>
        </tr>
        {% endfor %}
    </table>
</div>
{{ pager }}
{% else %}
<p>沒有符合條件的異動紀錄。</p>
{% endif %}
"""

PAGE_SUPPLIERS = """
<h1>供應商管理</h1>
<div class="action-links">
    <a class="primary" href="{{ url_for('supplier_new') }}">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.9" stroke-linecap="round" aria-hidden="true"><path d="M12 5v14"/><path d="M5 12h14"/></svg>新增供應商</a>
</div>
{% if rows %}
<div class="table-scroll">
    <table>
        <tr><th>名稱</th><th>聯絡人</th><th>電話</th><th>備註</th><th>商品數</th><th>操作</th></tr>
        {% for s in rows %}
        <tr>
            <td>{{ s['name'] }}</td>
            <td>{{ s['contact'] }}</td>
            <td>{{ s['phone'] }}</td>
            <td>{{ s['note'] }}</td>
            <td>{{ s['product_count'] }}</td>
            <td>
                <a class="plain" href="{{ url_for('supplier_edit', sid=s['id']) }}">編輯</a>
                {% if session.get('is_admin') %}
                <form class="inline" method="post" action="{{ url_for('supplier_delete', sid=s['id']) }}"
                      onsubmit="return confirm('確定刪除供應商「{{ s['name'] }}」?其商品的供應商欄位將被清空。');">
                    <button class="small-btn icon-btn" type="submit" title="刪除供應商" aria-label="刪除供應商"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M4 7h16"/><path d="M9.5 7V5.5a1 1 0 0 1 1-1h3a1 1 0 0 1 1 1V7"/><path d="M6.5 7l.8 12a1.5 1.5 0 0 0 1.5 1.4h6.4a1.5 1.5 0 0 0 1.5-1.4l.8-12"/><path d="M10.5 11v6M13.5 11v6"/></svg></button>
                </form>
                {% endif %}
            </td>
        </tr>
        {% endfor %}
    </table>
</div>
{% else %}
<p>目前沒有供應商資料。</p>
{% endif %}
"""

PAGE_SUPPLIER_FORM = """
<h1>{{ title }}</h1>
<form method="post">
    <label>供應商名稱 *</label><input type="text" name="name" value="{{ f['name'] }}">
    <label>聯絡人</label><input type="text" name="contact" value="{{ f['contact'] }}">
    <label>電話</label><input type="text" name="phone" value="{{ f['phone'] }}">
    <label>備註</label><input type="text" name="note" value="{{ f['note'] }}">
    <input type="submit" value="儲存">
</form>
"""

PAGE_REPORT = """
<h1>庫存報表</h1>
<div class="pane">
    <div class="pane-h">這份報表在說什麼<span class="r">
        <form method="get" class="filters" style="margin:0">
            起 <input type="date" name="start" value="{{ start }}">
            迄 <input type="date" name="end" value="{{ end }}">
            <input type="submit" value="套用區間">
            <a class="plain" href="{{ url_for('report') }}">清除</a>
        </form></span></div>
    <div class="pane-b">
    <div class="stat-row">
        {% for k in verdicts %}
        <div class="stat-box {{ k['tone'] }}">
            <div class="stat-cap">{{ k['label'] }}</div>
            <div class="stat-num"{% if k['id'] %} id="{{ k['id'] }}"{% endif %}{% if k['small'] %} style="font-size:18px"{% endif %}>{{ k['value'] }}</div>
            <div class="stat-cap">{{ k['say'] }}</div>
        </div>
        {% endfor %}
    </div>
    <p class="note">入庫/出庫總量統計{% if start or end %}套用上方日期區間{% else %}為全部期間{% endif %};「目前庫存」「庫存價值」「平均成本」一律為現時狀態。</p>
    </div>
</div>

<div class="pane">
    <div class="pane-h">庫齡分析(Inventory Aging)<span class="r">依各批次入庫時間;90 天以上通常代表呆滯風險</span></div>
    <div class="pane-b">
    <div class="bars">
        {% for a in aging %}
        <div class="bar">
            <span>{{ a['label'] }}</span>
            <span class="track"><span class="fill" style="width:{{ a['pct'] }}%;background:{{ a['color'] }}"></span></span>
            <span class="num mono">{{ a['count'] }} 批</span>
        </div>
        {% endfor %}
    </div>
    </div>
</div>
<div class="table-scroll">
    <table>
        <tr><th>SKU</th><th>名稱</th><th class="num">入庫總量</th><th class="num">出庫總量</th><th class="num">淨變動</th><th class="num">目前庫存</th><th class="num">單價</th><th class="num">平均成本</th><th class="num">庫存價值</th></tr>
        {% for r in rows %}
        <tr>
            <td>{{ r['sku'] }}</td>
            <td>{{ r['name'] }}</td>
            <td class="num" id="in-{{ r['id'] }}">{{ r['total_in'] }}</td>
            <td class="num" id="out-{{ r['id'] }}">{{ r['total_out'] }}</td>
            <td class="num">{{ r['net'] }}</td>
            <td class="num" id="qty-{{ r['id'] }}">{{ r['quantity'] }}</td>
            <td class="num">{{ r['unit_price_str'] }}</td>
            <td class="num">{{ r['avg_cost_str'] }}</td>
            <td class="num" id="value-{{ r['id'] }}">{{ r['value_str'] }}</td>
        </tr>
        {% endfor %}
        <tfoot>
        <tr>
            <td colspan="2">總計</td>
            <td class="num" id="total-in">{{ total_in }}</td>
            <td class="num" id="total-out">{{ total_out }}</td>
            <td class="num">{{ total_net }}</td>
            <td class="num" id="total-qty">{{ total_qty }}</td>
            <td></td>
            <td></td>
            <td class="num" id="total-value">{{ total_value_str }}</td>
        </tr>
            </tfoot>
</table>
</div>
<p class="note">平均成本為加權平均成本(存貨計價):僅以尚有剩餘且有登記成本的批次計算,「—」表示無成本資料。</p>

<div class="detail-section">
</div>

<div class="detail-section">
    <h2>ABC 分析(柏拉圖法則)</h2>
    <div class="table-scroll">
        <table>
            <tr><th>ABC</th><th>XYZ</th><th>組合</th><th>SKU</th><th>名稱</th><th class="num">庫存價值</th><th class="num">價值占比</th><th class="num">累積占比</th><th class="num">變異係數</th></tr>
            {% for r in abc_rows %}
            <tr>
                <td><strong>{{ r['abc'] }}</strong></td>
                <td>{{ r['xyz'] }}</td>
                <td><strong>{{ r['abc_xyz'] }}</strong></td>
                <td>{{ r['sku'] }}</td>
                <td>{{ r['name'] }}</td>
                <td class="num">{{ r['value_str'] }}</td>
                <td class="num">{{ r['pct_str'] }}%</td>
                <td class="num">{{ r['cum_pct_str'] }}%</td>
                <td class="num">{{ r['cv_str'] }}</td>
            </tr>
            {% endfor %}
        </table>
    </div>
    <p class="note">
        A 級(累積價值 ≤80%)是最該重點盤點與控管的少數關鍵料號;B 級(≤95%)次之;C 級數量多但價值低,可放寬管理頻率。<br>
        XYZ 依需求變異係數分級:X 穩定(CV&lt;0.5)、Y 中等(0.5–1.0)、Z 高度波動(&gt;1.0)。
        組合的用法:<strong>AX</strong> 可壓低庫存高頻補貨,<strong>AZ</strong> 最棘手(要備量又怕呆滯),
        <strong>CZ</strong> 乾脆多備一點——管理成本高於料件本身。
    </p>
</div>

<div class="detail-section">
    <h2>庫存準確率(帳實相符)</h2>
    {% if accuracy_info %}
    <div class="stat-row">
        <div class="stat-box">
            <div class="stat-num" id="accuracy-value">{{ accuracy_info['accuracy_str'] }}%</div>
            <div class="stat-cap">最近一次盤點的相符比例</div>
        </div>
        <div class="stat-box">
            <div class="stat-num" style="font-size:16px">{{ accuracy_info['name'] }}</div>
            <div class="stat-cap">過帳於 {{ accuracy_info['posted_local'] }}</div>
        </div>
    </div>
    <p class="note">業界基準為 95–99%。低於此區間代表現場作業與系統登記之間有系統性落差,應提高盤點頻率並追查原因。</p>
    {% else %}
    <p>尚無已過帳的盤點單。<a class="plain" href="{{ url_for('counts_page') }}">建立第一張盤點單</a>後,這裡會顯示庫存準確率。</p>
    {% endif %}
</div>
"""

PAGE_PRODUCT_DETAIL = """
<h1>商品詳細:{{ p['name'] }}</h1>
{# 動作放在規格表之前。這頁最常見的來源是現場掃料架上的 QR,掃完就是要登記進出;
   放在表格後面時實測手機上「入庫」在 966px 處,得先捲過一整屏才看得到。 #}
<div class="action-links">
    <a class="primary" href="{{ url_for('stock_in', product_id=p['id']) }}">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M12 3v10"/><path d="M8 9.5l4 4 4-4"/><path d="M4 16v3.5h16V16"/></svg>入庫</a>
    <a class="primary out" href="{{ url_for('stock_out', product_id=p['id']) }}">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M12 14V4"/><path d="M8 7.5l4-4 4 4"/><path d="M4 16v3.5h16V16"/></svg>出庫</a>
    <a href="{{ url_for('product_edit', pid=p['id']) }}">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M4 20h4l10-10-4-4L4 16z"/><path d="M13.5 6.5l4 4"/></svg>編輯基本資料</a>
    <a href="{{ url_for('history', product_id=p['id']) }}">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M12 7v5l3.5 2"/><circle cx="12" cy="12" r="8.5"/></svg>完整異動歷史</a>
</div>
<div class="table-scroll">
    <table class="cards">
        <tr><th>SKU</th><th>儲位</th><th>分類</th><th class="num">現貨</th><th class="num">可用</th><th>單位</th><th class="num">單價</th><th class="num">低庫存門檻</th><th>出庫策略</th><th>供應商</th></tr>
        <tr>
            <td data-label="SKU">{{ p['sku'] }}</td>
            <td data-label="儲位">{{ p['location'] or '—' }}</td>
            <td data-label="分類">{{ p['category'] }}</td>
            <td data-label="現貨" class="num" id="qty-{{ p['id'] }}">{{ p['quantity'] }}</td>
            <td data-label="可用" class="num">{{ p['available'] }}{% if p['reserved'] %} <span class="resv-note">(預留 {{ p['reserved'] }})</span>{% endif %}</td>
            <td data-label="單位">{{ p['unit'] }}{% if p['purchase_unit'] %}(採購:1 {{ p['purchase_unit'] }} = {{ p['units_per_purchase'] }} {{ p['unit'] }}){% endif %}</td>
            <td data-label="單價" class="num">{{ p['unit_price_str'] }}</td>
            <td data-label="低庫存門檻" class="num">{{ p['low_stock_threshold'] }}</td>
            <td data-label="出庫策略">{{ p['issue_strategy'] }}</td>
            <td data-label="供應商">{{ p['supplier_name'] or '—' }}</td>
        </tr>
    </table>
</div>
<div class="qr-row">
    <img class="qr-img" src="{{ url_for('product_qr', pid=p['id']) }}" alt="料號 QR">
    <div>
        <div class="qr-cap">料架標籤 QR</div>
        <p class="note">列印貼在料架上,手機掃碼即可直接開啟本頁登記進出。
        <a class="plain" href="{{ url_for('labels_page') }}">批次列印所有標籤</a></p>
    </div>
</div>

<div class="detail-section">
    <h2>照片</h2>
    {% if images %}
    <div class="photo-wall">
        {% for img in images %}
        <div class="photo-item">
            <img src="{{ url_for('serve_image', filename=img['filename']) }}" alt="{{ p['name'] }}">
            {% if session.get('is_admin') %}
            <form class="inline" method="post" action="{{ url_for('image_delete', pid=p['id'], img_id=img['id']) }}"
                  onsubmit="return confirm('確定刪除這張照片?');">
                <button class="small-btn icon-btn" type="submit" title="刪除照片" aria-label="刪除照片"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M4 7h16"/><path d="M9.5 7V5.5a1 1 0 0 1 1-1h3a1 1 0 0 1 1 1V7"/><path d="M6.5 7l.8 12a1.5 1.5 0 0 0 1.5 1.4h6.4a1.5 1.5 0 0 0 1.5-1.4l.8-12"/><path d="M10.5 11v6M13.5 11v6"/></svg></button>
            </form>
            {% endif %}
        </div>
        {% endfor %}
    </div>
    {% else %}
    <p>尚未上傳照片。</p>
    {% endif %}
    <form method="post" enctype="multipart/form-data" action="{{ url_for('image_upload', pid=p['id']) }}">
        <input type="file" name="photo" accept="image/*">
        <input type="submit" value="上傳照片">
    </form>
</div>

<div class="detail-section">
    <h2>在途採購</h2>
    {% if open_pos %}
    <p class="note">這項料目前還有 <strong>{{ onorder_total }}</strong> 在路上。下單前先看這裡,避免重複採購。</p>
    <div class="table-scroll">
        <table class="cards">
            <tr><th>採購單</th><th>狀態</th><th>預計到貨</th><th class="num">訂購</th><th class="num">已收</th><th class="num">在途</th><th>操作</th></tr>
            {% for o in open_pos %}
            <tr>
                <td data-label="採購單">{{ o['po_no'] or '(未填單號)' }}</td>
                <td data-label="狀態"><span class="chip-po chip-{{ o['status'] }}">{{ o['status_label'] }}</span></td>
                <td data-label="預計到貨">{{ o['eta'] or '—' }}</td>
                <td data-label="訂購" class="num">{{ o['ordered_qty'] }}</td>
                <td data-label="已收" class="num">{{ o['received_qty'] }}</td>
                <td data-label="在途" class="num"><span class="onorder-cell">{{ o['onorder'] }}</span></td>
                <td data-label="操作"><a class="plain" href="{{ url_for('order_detail', oid=o['id']) }}">開啟</a></td>
            </tr>
            {% endfor %}
        </table>
    </div>
    {% else %}
    <p class="note">目前沒有這項料的在途採購單。</p>
    {% endif %}
</div>

<div class="detail-section">
    <h2>跨公司料號對照</h2>
    {% if aliases %}
    <div class="table-scroll">
        <table>
            <tr><th>公司</th><th>該公司料號</th><th>備註</th><th>操作</th></tr>
            {% for a in aliases %}
            <tr>
                <td>{{ a['company'] }}</td>
                <td>{{ a['alias_sku'] }}</td>
                <td>{{ a['note'] }}</td>
                <td>
                    {% if session.get('is_admin') %}
                    <form class="inline" method="post" action="{{ url_for('alias_delete', pid=p['id'], aid=a['id']) }}"
                          onsubmit="return confirm('確定刪除別名「{{ a['company'] }}:{{ a['alias_sku'] }}」?');">
                        <button class="small-btn icon-btn" type="submit" title="刪除別名" aria-label="刪除別名"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M4 7h16"/><path d="M9.5 7V5.5a1 1 0 0 1 1-1h3a1 1 0 0 1 1 1V7"/><path d="M6.5 7l.8 12a1.5 1.5 0 0 0 1.5 1.4h6.4a1.5 1.5 0 0 0 1.5-1.4l.8-12"/><path d="M10.5 11v6M13.5 11v6"/></svg></button>
                    </form>
                    {% else %}—{% endif %}
                </td>
            </tr>
            {% endfor %}
        </table>
    </div>
    {% else %}
    <p>尚無別名料號。不同公司對同一物料的料號都可以登記在這裡,搜尋時任一料號都找得到。</p>
    {% endif %}
    <form method="post" action="{{ url_for('alias_add', pid=p['id']) }}">
        <label>公司名稱</label><input type="text" name="company">
        <label>該公司料號</label><input type="text" name="alias_sku">
        <label>備註</label><input type="text" name="note">
        <input type="submit" value="新增別名">
    </form>
</div>

<div class="detail-section">
    <h2>批次庫存(FIFO 先進先出)</h2>
    {% if lots %}
    <div class="table-scroll">
        <table class="cards">
            <tr><th>批號</th><th>入庫時間(台灣)</th><th class="num">庫齡(天)</th><th>有效期</th><th class="num">剩餘 / 原始</th><th class="num">成本單價</th><th>備註</th></tr>
            {% for l in lots %}
            <tr{% if l['qty_remaining'] == 0 %} class="lot-empty"{% endif %}>
                <td data-label="批號">{{ l['lot_no'] }}</td>
                <td data-label="入庫時間">{{ l['received_local'] }}</td>
                <td data-label="庫齡(天)" class="num">{{ l['age_days'] }}</td>
                <td data-label="有效期">{{ l['expiry_date'] or '—' }}</td>
                <td data-label="剩餘/原始" class="num">{{ l['qty_remaining'] }} / {{ l['qty_received'] }}</td>
                <td data-label="成本單價" class="num">{{ l['cost_str'] }}</td>
                <td data-label="備註">{{ l['note'] }}</td>
            </tr>
            {% endfor %}
        </table>
    </div>
    <p class="note">出庫依 FIFO 先進先出原則自動從最早批次扣減;各筆出庫的批次消耗明細見異動歷史。</p>
    {% else %}
    <p>尚無批次紀錄,入庫後會自動建立批次。</p>
    {% endif %}
</div>

<div class="detail-section">
    <h2>近期異動</h2>
    {% if recent_tx %}
    <div class="table-scroll">
        <table>
            <tr><th>時間(台灣)</th><th>類型</th><th>數量</th><th>備註</th><th>操作人員</th></tr>
            {% for r in recent_tx %}
            <tr>
                <td>{{ r['created_local'] }}</td>
                <td>{% if r['type'] == 'in' %}入庫{% else %}出庫{% endif %}</td>
                <td>{{ r['quantity'] }}</td>
                <td>{{ r['note'] }}</td>
                <td>{{ r['username'] }}</td>
            </tr>
            {% endfor %}
        </table>
    </div>
    {% else %}
    <p>尚無異動紀錄。</p>
    {% endif %}
</div>
"""

PAGE_USERS = """
<h1>帳號管理</h1>
<div class="table-scroll">
    <table class="cards">
        <tr><th>帳號</th><th>角色</th><th>建立時間</th><th>操作</th></tr>
        {% for u in users %}
        <tr>
            <td data-label="帳號">{{ u['username'] }}</td>
            <td data-label="角色">{% if u['is_admin'] %}管理員{% else %}一般使用者{% endif %}</td>
            <td data-label="建立時間">{{ u['created_local'] }}</td>
            <td data-label="操作">
                <form class="inline" method="post" action="{{ url_for('user_delete', uid=u['id']) }}"
                      onsubmit="return confirm('確定刪除帳號「{{ u['username'] }}」?');">
                    <button class="small-btn icon-btn" type="submit" title="刪除帳號" aria-label="刪除帳號"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M4 7h16"/><path d="M9.5 7V5.5a1 1 0 0 1 1-1h3a1 1 0 0 1 1 1V7"/><path d="M6.5 7l.8 12a1.5 1.5 0 0 0 1.5 1.4h6.4a1.5 1.5 0 0 0 1.5-1.4l.8-12"/><path d="M10.5 11v6M13.5 11v6"/></svg></button>
                </form>
            </td>
        </tr>
        {% endfor %}
    </table>
</div>

<div class="detail-section">
    <h2>新增帳號</h2>
    <form method="post" action="{{ url_for('user_new') }}">
        <label>帳號</label><input type="text" name="username">
        <label>密碼(至少 8 碼)</label><input type="password" name="password">
        <label><input type="checkbox" name="is_admin" value="1"> 設為管理員(可刪除資料、CSV 匯入、管理帳號)</label>
        <input type="submit" value="建立帳號">
    </form>
</div>

<div class="detail-section">
    <h2>重設密碼</h2>
    <form method="post" action="{{ url_for('user_password') }}">
        <label>選擇帳號</label>
        <select name="user_id">
            {% for u in users %}
            <option value="{{ u['id'] }}">{{ u['username'] }}</option>
            {% endfor %}
        </select>
        <label>新密碼(至少 8 碼)</label><input type="password" name="password">
        <input type="submit" value="重設密碼">
    </form>
    <p class="note">同事忘記密碼時由管理員在此重設。系統不開放自助註冊,新同事的帳號請用上方表單建立。</p>
</div>
"""

PAGE_AUDIT = """
<h1>稽核軌跡</h1>
<p class="note">記錄商品、供應商、別名、照片的建檔維護,以及 CSV 匯入與帳號管理等操作,供事後追查。</p>
<form method="get" class="filters">
    <input type="text" name="q" placeholder="搜尋操作者或內容" value="{{ q }}">
    <input type="submit" value="搜尋">
    <a class="plain" href="{{ url_for('audit_page') }}">清除</a>
</form>
{% if rows %}
<div class="table-scroll">
    <table>
        <tr><th>時間(台灣)</th><th>操作者</th><th>動作</th><th>對象</th><th>內容</th></tr>
        {% for r in rows %}
        <tr>
            <td>{{ r['created_local'] }}</td>
            <td>{{ r['username'] }}</td>
            <td>{{ r['action'] }}</td>
            <td>{{ r['target_type'] }}{% if r['target_id'] %} #{{ r['target_id'] }}{% endif %}</td>
            <td>{{ r['detail'] }}</td>
        </tr>
        {% endfor %}
    </table>
</div>
{{ pager }}
{% else %}
<p>目前沒有稽核紀錄。</p>
{% endif %}
"""

PAGE_COUNTS = """
<h1>循環盤點</h1>
<p class="note">
    業界實務:依 ABC 分級滾動盤點(A 類每週、B 類每月、C 類每季),而非一年一次大盤。
    研究顯示未持續盤點的系統,帳面會單向偏離且不會自行修正。
</p>

{% if session.get('is_admin') %}
<div class="detail-section">
    <h2>建立盤點單</h2>
    <form method="post" action="{{ url_for('count_new') }}">
        <label>盤點單名稱</label><input type="text" name="name" placeholder="例:{{ today }} A類週盤">
        <label>盤點範圍</label>
        <select name="scope">
            <option value="all">全部商品</option>
            <option value="A">僅 A 類(高價值,建議每週)</option>
            <option value="B">僅 B 類(建議每月)</option>
            <option value="C">僅 C 類(建議每季)</option>
            <option value="category">指定分類</option>
        </select>
        <label for="cloc">儲位前綴(範圍選「指定儲位」時適用)</label>
    <input id="cloc" class="w-code" type="text" name="location" placeholder="例:CE、D-15">
    <label>分類(範圍選「指定分類」時適用)</label>
        <select name="category">
            <option value="">(不指定)</option>
            {% for c in categories %}<option value="{{ c }}">{{ c }}</option>{% endfor %}
        </select>
        <input type="submit" value="建立盤點單">
    </form>
</div>
{% endif %}

<div class="detail-section">
    <h2>盤點單列表</h2>
    {% if rows %}
    <div class="table-scroll">
        <table class="cards">
            <tr><th>名稱</th><th>範圍</th><th>建立時間</th><th>建立者</th><th>狀態</th><th class="num">準確率</th><th>操作</th></tr>
            {% for r in rows %}
            <tr>
                <td data-label="名稱">{{ r['name'] }}</td>
                <td data-label="範圍">{{ r['scope'] }}</td>
                <td data-label="建立時間">{{ r['created_local'] }}</td>
                <td data-label="建立者">{{ r['username'] }}</td>
                <td data-label="狀態">{% if r['status'] == 'posted' %}<span class="chip-posted">已過帳</span>{% else %}<span class="chip-open">盤點中</span>{% endif %}</td>
                <td data-label="準確率" class="num">{% if r['accuracy'] is not none %}{{ r['accuracy_str'] }}%{% else %}—{% endif %}</td>
                <td data-label="操作"><a class="plain" href="{{ url_for('count_detail', cid=r['id']) }}">開啟</a></td>
            </tr>
            {% endfor %}
        </table>
    </div>
    {% else %}
    <p>尚無盤點單。{% if session.get('is_admin') %}用上方表單建立第一張。{% else %}請洽管理員建立。{% endif %}</p>
    {% endif %}
</div>
"""

PAGE_COUNT_DETAIL = """
<div class="head">
    <h1>盤點單:{{ c['name'] }}</h1>
    <div class="acts">
        <span class="chip {{ 'ok' if c['status'] == 'posted' else 'act' }}">{{ '已過帳' if c['status'] == 'posted' else '進行中' }}</span>
        <span>建立於 {{ c['created_local'] }}</span>
        {% if c['status'] == 'posted' %}<span>準確率 {{ c['accuracy_str'] }}%</span>{% endif %}
    </div>
</div>
<div class="stat-row">
    <div class="stat-box"><div class="stat-cap">明細品項</div><div class="stat-num">{{ total }}</div></div>
    <div class="stat-box {{ 'good' if counted == total else 'warn' }}"><div class="stat-cap">已盤點</div><div class="stat-num">{{ counted }}</div></div>
    <div class="stat-box {{ 'bad' if diff_count else 'good' }}"><div class="stat-cap">有差異</div><div class="stat-num">{{ diff_count }}</div></div>
    <div class="stat-box"><div class="stat-cap">尚未盤點</div><div class="stat-num">{{ total - counted }}</div></div>
</div>

{% if c['status'] != 'posted' %}
<div class="btnrow">
    <form class="inline" method="post" action="{{ url_for('count_fill', cid=c['id']) }}">
        <button class="btn ghost" type="submit">無差異全部確認</button>
    </form>
    <span class="note">把還沒盤的列一次填成系統帳。帳面本來就正確時最常用。</span>
</div>
{% endif %}

<div class="pane">
    <div class="pane-h">明細<span class="r">
        <a href="{{ url_for('count_detail', cid=c['id'], sort=('loc' if sort != 'loc' else None), filter=filter) }}">{{ '依儲位排序' if sort != 'loc' else '依建檔順序' }}</a>
        <a href="{{ url_for('count_detail', cid=c['id'], sort=sort, filter=('todo' if filter != 'todo' else None)) }}">{{ '只看未盤' if filter != 'todo' else '看全部' }}</a>
        <a href="{{ url_for('count_detail', cid=c['id'], sort=sort, filter=('diff' if filter != 'diff' else None)) }}">{{ '只看有差異' if filter != 'diff' else '看全部' }}</a>
        {% if not printing %}<a href="{{ url_for('count_detail', cid=c['id'], sort=sort, filter=filter, print=1) }}">列印整張單</a>
        第 {{ page }} 頁{% else %}全 {{ items|length }} 列{% endif %}</span></div>
    {% if printing %}<p class="note no-print">這是整張單的列印版({{ items|length }} 列,不分頁)。按瀏覽器的列印(Ctrl+P)輸出,實盤數欄會印成空格供現場手寫。</p>{% endif %}
    <form method="post" action="{{ url_for('count_record', cid=c['id']) }}">
        <input type="hidden" name="page" value="{{ page }}">
        <input type="hidden" name="sort" value="{{ sort }}">
        <input type="hidden" name="filter" value="{{ filter }}">
        <div class="table-scroll">
            <table class="cards">
                <tr><th>料號</th><th>名稱</th><th>儲位</th><th class="num">系統帳</th>
                    <th class="num">實盤數</th><th class="num">差異</th><th>備註</th>
                    {% if c['status'] != 'posted' %}<th class="no-print">單列</th>{% endif %}</tr>
                {% for i in items %}
                <tr id="i{{ i['id'] }}"{% if i['diff'] is not none and i['diff'] != 0 %} class="has-diff"{% endif %}>
                    <td data-label="料號" class="mono">{{ i['sku'] }}</td>
                    <td data-label="名稱"><a class="plain" href="{{ url_for('product_detail', pid=i['product_id']) }}">{{ i['name'] }}</a></td>
                    <td data-label="儲位" class="mono">{{ i['location'] or '—' }}</td>
                    <td data-label="系統帳" class="num">{{ i['system_qty'] }}</td>
                    <td data-label="實盤數" class="num">
                        {% if c['status'] == 'posted' %}{{ i['counted_qty'] if i['counted_qty'] is not none else '—' }}
                        {% else %}
                        <label class="sr-only" for="q{{ i['id'] }}">第 {{ loop.index }} 列實盤數</label>
                        <input id="q{{ i['id'] }}" class="count-input" type="number" name="qty_{{ i['id'] }}" min="0"
                               inputmode="numeric" value="{{ i['counted_qty'] if i['counted_qty'] is not none else '' }}">
                        {% endif %}
                    </td>
                    <td data-label="差異" class="num">
                        {% if i['diff'] is none %}<span class="dim">—</span>
                        {% elif i['diff'] == 0 %}<span class="chip ok">相符</span>
                        {% else %}<span class="chip bad">{{ '+' if i['diff'] > 0 else '' }}{{ i['diff'] }}</span>{% endif %}
                    </td>
                    <td data-label="備註">
                        {% if c['status'] == 'posted' %}{{ i['note'] }}
                        {% else %}
                        <label class="sr-only" for="n{{ i['id'] }}">第 {{ loop.index }} 列備註</label>
                        <input id="n{{ i['id'] }}" class="count-note" type="text" name="note_{{ i['id'] }}"
                               placeholder="差異原因" value="{{ i['note'] }}">
                        {% endif %}
                    </td>
                    {% if c['status'] != 'posted' %}
                    <td data-label="單列" class="no-print"><button class="small-btn ok-btn" type="submit"
                        formaction="{{ url_for('count_record', cid=c['id'], only=i['id']) }}">存這列</button></td>
                    {% endif %}
                </tr>
                {% endfor %}
            </table>
        </div>
        {% if c['status'] != 'posted' %}
        <div class="savebar">
            <span class="cnt">這一頁已填 <b>{{ page_filled }}</b> / {{ items|length }} 列　·　整張單已盤 <b>{{ counted }}</b> / {{ total }} 列</span>
            <button class="btn" type="submit">儲存這一頁</button>
        </div>
        {% endif %}
    </form>
</div>
{{ pager }}

{% if c['status'] != 'posted' %}
<div class="pane">
    <div class="pane-h">過帳</div>
    <div class="pane-b">
    {% if total - counted %}
    <p class="note"><b>還有 {{ total - counted }} 列沒有盤點。</b>過帳只會調整已盤點的列,沒盤的列會被略過,而且過帳後整張單就鎖住不能再改。</p>
    {% else %}
    <p class="note">全部 {{ total }} 列都已盤點。過帳會依差異產生「盤點調整」異動並同步批次帳,之後這張單就鎖住不能再改。</p>
    {% endif %}
    {% if session.get('is_admin') %}
    <form method="post" action="{{ url_for('count_post', cid=c['id']) }}"
          onsubmit="return confirm('確定過帳?將依差異調整庫存,且不可復原。');">
        <button class="btn" type="submit">過帳並調整庫存</button>
    </form>
    {% else %}
    <p class="note">過帳需要管理員權限。</p>
    {% endif %}
    </div>
</div>
{% endif %}
"""


PO_STATUS_LABEL = {"ordered": "已下訂", "shipped": "已出貨", "arrived": "已到貨待驗",
                   "closed": "已入庫結案", "cancelled": "已作廢"}

PAGE_ORDER_NEW = """
<h1>建立採購單</h1>
<p class="note">這一版一張單一個品項。要一次開多個品項,仍請用<a class="plain" href="{{ url_for('orders_page') }}">上傳明細檔</a>。</p>
{% if prod %}
<div class="pickcard">
    <div>
        <div class="nm">{{ prod['name'] }}</div>
        <div class="meta">{{ prod['sku'] }}　·　儲位 {{ prod['location'] or '未設定' }}　·　{{ prod['supplier_name'] or '無供應商' }}</div>
    </div>
    <div class="right">
        <div class="big">{{ prod['available'] }}</div>
        <div class="meta">目前可用({{ prod['unit'] }}){% if prod['onorder'] %}　·　已在途 {{ prod['onorder'] }}{% endif %}　·　<a href="{{ url_for('pick', to='order_new') }}">換一項料</a></div>
    </div>
</div>
<form method="post">
    <input type="hidden" name="product_id" value="{{ prod['id'] }}">
    <label for="oq">訂購數量 <span class="req">必填</span></label>
    <input id="oq" class="w-qty" type="number" name="ordered_qty" min="1" inputmode="numeric" value="{{ qty }}">
    <label for="pn">採購單號 <span class="opt">選填</span></label>
    <p class="field-help">不填會自動編號。</p>
    <input id="pn" class="w-code" type="text" name="po_no" placeholder="例:PO-20260903-01">
    <label for="sp">供應商 <span class="opt">選填</span></label>
    <input id="sp" class="w-code" type="text" name="supplier_name" value="{{ prod['supplier_name'] or '' }}">
    <label for="et">預計到貨日 <span class="opt">選填</span></label>
    <input id="et" class="w-code" type="date" name="eta" value="{{ eta }}">
    <label for="uc">單價 <span class="opt">選填</span></label>
    <input id="uc" class="w-qty" type="text" name="unit_cost" inputmode="decimal">
    <label for="nt">備註 <span class="opt">選填</span></label>
    <input id="nt" type="text" name="note">
    <input type="submit" value="建立採購單">
    <p class="note" style="margin-top:12px">建立後這一項會立刻算進「在途」,首頁與商品明細都看得到,避免有人重複下單。</p>
</form>
{% else %}
<div class="pane"><div class="pane-b"><div class="empty">
    <div class="h">先找到要採購的料</div>
    <div class="p">用料號、品名或儲位找;找到之後數量與供應商會自動帶入。</div>
    <div class="btnrow" style="justify-content:center">
        <a class="btn" href="{{ url_for('pick', to='order_new') }}">先找料</a>
        <a class="btn ghost" href="{{ url_for('alerts') }}">從短缺佇列挑</a>
    </div>
</div></div></div>
{% endif %}
"""

PAGE_ORDERS = """
<h1>採購訂單</h1>
<p class="note">
    訂單只登記一次,後面每個階段自動往下帶:<strong>下訂 → 出貨 → 到貨 → 檢查 → 入庫</strong>。
    標記到貨時系統直接把明細變成收貨單,現場只要核對數量,不必重打料號。
    下訂到入庫之間的量叫「在途」,它讓你在下單前先看到「其實已經在路上了」。
</p>

<div class="po-flow">
    {% for st in ['ordered','shipped','arrived'] %}
    <a class="po-step po-{{ st }}" href="{{ url_for('orders_page', status=st) }}">
        <span class="po-step-label">{{ labels[st] }}</span>
        <span class="po-step-n">{{ summary[st]['orders'] }}<small> 張單</small></span>
        <span class="po-step-sub">{{ summary[st]['items'] }} 品項・在途 {{ summary[st]['qty'] }}</span>
    </a>
    {% endfor %}
</div>

<div class="detail-section">
    <h2>上傳採購明細建立訂單</h2>
    <div class="import-help">
        <p>欄位順序(第一列為標題列,會被略過):<br>
        <code>料號,品名,數量,單價,預計到貨日,備註</code><br>
        料號可以是<strong>我方料號,也可以是供應商自己的料號</strong>——系統會透過跨公司料號對照自動找到我方商品。</p>
        <p><a class="plain" href="{{ url_for('order_template') }}">下載空白範例檔</a></p>
    </div>
    <form method="post" action="{{ url_for('order_upload') }}" enctype="multipart/form-data">
        <label>採購單號(選填)</label>
        <input type="text" name="po_no" placeholder="例:PO-20260808-01">
        <label>供應商</label>
        <select name="supplier_id">
            <option value="">(未指定)</option>
            {% for s in suppliers %}<option value="{{ s['id'] }}">{{ s['name'] }}</option>{% endfor %}
        </select>
        <label>預計到貨日(選填)</label><input type="date" name="eta">
        <label>明細檔案</label>
        <input type="file" name="file" accept=".csv,.xlsx,.xlsm,.tsv,.txt,text/csv">
        <label>備註(選填)</label><input type="text" name="note">
        <input type="submit" value="建立採購單">
    </form>
</div>

<div class="detail-section">
    <h2>訂單列表{% if status_filter %}:{{ labels.get(status_filter, status_filter) }}{% endif %}</h2>
    <p class="filters">
        <a class="chip-link{% if not status_filter %} on{% endif %}" href="{{ url_for('orders_page') }}">全部</a>
        {% for st, lb in labels.items() %}
        <a class="chip-link{% if status_filter == st %} on{% endif %}" href="{{ url_for('orders_page', status=st) }}">{{ lb }}</a>
        {% endfor %}
    </p>
    {% if rows %}
    <div class="table-scroll">
    <table class="cards">
        <tr><th>單號</th><th>供應商</th><th>建立</th><th>預計到貨</th>
            <th class="num">品項</th><th class="num">訂購</th><th class="num">已收</th><th>狀態</th><th>操作</th></tr>
        {% for r in rows %}
        <tr>
            <td data-label="單號">{{ r['po_no'] or '(未填單號)' }}</td>
            <td data-label="供應商">{{ r['supplier_name'] or '—' }}</td>
            <td data-label="建立">{{ r['created_local'] }}</td>
            <td data-label="預計到貨">{{ r['eta'] or '—' }}</td>
            <td data-label="品項" class="num">{{ r['item_count'] }}</td>
            <td data-label="訂購" class="num">{{ r['ordered_total'] }}</td>
            <td data-label="已收" class="num">{{ r['received_total'] }}</td>
            <td data-label="狀態"><span class="chip-po chip-{{ r['status'] }}">{{ labels[r['status']] }}</span></td>
            <td data-label="操作"><a class="plain" href="{{ url_for('order_detail', oid=r['id']) }}">開啟</a></td>
        </tr>
        {% endfor %}
    </table>
    </div>
    {% else %}
    <p>沒有符合條件的採購單。用上方表單建立第一張。</p>
    {% endif %}
</div>
"""

PAGE_ORDER_DETAIL = """
<h1>採購單:{{ o['po_no'] or '(未填單號)' }}</h1>
<p class="note">
    供應商 {{ o['supplier_name'] or '未指定' }}・建立於 {{ o['created_local'] }}({{ o['username'] }})
    {% if o['eta'] %}・預計到貨 {{ o['eta'] }}{% endif %}
    {% if o['shipped_at'] %}<br>已於 {{ o['shipped_at'] }} 出貨{% if o['tracking_no'] %},追蹤號 {{ o['tracking_no'] }}{% endif %}{% endif %}
    {% if o['arrived_at'] %}<br>已於 {{ o['arrived_local'] }} 登記到貨{% endif %}
    {% if o['note'] %}<br>備註:{{ o['note'] }}{% endif %}
</p>

<div class="po-track">
    {% for st in ['ordered','shipped','arrived','closed'] %}
    <div class="po-node{% if step_index >= loop.index0 %} done{% endif %}{% if step_index == loop.index0 %} now{% endif %}">
        <span class="po-dot"></span>
        <span class="po-node-label">{{ labels[st] }}</span>
    </div>
    {% endfor %}
</div>

<div class="stat-row">
    <div class="stat-box"><div class="stat-num">{{ total }}</div><div class="stat-cap">明細品項</div></div>
    <div class="stat-box"><div class="stat-num">{{ ordered_total }}</div><div class="stat-cap">訂購總量</div></div>
    <div class="stat-box"><div class="stat-num">{{ received_total }}</div><div class="stat-cap">已入庫</div></div>
    <div class="stat-box"><div class="stat-num"{% if onorder_total %} style="color:#b45309"{% endif %}>{{ onorder_total }}</div><div class="stat-cap">在途未到</div></div>
</div>

<div class="table-scroll">
<table class="cards">
    <tr><th>行</th><th>檔案料號</th><th>對應商品</th><th class="num">訂購</th>
        <th class="num">已收</th><th class="num">在途</th><th class="num">單價</th><th>備註</th></tr>
    {% for i in items %}
    <tr{% if i['product_id'] is none %} class="has-diff"{% endif %}>
        <td data-label="行">{{ i['line_no'] }}</td>
        <td data-label="檔案料號">
            <span class="loc-cell">{{ i['raw_sku'] or '—' }}</span>
            {% if i['raw_name'] %}<br><span class="alias-cell">{{ i['raw_name'] }}</span>{% endif %}
        </td>
        <td data-label="對應商品">
            {% if i['product_id'] %}
                <a class="plain" href="{{ url_for('product_detail', pid=i['product_id']) }}">{{ i['pname'] }}</a>
                <br><span class="alias-cell">{{ i['psku'] }}・{{ i['match_label'] }}</span>
            {% elif o['status'] in ('ordered','shipped','arrived') %}
                <span class="badge-low">未對應</span>
                <a class="small-btn ok-btn" href="{{ url_for('pick', to='order:' ~ o['id'] ~ ':' ~ i['id'], q=i['raw_sku']) }}">指定我方商品</a>
            {% else %}<span class="badge-low">未對應</span>{% endif %}
        </td>
        <td data-label="訂購" class="num">{{ i['ordered_qty'] }}</td>
        <td data-label="已收" class="num">{{ i['received_qty'] }}</td>
        <td data-label="在途" class="num">{{ i['onorder'] if i['onorder'] else '—' }}</td>
        <td data-label="單價" class="num">{{ i['unit_cost'] if i['unit_cost'] is not none else '—' }}</td>
        <td data-label="備註">{{ i['note'] }}</td>
    </tr>
    {% endfor %}
</table>
</div>

{% if receipts %}
<div class="detail-section">
    <h2>由本訂單產生的收貨單</h2>
    <div class="table-scroll">
        <table class="cards">
            <tr><th>收貨單號</th><th>建立時間</th><th>狀態</th><th>操作</th></tr>
            {% for r in receipts %}
            <tr>
                <td data-label="收貨單號">{{ r['ref_no'] or '(未填單號)' }}</td>
                <td data-label="建立時間">{{ r['created_local'] }}</td>
                <td data-label="狀態">
                    {% if r['status'] == 'posted' %}<span class="chip-posted">已放行</span>
                    {% elif r['status'] == 'cancelled' %}<span class="chip-void">已作廢</span>
                    {% else %}<span class="chip-open">待核對</span>{% endif %}
                </td>
                <td data-label="操作"><a class="plain" href="{{ url_for('receipt_detail', rid=r['id']) }}">開啟</a></td>
            </tr>
            {% endfor %}
        </table>
    </div>
</div>
{% endif %}

{% if o['status'] in ('ordered','shipped') %}
<div class="detail-section">
    <h2>推進狀態</h2>
    {% if o['status'] == 'ordered' %}
    <p class="note">供應商通知出貨後,在這裡登記出貨日與追蹤號。若貨已直接送到,也可以跳過這步直接標記到貨。</p>
    <form method="post" action="{{ url_for('order_ship', oid=o['id']) }}">
        <label>出貨日</label><input type="date" name="shipped_at" value="{{ today }}">
        <label>追蹤號 / 提單號(選填)</label><input type="text" name="tracking_no">
        <input type="submit" value="標記為已出貨">
    </form>
    {% endif %}
    <form method="post" action="{{ url_for('order_arrive', oid=o['id']) }}"
          onsubmit="return confirm('確定登記到貨?系統會自動建立收貨單,現場核對數量後才會入庫。');">
        <p class="note">
            貨到了按這裡。系統會<strong>自動把這張訂單的明細變成收貨單</strong>,不必重新上傳或重打;
            庫存此時仍不變,要等收貨單核對放行才入庫。
        </p>
        <input type="submit" value="登記到貨並自動建立收貨單">
    </form>
</div>
{% endif %}

{% if o['status'] in ('ordered','shipped','arrived') and session.get('is_admin') %}
<div class="detail-section">
    <h2>作廢</h2>
    <p class="note">訂單取消或重下時使用。作廢後不再計入在途量,也不影響庫存。</p>
    <form method="post" action="{{ url_for('order_cancel', oid=o['id']) }}"
          onsubmit="return confirm('確定作廢這張採購單?');">
        <button class="small-btn" type="submit">作廢此採購單</button>
    </form>
</div>
{% endif %}
"""

PAGE_RECEIPTS = """
<h1>收貨單(到貨預先登記)</h1>
<p class="note">
    業界作法叫 ASN(預先到貨通知):供應商的送貨明細先進系統成為待核對單據,
    現場只需核對數量、不必重打料號與品名。<strong>放行之前庫存完全不變</strong>,
    確認無誤按下放行,才產生正式入庫與批次。
</p>

<div class="detail-section">
    <h2>上傳到貨明細(Excel / CSV)</h2>
    <div class="import-help">
        <p>欄位順序(第一列為標題列,會被略過):<br>
        <code>料號,品名,數量,批號,效期,單價,備註</code><br>
        料號可以是<strong>我方料號,也可以是供應商自己的料號</strong>——系統會透過跨公司料號對照自動找到我方商品。
        對不上的列會標示「未對應」,可在明細頁手動指定,並選擇記住這個對應,下次就自動對上。</p>
        <p><a class="plain" href="{{ url_for('receipt_template') }}">下載空白範例檔</a>(可直接轉給供應商填)</p>
    </div>
    <form method="post" action="{{ url_for('receipt_upload') }}" enctype="multipart/form-data">
        <label>送貨單號 / 通知單號(選填)</label>
        <input type="text" name="ref_no" placeholder="例:DN-20260807-01">
        <label>供應商</label>
        <select name="supplier_id">
            <option value="">(未指定)</option>
            {% for s in suppliers %}<option value="{{ s['id'] }}">{{ s['name'] }}</option>{% endfor %}
        </select>
        <label>明細檔案</label>
        <input type="file" name="file" accept=".csv,.xlsx,.xlsm,.tsv,.txt,text/csv">
        <label>備註(選填)</label><input type="text" name="note">
        <input type="submit" value="建立收貨單">
    </form>
</div>

<div class="detail-section">
    <h2>收貨單列表</h2>
    {% if rows %}
    <div class="table-scroll">
        <table class="cards">
            <tr><th>單號</th><th>供應商</th><th>建立時間</th><th>建立者</th>
                <th class="num">品項</th><th class="num">已核對</th><th>狀態</th><th>操作</th></tr>
            {% for r in rows %}
            <tr>
                <td data-label="單號">{{ r['ref_no'] or '(未填單號)' }}</td>
                <td data-label="供應商">{{ r['supplier_name'] or '—' }}</td>
                <td data-label="建立時間">{{ r['created_local'] }}</td>
                <td data-label="建立者">{{ r['username'] }}</td>
                <td data-label="品項" class="num">{{ r['item_count'] }}</td>
                <td data-label="已核對" class="num">{{ r['checked_count'] }}</td>
                <td data-label="狀態">
                    {% if r['status'] == 'posted' %}<span class="chip-posted">已放行</span>
                    {% elif r['status'] == 'cancelled' %}<span class="chip-void">已作廢</span>
                    {% else %}<span class="chip-open">待核對</span>{% endif %}
                </td>
                <td data-label="操作"><a class="plain" href="{{ url_for('receipt_detail', rid=r['id']) }}">開啟</a></td>
            </tr>
            {% endfor %}
        </table>
    </div>
    {% else %}
    <p>尚無收貨單。用上方表單上傳供應商的送貨明細建立第一張。</p>
    {% endif %}
</div>
"""

PAGE_RECEIPT_DETAIL = """
<h1>收貨單:{{ r['ref_no'] or '(未填單號)' }}</h1>
<p class="note">
    供應商 {{ r['supplier_name'] or '未指定' }}・來源 {{ r['source'] }}・
    建立於 {{ r['created_local'] }}({{ r['username'] }})・
    {% if r['status'] == 'posted' %}<strong>已於 {{ r['posted_local'] }} 放行入庫</strong>
    {% elif r['status'] == 'cancelled' %}<strong>已作廢</strong>
    {% else %}狀態:待核對(核對實收數量後按下方「放行」才會入庫){% endif %}
    {% if r['note'] %}<br>備註:{{ r['note'] }}{% endif %}
</p>

<div class="stat-row">
    <div class="stat-box"><div class="stat-num">{{ total }}</div><div class="stat-cap">明細品項</div></div>
    <div class="stat-box"><div class="stat-num">{{ checked }}</div><div class="stat-cap">已核對</div></div>
    <div class="stat-box"><div class="stat-num"{% if unmatched %} style="color:#b91c1c"{% endif %}>{{ unmatched }}</div><div class="stat-cap">未對應料號</div></div>
    <div class="stat-box"><div class="stat-num">{{ total_qty }}</div><div class="stat-cap">預計收料總量</div></div>
</div>

{% if r['status'] == 'open' %}
<div class="btnrow">
    <form class="inline" method="post" action="{{ url_for('receipt_fill', rid=r['id']) }}">
        <button class="btn ghost" type="submit">全部照通知量核對</button>
    </form>
    <span class="note">照單全收是最常見的情況;按一次就把所有已對應列的實收數填成通知量。</span>
</div>
{% endif %}
<div class="pane-h">明細<span class="r">
    {% if printing %}全 {{ total }} 列(列印版){% else %}<a href="{{ url_for('receipt_detail', rid=r['id'], print=1) }}">列印這張單</a>{% endif %}
</span></div>
{% if printing %}<p class="note no-print">這是列印版({{ total }} 列)。按瀏覽器的列印(Ctrl+P)輸出,實收數欄會印成空格供現場手寫核對。</p>{% endif %}
<form method="post" action="{{ url_for('receipt_check_all', rid=r['id']) }}">
<div class="table-scroll">
    <table class="cards">
        <tr><th>行</th><th>檔案料號</th><th>對應商品</th><th class="num">通知量</th>
            <th class="num">實收數</th><th>備註</th><th>批號 / 效期</th>
            {% if r['status'] == 'open' %}<th class="no-print">單列</th>{% endif %}</tr>
        {% for i in items %}
        <tr id="i{{ i['id'] }}"{% if i['product_id'] is none %} class="has-diff"{% endif %}>
            <td data-label="行">{{ i['line_no'] }}</td>
            <td data-label="檔案料號">
                <span class="loc-cell">{{ i['raw_sku'] or '—' }}</span>
                {% if i['raw_name'] %}<br><span class="alias-cell">{{ i['raw_name'] }}</span>{% endif %}
            </td>
            <td data-label="對應商品">
                {% if i['product_id'] %}
                    <a class="plain" href="{{ url_for('product_detail', pid=i['product_id']) }}">{{ i['pname'] }}</a>
                    <br><span class="alias-cell">{{ i['psku'] }}・{{ i['match_label'] }}</span>
                {% elif r['status'] == 'open' %}
                    <span class="badge-low">未對應</span>
                    <a class="small-btn ok-btn" href="{{ url_for('pick', to='receipt:' ~ r['id'] ~ ':' ~ i['id'], q=i['raw_sku']) }}">指定我方商品</a>
                {% else %}
                    <span class="badge-low">未對應</span>
                {% endif %}
            </td>
            <td data-label="通知量" class="num">{{ i['expected_qty'] }}</td>
            <td data-label="實收數" class="num">
                {% if r['status'] != 'open' %}{{ i['received_qty'] if i['received_qty'] is not none else '—' }}
                {% else %}
                <label class="sr-only" for="q{{ i['id'] }}">第 {{ i['line_no'] }} 列實收數</label>
                <input id="q{{ i['id'] }}" class="count-input" type="number" name="qty_{{ i['id'] }}" min="0"
                       inputmode="numeric" value="{{ i['received_qty'] if i['received_qty'] is not none else '' }}">
                {% endif %}
            </td>
            <td data-label="備註">
                {% if r['status'] != 'open' %}{{ i['note'] }}
                {% else %}
                <label class="sr-only" for="n{{ i['id'] }}">第 {{ i['line_no'] }} 列備註</label>
                <input id="n{{ i['id'] }}" class="count-note" type="text" name="note_{{ i['id'] }}"
                       placeholder="備註" value="{{ i['note'] }}">
                {% endif %}
            </td>
            <td data-label="批號 / 效期">
                {{ i['lot_no'] or '(自動編號)' }}{% if i['expiry_date'] %}<br><span class="alias-cell">效期 {{ i['expiry_date'] }}</span>{% endif %}
            </td>
            {% if r['status'] == 'open' %}
            <td data-label="單列" class="no-print"><button class="small-btn ok-btn" type="submit"
                formaction="{{ url_for('receipt_check_all', rid=r['id'], only=i['id']) }}">存這列</button></td>
            {% endif %}
        </tr>
        {% endfor %}
    </table>
</div>
{% if r['status'] == 'open' %}
<div class="savebar">
    <span class="cnt">已填 <b>{{ checked }}</b> / {{ total }} 列{% if unmatched %}　·　<b>{{ unmatched }}</b> 列還對不到料號{% endif %}</span>
    <button class="btn" type="submit">儲存這一頁</button>
</div>
{% endif %}
</form>

{% if r['status'] == 'open' %}
<div class="detail-section">
    <h2>放行入庫</h2>
    {% if skipped %}
    <p class="msg error" style="border-left-width:4px"><b>第 {{ skipped }} 列尚未核對或對不到料號,放行後會被略過,而且不能再修改。</b>
    貨進了倉庫卻沒進帳,是庫存失準最常見的起點——請先把那幾列處理完再放行。</p>
    {% endif %}
    <p class="note">
        放行會把每一列「已對應且實收 &gt; 0」的品項寫成正式入庫異動並建立批次,庫存這時才會增加。
        實收填 0 或未核對的列會被略過。<strong>尚有未對應料號且實收 &gt; 0 時無法放行</strong>——
        有料進來卻沒有帳,是庫存失準最常見的起點。放行後不可再修改。
    </p>
    <form method="post" action="{{ url_for('receipt_post', rid=r['id']) }}"
          onsubmit="return confirm('確定放行?將依實收數量正式入庫,且不可復原。');">
        <input type="submit" value="放行並入庫">
    </form>
</div>
{% if session.get('is_admin') %}
<div class="detail-section">
    <h2>作廢</h2>
    <p class="note">送錯檔案或整批退回時使用。作廢不會影響庫存(本來就還沒入庫)。</p>
    <form method="post" action="{{ url_for('receipt_cancel', rid=r['id']) }}"
          onsubmit="return confirm('確定作廢這張收貨單?');">
        <button class="small-btn" type="submit">作廢此收貨單</button>
    </form>
</div>
{% endif %}
{% endif %}
"""

PAGE_RESERVATIONS = """
<h1>庫存預留</h1>
<p class="note">
    預留讓「已被承諾出去」的數量從可用量中扣除,避免兩個人同時把同一批料承諾給不同用途。
    現貨量不受影響——實際領走時才走出庫。
</p>

<div class="detail-section">
    <h2>建立預留</h2>
    {% if picked_product %}
    <form method="post" action="{{ url_for('reservation_new') }}">
        <input type="hidden" name="product_id" value="{{ picked_product['id'] }}">
        <div class="pickcard">
            <div><div class="nm">{{ picked_product['name'] }}</div>
                 <div class="meta">{{ picked_product['sku'] }}　·　儲位 {{ picked_product['location'] or '未設定' }}</div></div>
            <div class="right"><div class="big">{{ picked_product['available'] }}</div>
                 <div class="meta">目前可用({{ picked_product['unit'] }})　·　<a href="{{ url_for('pick', to='reserve') }}">換一項料</a></div></div>
        </div>
        <label>預留數量</label><input type="number" name="quantity" min="1" inputmode="numeric">
        <label>用途(工單／專案／客戶)</label><input type="text" name="purpose" placeholder="例:WO-1001">
        <input type="submit" value="建立預留">
    </form>
    {% else %}
    <p class="note">預留是把在庫的量綁給某張工單,綁住的量就不能被別人領走。</p>
    <p><a class="btn" href="{{ url_for('pick', to='reserve') }}">先找要預留的料</a></p>
    {% endif %}
</div>

<div class="detail-section">
    <h2>有效預留</h2>
    {% if rows %}
    <div class="table-scroll">
        <table class="cards">
            <tr><th>商品</th><th>SKU</th><th class="num">預留量</th><th>用途</th><th>建立者</th><th>建立時間</th><th>操作</th></tr>
            {% for r in rows %}
            <tr>
                <td data-label="商品">{{ r['name'] }}</td>
                <td data-label="SKU">{{ r['sku'] }}</td>
                <td data-label="預留量" class="num">{{ r['quantity'] }}</td>
                <td data-label="用途">{{ r['purpose'] or '—' }}</td>
                <td data-label="建立者">{{ r['username'] }}</td>
                <td data-label="建立時間">{{ r['created_local'] }}</td>
                <td data-label="操作">
                    <form class="inline" method="post" action="{{ url_for('reservation_release', rid=r['id']) }}">
                        <button class="small-btn ok-btn" type="submit">釋放</button>
                    </form>
                </td>
            </tr>
            {% endfor %}
        </table>
    </div>
    {% else %}
    <p>目前沒有有效的預留。</p>
    {% endif %}
</div>
"""

PAGE_PLANNING = """
<h1>存貨規劃:安全庫存建議</h1>
<div class="import-help">
    <p>
        低庫存門檻不該憑印象填。系統以最近 {{ window }} 天的實際出庫資料算出<strong>日均用量</strong>與
        <strong>標準差</strong>,套安全庫存公式推導建議值:
    </p>
    <p><code>安全庫存 = Z × 用量標準差 × √前置期</code>,<code>再訂購點 = 日均用量 × 前置期 + 安全庫存</code></p>
    <p>
        Z 由該商品的<strong>目標服務水準</strong>決定(95% → Z=1.65),前置期與服務水準可在商品編輯頁調整。
        沒有出庫歷史的商品無法推導,顯示「—」。
    </p>
</div>

<div class="table-scroll">
    <table class="cards">
        <tr>
            <th>SKU</th><th>名稱</th><th class="num">日均用量</th><th class="num">標準差</th>
            <th class="num">前置期</th><th class="num">服務水準</th>
            <th class="num">建議安全庫存</th><th class="num">再訂購點</th><th class="num">目前門檻</th><th>XYZ</th>
        </tr>
        {% for r in rows %}
        <tr>
            <td data-label="SKU">{{ r['sku'] }}</td>
            <td data-label="名稱"><a class="plain" href="{{ url_for('product_detail', pid=r['id']) }}">{{ r['name'] }}</a></td>
            <td data-label="日均用量" class="num">{{ r['mean_str'] }}</td>
            <td data-label="標準差" class="num">{{ r['sd_str'] }}</td>
            <td data-label="前置期" class="num">{{ r['lead_time_days'] }} 天</td>
            <td data-label="服務水準" class="num">{{ r['service_level_str'] }}%</td>
            <td data-label="建議安全庫存" class="num"><strong>{{ r['ss_str'] }}</strong></td>
            <td data-label="再訂購點" class="num">{{ r['rop_str'] }}</td>
            <td data-label="目前門檻" class="num">{{ r['low_stock_threshold'] }}</td>
            <td data-label="XYZ">{{ r['xyz'] }}</td>
        </tr>
        {% endfor %}
    </table>
</div>

{% if has_suggestion and session.get('is_admin') %}
<div class="detail-section">
    <h2>套用建議</h2>
    <p class="note">將所有可推導商品的低庫存門檻,一次更新為上表的建議安全庫存。個別商品仍可在編輯頁手動調整。</p>
    <form method="post" action="{{ url_for('planning_apply') }}"
          onsubmit="return confirm('確定將建議值套用到所有可推導的商品門檻?');">
        <input type="submit" value="一鍵套用建議門檻">
    </form>
</div>
{% endif %}
"""

PAGE_LABELS = """
<h1>料架標籤列印</h1>
<p class="note no-print">
    列印後貼在料架上,現場用手機掃 QR 即可直接開啟該料號頁面登記進出,不必打字找料號。
    按瀏覽器的列印功能(Ctrl+P / 手機分享選單)即可輸出。
</p>
{% if not has_qrcode %}
<div class="msg error">此功能需要安裝 qrcode 套件(pip install qrcode)後重新啟動系統。</div>
{% else %}
<form class="filters no-print" method="get">
    <label for="lb-loc">儲位開頭</label>
    <input id="lb-loc" type="text" name="loc" value="{{ loc }}" placeholder="例如 A- 或 防潮箱" size="12">
    <label for="lb-cat">分類</label>
    <select id="lb-cat" name="category">
        <option value="">全部分類</option>
        {% for c in cats %}<option value="{{ c }}"{% if c == category %} selected{% endif %}>{{ c }}</option>{% endfor %}
    </select>
    <label for="lb-sku">單一料號</label>
    <input id="lb-sku" type="text" name="sku" value="{{ sku }}" placeholder="補印一張時用" size="12">
    <input type="submit" value="篩選">
    {% if loc or category or sku %}<a class="plain" href="{{ url_for('labels_page') }}">清除</a>{% endif %}
</form>
<p class="note no-print">
    快速選一排貨架:
    {% for pre in prefixes %}<a class="plain" href="{{ url_for('labels_page', loc=pre) }}">{{ pre }}</a>{% if not loop.last %}　{% endif %}{% endfor %}
</p>
<div class="msg ok no-print">
    這次會印 <b>{{ count }}</b> 張標籤,約 <b>{{ sheets }}</b> 頁 A4。
    QR 內容指向 <span class="mono">{{ qr_base }}</span> —— 這串位址要是手機在公司網路上打得開的,標籤才掃得動。
</div>
{% if not products %}
<p>目前的篩選條件找不到商品。</p>
{% endif %}
<div class="label-sheet">
    {% for p in products %}
    <div class="label">
        <img src="{{ url_for('product_qr', pid=p['id']) }}" alt="QR">
        <div class="label-text">
            <div class="label-sku">{{ p['sku'] }}</div>
            <div class="label-name">{{ p['name'] }}</div>
            <div class="label-loc">儲位:{{ p['location'] or '未設定' }}</div>
        </div>
    </div>
    {% endfor %}
</div>
{% endif %}
"""

PAGE_IMAGE_SEARCH = """
<h1>以圖搜圖</h1>
<p>上傳一張物料照片,系統會比對庫內所有照片,列出外觀最相似的料號。</p>
{% if not has_pil %}
<div class="msg error">此功能需要安裝 Pillow 套件(pip install Pillow)後重新啟動系統。</div>
{% else %}
<form method="post" enctype="multipart/form-data">
    <input type="file" name="photo" accept="image/*">
    <input type="submit" value="搜尋相似物料">
</form>
{% endif %}
{% if results is not none %}
<div class="detail-section">
    <h2>搜尋結果(相似度由高至低)</h2>
    {% if results %}
    <div class="table-scroll">
        <table>
            <tr><th>照片</th><th>SKU</th><th>名稱</th><th>相似度</th><th>目前庫存</th></tr>
            {% for r in results %}
            <tr>
                <td><img class="thumb" src="{{ url_for('serve_image', filename=r['filename']) }}" alt=""></td>
                <td>{{ r['sku'] }}</td>
                <td><a class="plain" href="{{ url_for('product_detail', pid=r['product_id']) }}">{{ r['name'] }}</a></td>
                <td>{{ r['similarity'] }}%</td>
                <td>{{ r['quantity'] }} {{ r['unit'] }}</td>
            </tr>
            {% endfor %}
        </table>
    </div>
    {% else %}
    <p>庫內尚無照片可比對,請先到商品詳細頁上傳物料照片。</p>
    {% endif %}
</div>
{% endif %}
"""

PAGE_IMPORT = """
<h1>大量匯入(Excel / CSV)</h1>
<div class="import-help">
    <p><strong>商品匯入</strong>欄位順序(第一列為標題列,會被略過):<br>
    <code>SKU,名稱,分類,單位,單價,低庫存門檻,供應商,初始庫存,儲位</code><br>
    SKU 與名稱必填,其餘可留空(最後的<strong>儲位為選填</strong>,舊系統搬遷時可一次帶入);
    供應商不存在時自動建立;初始庫存會寫成一筆入庫異動(備註「CSV匯入」)。</p>
    <p><strong>別名匯入</strong>欄位順序:<br>
    <code>我方SKU,公司,別名料號,備註</code></p>
    <p><strong>支援格式</strong>:Excel(<code>.xlsx</code>)、CSV、Tab 分隔(<code>.tsv</code>/<code>.txt</code>)。
    CSV 編碼支援 UTF-8 與 Big5(台灣 Excel 另存 CSV 的預設編碼),系統自動判斷。
    {% if not has_openpyxl %}<br><strong>注意:</strong>此主機未安裝 openpyxl,目前只能讀 CSV;
    執行 <code>pip install -r requirements.txt</code> 後重新啟動即可支援 Excel。{% endif %}</p>
    <p class="note">要登記「到貨收料」請改用<a class="plain" href="{{ url_for('receipts_page') }}">收貨單</a>——
    那裡是先預先登記、核對無誤才放行入庫,不會直接動到庫存。</p>
</div>
<form method="post" enctype="multipart/form-data">
    <label>匯入類型</label>
    <select name="mode">
        <option value="products">商品匯入</option>
        <option value="aliases">別名匯入</option>
    </select>
    <label>檔案(Excel 或 CSV)</label>
    <input type="file" name="csv_file" accept=".csv,.xlsx,.xlsm,.tsv,.txt,text/csv">
    <input type="submit" value="開始匯入">
</form>
{% if report_rows is not none %}
<div class="detail-section">
    <h2>匯入結果:成功匯入 {{ ok_count }} 筆,跳過 {{ skip_count }} 筆</h2>
    <div class="table-scroll">
        <table>
            <tr><th>行號</th><th>內容</th><th>結果</th></tr>
            {% for r in report_rows %}
            <tr><td>{{ r['line'] }}</td><td>{{ r['label'] }}</td><td>{{ r['status'] }}</td></tr>
            {% endfor %}
        </table>
    </div>
</div>
{% endif %}
"""


# ---------------------------------------------------------------------------
# 帳號:註冊(僅首位)/ 登入 / 登出 / 帳號管理
# ---------------------------------------------------------------------------

MIN_PASSWORD_LEN = 8
LOGIN_MAX_FAILS = 5
LOGIN_LOCK_SEC = 300
_login_fails = {}  # username -> (失敗次數, 最後一次失敗時間);程序內記憶,重啟即清空


def check_password_policy(password):
    if len(password) < MIN_PASSWORD_LEN:
        return f"密碼至少 {MIN_PASSWORD_LEN} 碼,請重新設定"
    return None


def user_count():
    return get_db().execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"]


def create_user(username, password, is_admin):
    # 共用建帳邏輯(首位註冊與管理員新增帳號都走這裡)
    get_db().execute(
        "INSERT INTO users (username, password_hash, is_admin, created_at) VALUES (?, ?, ?, ?)",
        (username, generate_password_hash(password), 1 if is_admin else 0, now_str()),
    )


@app.route("/register", methods=["GET", "POST"])
def register():
    # 只開放給「系統第一位使用者」建立管理員帳號;之後帳號一律由管理員在 /users 新增,
    # 避免任何連得到內網的人自行建帳進入系統。
    error = None
    username = ""
    if user_count() > 0:
        return render_page(PAGE_REGISTER, closed=True,
                           error="系統已完成初始設定,不開放自助註冊。需要帳號請洽管理員建立。", page_title="建立管理員帳號")
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if not username or not password:
            error = "帳號與密碼不可為空"
        else:
            error = check_password_policy(password)
            if not error:
                db = get_db()
                try:
                    create_user(username, password, is_admin=True)  # 首位即管理員
                    db.commit()
                    return redirect(url_for("login"))
                except sqlite3.IntegrityError:
                    error = "帳號已存在,請改用其他名稱"
    return render_page(PAGE_REGISTER, error=error, username=username, closed=False, page_title="建立管理員帳號")


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    username = ""
    # 首次啟動(一個帳號都還沒有)直接帶到建立管理員頁:
    # 全新安裝的人卡在登入畫面卻沒有帳號可填,是實機裝機時最常見的卡關點。
    if request.method == "GET" and user_count() == 0:
        return redirect(url_for("register"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        fails, last = _login_fails.get(username, (0, 0.0))
        if fails >= LOGIN_MAX_FAILS and (time.time() - last) < LOGIN_LOCK_SEC:
            remain = int((LOGIN_LOCK_SEC - (time.time() - last)) / 60) + 1
            error = f"嘗試次數過多,請於約 {remain} 分鐘後再試"
        else:
            if fails >= LOGIN_MAX_FAILS:
                _login_fails.pop(username, None)  # 鎖定期已過,重新計數
            user = get_db().execute(
                "SELECT * FROM users WHERE username = ?", (username,)
            ).fetchone()
            if user and check_password_hash(user["password_hash"], password):
                _login_fails.pop(username, None)
                session.clear()
                session["user_id"] = user["id"]
                session["username"] = user["username"]
                session["is_admin"] = bool(user["is_admin"])  # 權限判斷依據
                return redirect(url_for("index"))
            prev = _login_fails.get(username, (0, 0.0))[0]
            _login_fails[username] = (prev + 1, time.time())
            error = "帳號或密碼錯誤"
    return render_page(PAGE_LOGIN, error=error, username=username, page_title="登入")


@app.route("/users")
@admin_required
def users_page(error=None, msg=None):
    rows = []
    for u in get_db().execute("SELECT * FROM users ORDER BY id").fetchall():
        d = dict(u)
        d["created_local"] = fmt_local(u["created_at"])
        rows.append(d)
    return render_page(PAGE_USERS, users=rows, error=error, msg=msg, page_title="帳號管理")


@app.route("/users/new", methods=["POST"])
@admin_required
def user_new():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    is_admin = bool(request.form.get("is_admin"))
    if not username or not password:
        return users_page(error="帳號與密碼不可為空")
    policy_err = check_password_policy(password)
    if policy_err:
        return users_page(error=policy_err)
    db = get_db()
    try:
        create_user(username, password, is_admin)
        audit("新增帳號", "帳號", username, "管理員" if is_admin else "一般使用者")
        db.commit()
    except sqlite3.IntegrityError:
        return users_page(error="帳號已存在,請改用其他名稱")
    return redirect(url_for("users_page"))


@app.route("/users/password", methods=["POST"])
@admin_required
def user_password():
    uid = safe_int(request.form.get("user_id", ""))
    password = request.form.get("password", "")
    if uid is None:
        return users_page(error="請選擇要重設密碼的帳號")
    policy_err = check_password_policy(password)
    if policy_err:
        return users_page(error=policy_err)
    db = get_db()
    row = db.execute("SELECT username FROM users WHERE id = ?", (uid,)).fetchone()
    if row is None:
        return users_page(error="找不到指定的帳號")
    db.execute("UPDATE users SET password_hash = ? WHERE id = ?",
               (generate_password_hash(password), uid))
    _login_fails.pop(row["username"], None)  # 重設密碼同時解除鎖定
    audit("重設密碼", "帳號", row["username"])
    db.commit()
    return users_page(msg=f"已重設 {row['username']} 的密碼")


@app.route("/users/<int:uid>/delete", methods=["POST"])
@admin_required
def user_delete(uid):
    db = get_db()
    if uid == session.get("user_id"):
        return users_page(error="不可刪除自己的帳號")
    row = db.execute("SELECT * FROM users WHERE id = ?", (uid,)).fetchone()
    if row is None:
        return users_page(error="找不到指定的帳號")
    if row["is_admin"]:
        admins = db.execute("SELECT COUNT(*) AS c FROM users WHERE is_admin = 1").fetchone()["c"]
        if admins <= 1:
            return users_page(error="不可刪除最後一位管理員")
    # 保留其異動歷史(transactions.user_id 外鍵),僅停用帳號:改名並鎖定密碼
    has_tx = db.execute("SELECT COUNT(*) AS c FROM transactions WHERE user_id = ?",
                        (uid,)).fetchone()["c"] > 0
    if has_tx:
        db.execute("UPDATE users SET username = ?, password_hash = ?, is_admin = 0 WHERE id = ?",
                   (f"{row['username']}(已停用)", generate_password_hash(secrets.token_hex(16)), uid))
        audit("停用帳號", "帳號", row["username"], "有異動紀錄,保留歷史故改為停用")
    else:
        db.execute("DELETE FROM users WHERE id = ?", (uid,))
        audit("刪除帳號", "帳號", row["username"])
    db.commit()
    return redirect(url_for("users_page"))


@app.route("/audit")
@admin_required
def audit_page():
    q = request.args.get("q", "").strip()
    page = max(1, safe_int(request.args.get("page", "1"), 1) or 1)
    per = 200
    sql = "SELECT * FROM audit_log WHERE 1=1"
    params = []
    if q:
        sql += " AND (username LIKE ? OR action LIKE ? OR detail LIKE ? OR target_type LIKE ?)"
        params += [f"%{q}%"] * 4
    sql += " ORDER BY id DESC LIMIT ? OFFSET ?"
    params += [per + 1, (page - 1) * per]
    fetched = get_db().execute(sql, params).fetchall()
    has_next = len(fetched) > per
    rows = []
    for r in fetched[:per]:
        d = dict(r)
        d["created_local"] = fmt_local(r["created_at"])
        rows.append(d)
    return render_page(PAGE_AUDIT, rows=rows, q=q,
                       pager=build_pager("audit_page", page, has_next, q=q), page_title="稽核紀錄")


# 登出用 GET 是為了 curl 可測性的刻意簡化;本系統未實作 CSRF token,
# 屬已知取捨(內部工具、無跨站表單情境),部署對外時應補上。
@app.route("/logout")
@login_required
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------------------------------------------------------------------
# 庫存總覽(首頁)與低庫存警示
# ---------------------------------------------------------------------------

def query_products(q="", category="", limit=None, offset=0, missing=False):
    # 搜尋同時比對我方 SKU、名稱與各公司別名料號(跨公司料號整合的核心)
    sql = """
        SELECT p.*, s.name AS supplier_name,
               (SELECT GROUP_CONCAT(a.company || ':' || a.alias_sku, ' / ')
                FROM part_aliases a WHERE a.product_id = p.id) AS alias_text
        FROM products p LEFT JOIN suppliers s ON p.supplier_id = s.id
        WHERE 1=1
    """
    params = []
    if q:
        sql += """ AND (p.name LIKE ? OR p.sku LIKE ? OR p.location LIKE ? OR EXISTS (
                     SELECT 1 FROM part_aliases a WHERE a.product_id = p.id
                       AND (a.alias_sku LIKE ? OR a.company LIKE ?)))"""
        params += [f"%{q}%"] * 5
    if category:
        sql += " AND p.category = ?"
        params.append(category)
    if missing:
        # 資料待補:沒有儲位(現場找不到東西在哪)或還掛著搬遷時自動編的臨時料號
        sql += " AND (p.location IS NULL OR p.location = '' OR p.sku LIKE 'TMP-%')"
    sql += " ORDER BY p.id"
    if limit is not None:      # 匯出端點傳 limit=None 取得完整資料
        sql += " LIMIT ? OFFSET ?"
        params += [limit + 1, offset]   # 多取一筆用來判斷是否還有下一頁
    rows = [dict(r) for r in get_db().execute(sql, params).fetchall()]
    resv = reserved_map()
    onord = onorder_map()
    for r in rows:
        r["unit_price_str"] = fmt_money(r["unit_price"])
        r["reserved"] = resv.get(r["id"], 0)
        r["available"] = r["quantity"] - r["reserved"]   # 業界的 on-hand vs available 區分
        r["onorder"] = onord.get(r["id"], 0)             # 已下訂尚未入庫(在途)
        r["low"] = bool(r["low_stock_threshold"]) and r["available"] <= r["low_stock_threshold"]
    return rows


INDEX_PER_PAGE = 100


def render_index(error=None, msg=None):
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    cols_full = request.args.get("cols", "") == "full"
    missing = request.args.get("missing", "") == "1"
    page = max(1, safe_int(request.args.get("page", "1"), 1) or 1)
    db = get_db()
    fetched = query_products(q, category, limit=INDEX_PER_PAGE,
                             offset=(page - 1) * INDEX_PER_PAGE,
                             missing=missing)
    has_next = len(fetched) > INDEX_PER_PAGE
    rows = fetched[:INDEX_PER_PAGE]
    categories = [c["category"] for c in db.execute(
        "SELECT DISTINCT category FROM products WHERE category != '' ORDER BY category"
    ).fetchall()]
    return render_page(PAGE_INDEX, rows=rows, q=q, category=category,
                       categories=categories, cols_full=cols_full, missing=missing,
                       page=page,
                       pager=build_pager("index", page, has_next, q=q, category=category,
                                         **({"cols": "full"} if cols_full else {}),
                                         **({"missing": "1"} if missing else {})),
                       error=error, msg=msg, page_title="庫存總覽")


@app.route("/")
@login_required
def index():
    return render_index()


@app.route("/alerts")
@login_required
def alerts():
    db = get_db()
    resv = reserved_map()
    onord = onorder_map()
    rows = []
    # 低庫存以「可用量」判斷:已被預留的量不該算成可動用庫存
    for r in db.execute("""
            SELECT p.*, s.name AS supplier_name
            FROM products p LEFT JOIN suppliers s ON p.supplier_id = s.id
            WHERE p.low_stock_threshold > 0
            ORDER BY p.quantity ASC
        """).fetchall():
        d = dict(r)
        d["reserved"] = resv.get(r["id"], 0)
        d["available"] = r["quantity"] - d["reserved"]
        d["onorder"] = onord.get(r["id"], 0)
        if d["available"] <= r["low_stock_threshold"]:
            # 建議補量:以補貨規劃推導的再訂購點為目標,扣掉可用量與已在途的量
            _ss, rop = suggest_safety_stock(r, usage_stats(r["id"]))
            target = rop or r["low_stock_threshold"]
            d["suggest"] = max(0, target - d["available"] - d["onorder"])
            rows.append(d)
    # 即將到期批次(30 天內)與已過期批次:FEFO 料件的呆滯與報廢風險
    horizon = (today_local() + timedelta(days=30)).isoformat()
    today = today_local().isoformat()
    expiring = []
    for r in db.execute("""
            SELECT l.*, p.name, p.sku, p.unit FROM lots l JOIN products p ON l.product_id = p.id
            WHERE l.qty_remaining > 0 AND l.expiry_date != '' AND l.expiry_date <= ?
            ORDER BY l.expiry_date
        """, (horizon,)).fetchall():
        d = dict(r)
        d["expired"] = r["expiry_date"] < today
        try:
            d["days_left"] = (datetime.strptime(r["expiry_date"], "%Y-%m-%d").date()
                              - today_local()).days
        except ValueError:
            d["days_left"] = 0
        expiring.append(d)
    has_threshold = db.execute(
        "SELECT COUNT(*) AS c FROM products WHERE low_stock_threshold > 0").fetchone()["c"]
    return render_page(PAGE_ALERTS, products=rows, expiring=expiring,
                       has_threshold=has_threshold, page_title="短缺與效期")


# ---------------------------------------------------------------------------
# 商品管理
# ---------------------------------------------------------------------------

def parse_product_form():
    # 回傳 (欄位 dict, 錯誤訊息或 None)
    f = {
        "name": request.form.get("name", "").strip(),
        "sku": request.form.get("sku", "").strip(),
        "category": request.form.get("category", "").strip(),
        "unit": request.form.get("unit", "").strip() or "個",
        "unit_price": request.form.get("unit_price", "").strip() or "0",
        "low_stock_threshold": request.form.get("low_stock_threshold", "").strip() or "0",
        "supplier_id": request.form.get("supplier_id", "").strip(),
        "location": request.form.get("location", "").strip(),
        "purchase_unit": request.form.get("purchase_unit", "").strip(),
        "units_per_purchase": request.form.get("units_per_purchase", "").strip() or "1",
        "lead_time_days": request.form.get("lead_time_days", "").strip() or "7",
        "service_level": request.form.get("service_level", "").strip() or "95",
        "issue_strategy": request.form.get("issue_strategy", "FIFO").strip(),
    }
    if not f["name"] or not f["sku"]:
        return f, "商品名稱與 SKU 不可為空"
    upp = safe_int(f["units_per_purchase"])
    if upp is None or upp < 1 or upp > MAX_QUANTITY:
        return f, "每採購單位的換算數量必須為 1 以上的整數"
    f["upp_val"] = upp
    lead = safe_int(f["lead_time_days"])
    if lead is None or lead < 1 or lead > 3650:
        return f, "採購前置期必須為 1~3650 天"
    f["lead_val"] = lead
    try:
        f["service_val"] = float(f["service_level"])
        if not math.isfinite(f["service_val"]) or not (50 <= f["service_val"] <= 99.9):
            raise ValueError
    except (ValueError, TypeError):
        return f, "服務水準必須介於 50 與 99.9 之間"
    if f["issue_strategy"] not in ("FIFO", "FEFO"):
        f["issue_strategy"] = "FIFO"
    try:
        f["unit_price_val"] = float(f["unit_price"])
        # math.isfinite 擋掉 inf / nan:否則會污染報表金額與 ABC 分級
        if not math.isfinite(f["unit_price_val"]) or f["unit_price_val"] < 0:
            raise ValueError
        if f["unit_price_val"] > MAX_QUANTITY:
            return f, "單價過大,請確認輸入"
    except (ValueError, TypeError, OverflowError):
        return f, "單價必須為有效的非負數字"
    try:
        f["threshold_val"] = int(f["low_stock_threshold"])
        if f["threshold_val"] < 0:
            raise ValueError
        if f["threshold_val"] > MAX_QUANTITY:
            return f, "低庫存門檻過大,請確認輸入"
    except (ValueError, TypeError, OverflowError):
        return f, "低庫存門檻必須為非負整數"
    f["supplier_val"] = safe_int(f["supplier_id"])
    return f, None


def supplier_options():
    return get_db().execute("SELECT id, name FROM suppliers ORDER BY name").fetchall()


EMPTY_PRODUCT_FORM = {"name": "", "sku": "", "category": "", "unit": "個",
                      "unit_price": "0", "low_stock_threshold": "0", "supplier_id": "",
                      "location": "", "purchase_unit": "", "units_per_purchase": "1",
                      "lead_time_days": "7", "service_level": "95", "issue_strategy": "FIFO"}


@app.route("/products/new", methods=["GET", "POST"])
@login_required
def product_new():
    f = dict(EMPTY_PRODUCT_FORM)
    error = None
    if request.method == "POST":
        f, error = parse_product_form()
        if not error:
            db = get_db()
            try:
                cur = db.execute("""
                    INSERT INTO products (name, sku, category, unit, unit_price,
                                          low_stock_threshold, supplier_id, created_at,
                                          location, purchase_unit, units_per_purchase,
                                          lead_time_days, service_level, issue_strategy)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (f["name"], f["sku"], f["category"], f["unit"], f["unit_price_val"],
                      f["threshold_val"], f["supplier_val"], now_str(),
                      f["location"], f["purchase_unit"], f["upp_val"],
                      f["lead_val"], f["service_val"], f["issue_strategy"]))
                audit("新增商品", "商品", cur.lastrowid, f"{f['sku']} {f['name']}")
                db.commit()
                return redirect(url_for("index"))
            except sqlite3.IntegrityError:
                error = "SKU 已存在,請改用其他編號"
    return render_page(PAGE_PRODUCT_FORM, title="新增商品", f=f,
                       supplier_list=supplier_options(), error=error)


@app.route("/products/<int:pid>/edit", methods=["GET", "POST"])
@login_required
def product_edit(pid):
    db = get_db()
    row = db.execute("SELECT * FROM products WHERE id = ?", (pid,)).fetchone()
    if row is None:
        return render_index(error="找不到指定的商品")
    error = None
    if request.method == "POST":
        f, error = parse_product_form()
        if not error:
            try:
                db.execute("""
                    UPDATE products SET name=?, sku=?, category=?, unit=?, unit_price=?,
                                        low_stock_threshold=?, supplier_id=?, location=?,
                                        purchase_unit=?, units_per_purchase=?,
                                        lead_time_days=?, service_level=?, issue_strategy=?
                    WHERE id=?
                """, (f["name"], f["sku"], f["category"], f["unit"], f["unit_price_val"],
                      f["threshold_val"], f["supplier_val"], f["location"],
                      f["purchase_unit"], f["upp_val"], f["lead_val"],
                      f["service_val"], f["issue_strategy"], pid))
                audit("編輯商品", "商品", pid,
                      f"{f['sku']} {f['name']} 單價={f['unit_price_val']} 門檻={f['threshold_val']}")
                db.commit()
                return redirect(url_for("index"))
            except sqlite3.IntegrityError:
                error = "SKU 已存在,請改用其他編號"
    else:
        f = {"name": row["name"], "sku": row["sku"], "category": row["category"],
             "unit": row["unit"], "unit_price": fmt_num(row["unit_price"]),
             "low_stock_threshold": str(row["low_stock_threshold"]),
             "supplier_id": "" if row["supplier_id"] is None else str(row["supplier_id"]),
             "location": row["location"] or "", "purchase_unit": row["purchase_unit"] or "",
             "units_per_purchase": str(row["units_per_purchase"] or 1),
             "lead_time_days": str(row["lead_time_days"] or 7),
             "service_level": fmt_num(row["service_level"] or 95),
             "issue_strategy": row["issue_strategy"] or "FIFO"}
    return render_page(PAGE_PRODUCT_FORM, title="編輯商品", f=f,
                       supplier_list=supplier_options(), error=error)


@app.route("/products/<int:pid>/delete", methods=["POST"])
@admin_required
def product_delete(pid):
    db = get_db()
    has_tx = db.execute(
        "SELECT COUNT(*) AS c FROM transactions WHERE product_id = ?", (pid,)
    ).fetchone()["c"] > 0
    if has_tx:
        # 有異動紀錄的商品不可刪除,以保留完整歷史
        return render_index(error="此商品已有異動紀錄,無法刪除")
    prow = db.execute("SELECT name, sku FROM products WHERE id = ?", (pid,)).fetchone()
    # 先刪實體照片檔(DB 列由 FK CASCADE 處理,但檔案系統不會自動清)
    for img in db.execute("SELECT filename FROM product_images WHERE product_id = ?", (pid,)).fetchall():
        path = os.path.join(IMAGE_DIR, img["filename"])
        try:
            if os.path.exists(path):
                os.remove(path)
        except OSError:
            pass
    db.execute("DELETE FROM products WHERE id = ?", (pid,))
    if prow:
        audit("刪除商品", "商品", pid, f"{prow['sku']} {prow['name']}")
    db.commit()
    return redirect(url_for("index"))


# ---------------------------------------------------------------------------
# 商品詳細頁:照片、跨公司料號別名、近期異動
# ---------------------------------------------------------------------------

def render_product_detail(pid, error=None, msg=None):
    db = get_db()
    row = db.execute("""
        SELECT p.*, s.name AS supplier_name
        FROM products p LEFT JOIN suppliers s ON p.supplier_id = s.id
        WHERE p.id = ?
    """, (pid,)).fetchone()
    if row is None:
        return render_index(error="找不到指定的商品")
    p = dict(row)
    p["unit_price_str"] = fmt_money(p["unit_price"])
    p["reserved"] = reserved_qty(pid)
    p["available"] = p["quantity"] - p["reserved"]
    images = db.execute(
        "SELECT * FROM product_images WHERE product_id = ? ORDER BY id", (pid,)).fetchall()
    aliases = db.execute(
        "SELECT * FROM part_aliases WHERE product_id = ? ORDER BY company, alias_sku", (pid,)).fetchall()
    recent_tx = []
    for r in db.execute("""
            SELECT t.*, u.username FROM transactions t JOIN users u ON t.user_id = u.id
            WHERE t.product_id = ? ORDER BY t.id DESC LIMIT 10
        """, (pid,)).fetchall():
        d = dict(r)
        d["created_local"] = fmt_local(r["created_at"])
        recent_tx.append(d)
    # 批次庫存(FIFO 順序)+ 庫齡天數
    now_utc = datetime.now(timezone.utc)
    lots = []
    for l in db.execute(
            "SELECT * FROM lots WHERE product_id = ? ORDER BY received_at, id", (pid,)).fetchall():
        d = dict(l)
        d["age_days"] = (now_utc - parse_utc(l["received_at"])).days
        d["received_local"] = fmt_local(l["received_at"])
        d["cost_str"] = fmt_money(l["unit_cost"]) if l["unit_cost"] is not None else "—"
        lots.append(d)
    # 在途採購:這項料還有多少在路上、哪幾張單
    open_pos = []
    for r in db.execute(f"""
            SELECT o.id, o.po_no, o.status, o.eta, i.ordered_qty, i.received_qty
            FROM purchase_order_items i JOIN purchase_orders o ON i.po_id = o.id
            WHERE i.product_id = ? AND o.status IN ({','.join('?' * len(ONORDER_STATUSES))})
            ORDER BY o.id DESC
        """, (pid, *ONORDER_STATUSES)).fetchall():
        d = dict(r)
        d["onorder"] = max(0, r["ordered_qty"] - r["received_qty"])
        d["status_label"] = PO_STATUS_LABEL.get(r["status"], r["status"])
        open_pos.append(d)
    return render_page(PAGE_PRODUCT_DETAIL, p=p, images=images, aliases=aliases,
                       recent_tx=recent_tx, lots=lots, open_pos=open_pos,
                       onorder_total=sum(d["onorder"] for d in open_pos),
                       error=error, msg=msg, page_title="商品明細", back_url=url_for('index'), back_label="回庫存總覽")


@app.route("/products/<int:pid>")
@login_required
def product_detail(pid):
    return render_product_detail(pid)


@app.route("/products/<int:pid>/images", methods=["POST"])
@login_required
def image_upload(pid):
    db = get_db()
    if db.execute("SELECT 1 FROM products WHERE id = ?", (pid,)).fetchone() is None:
        return render_index(error="找不到指定的商品")
    file = request.files.get("photo")
    if not file or not file.filename:
        return render_product_detail(pid, error="請先選擇要上傳的照片檔案")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ALLOWED_IMAGE_EXTS:
        return render_product_detail(
            pid, error="不支援的檔案格式,僅接受 jpg / jpeg / png / gif / webp / bmp")
    data = file.read()
    phash = ""
    if HAS_PIL:
        try:
            phash = compute_dhash(io.BytesIO(data))
        except Exception:
            return render_product_detail(pid, error="無法讀取圖片內容,請確認檔案未損壞")
    fname = f"img_{pid}_{secrets.token_hex(8)}.{ext}"
    with open(os.path.join(IMAGE_DIR, fname), "wb") as f:
        f.write(data)
    db.execute("""
        INSERT INTO product_images (product_id, filename, original_name, phash, created_at)
        VALUES (?, ?, ?, ?, ?)
    """, (pid, fname, file.filename, phash, now_str()))
    audit("上傳照片", "商品", pid, file.filename)
    db.commit()
    return redirect(url_for("product_detail", pid=pid))


@app.route("/products/<int:pid>/images/<int:img_id>/delete", methods=["POST"])
@admin_required
def image_delete(pid, img_id):
    db = get_db()
    row = db.execute(
        "SELECT * FROM product_images WHERE id = ? AND product_id = ?", (img_id, pid)).fetchone()
    if row is not None:
        path = os.path.join(IMAGE_DIR, row["filename"])
        if os.path.exists(path):
            os.remove(path)
        db.execute("DELETE FROM product_images WHERE id = ?", (img_id,))
        audit("刪除照片", "商品", pid, row["filename"])
        db.commit()
    return redirect(url_for("product_detail", pid=pid))


@app.route("/images/<path:filename>")
@login_required
def serve_image(filename):
    return send_from_directory(IMAGE_DIR, filename)


@app.route("/products/<int:pid>/aliases", methods=["POST"])
@login_required
def alias_add(pid):
    db = get_db()
    if db.execute("SELECT 1 FROM products WHERE id = ?", (pid,)).fetchone() is None:
        return render_index(error="找不到指定的商品")
    company = request.form.get("company", "").strip()
    alias_sku = request.form.get("alias_sku", "").strip()
    note = request.form.get("note", "").strip()
    if not company or not alias_sku:
        return render_product_detail(pid, error="公司名稱與該公司料號不可為空")
    try:
        db.execute("""
            INSERT INTO part_aliases (product_id, company, alias_sku, note, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (pid, company, alias_sku, note, now_str()))
        audit("新增別名", "商品", pid, f"{company}:{alias_sku}")
        db.commit()
    except sqlite3.IntegrityError:
        return render_product_detail(pid, error="此公司+料號組合已存在,同一組別名不可重複登記")
    return redirect(url_for("product_detail", pid=pid))


@app.route("/products/<int:pid>/aliases/<int:aid>/delete", methods=["POST"])
@admin_required
def alias_delete(pid, aid):
    db = get_db()
    arow = db.execute("SELECT company, alias_sku FROM part_aliases WHERE id = ?", (aid,)).fetchone()
    db.execute("DELETE FROM part_aliases WHERE id = ? AND product_id = ?", (aid, pid))
    if arow:
        audit("刪除別名", "商品", pid, f"{arow['company']}:{arow['alias_sku']}")
    db.commit()
    return redirect(url_for("product_detail", pid=pid))


# ---------------------------------------------------------------------------
# 入庫 / 出庫
# ---------------------------------------------------------------------------

PICK_LIMIT = 20


def pick_products(q):
    """選料查詢:沿用首頁同一組條件(我方料號/品名/儲位/跨公司別名),最多 20 筆。
    取代舊的 product_dropdown()——那是無 WHERE、無 LIMIT 的全表查詢,
    實測在正式資料上產生 2,281 個 <option>,現場根本挑不到料。"""
    q = (q or "").strip()
    if not q:
        # 沒有關鍵字時給「最近登記過的料」,現場最常見的就是連續處理同一批
        return get_db().execute("""
            SELECT p.*, s.name AS supplier_name FROM products p
            LEFT JOIN suppliers s ON p.supplier_id = s.id
            WHERE p.id IN (SELECT product_id FROM transactions ORDER BY id DESC LIMIT 60)
            ORDER BY (SELECT MAX(id) FROM transactions t WHERE t.product_id = p.id) DESC
            LIMIT ?""", (PICK_LIMIT,)).fetchall()
    return query_products(q=q, limit=PICK_LIMIT)


def product_or_none(pid):
    p = safe_int(str(pid))
    if p is None:
        return None
    row = get_db().execute("""
        SELECT p.*, s.name AS supplier_name FROM products p
        LEFT JOIN suppliers s ON p.supplier_id = s.id WHERE p.id = ?""", (p,)).fetchone()
    if row is None:
        return None
    d = dict(row)
    d["reserved"] = reserved_map().get(d["id"], 0)
    d["available"] = d["quantity"] - d["reserved"]
    return d



# ---------------------------------------------------------------------------
# 選料(共用元件):先找料 → 再做事。取代全站五處的兩千項下拉選單
# ---------------------------------------------------------------------------

PICK_TARGETS = {
    "in":      ("入庫登記 · 先找到料", "stock_in"),
    "out":     ("出庫登記 · 先找到料", "stock_out"),
    "history": ("異動歷史 · 先找到料", "history"),
    "reserve": ("建立預留 · 先找到料", "reservations_page"),
    "order_new": ("建立採購單 · 先找到料", "order_new"),
}


@app.route("/pick")
@login_required
def pick():
    to = request.args.get("to", "in")
    q = request.args.get("q", "").strip()
    extra, post_to, remember = {}, None, False
    if to.startswith("order:"):
        _, oid, item_id = to.split(":")
        title = "指定我方商品 · 採購單"
        post_to = url_for("order_item_map", oid=int(oid), item_id=int(item_id))
        remember = True
    elif to.startswith("receipt:"):
        _, rid, item_id = to.split(":")
        title = "指定我方商品 · 收貨單"
        post_to = url_for("receipt_item_map", rid=int(rid), item_id=int(item_id))
        remember = True
    elif to in PICK_TARGETS:
        title, endpoint = PICK_TARGETS[to]
    else:
        to, (title, endpoint) = "in", PICK_TARGETS["in"]

    rows = pick_products(q)
    # 掃碼槍的情境:掃到唯一一筆就直接進下一步,不用再點一次
    if q and len(rows) == 1 and post_to is None:
        return redirect(url_for(PICK_TARGETS[to][1], product_id=rows[0]["id"]))
    link_base = "" if post_to else url_for(PICK_TARGETS[to][1]) + "?product_id="
    return render_page(PAGE_PICK, title=title, page_title=title, narrow=True,
                       to=to, q=q, rows=rows, extra=extra, post_to=post_to,
                       remember=remember, link_base=link_base,
                       truncated=(len(rows) >= PICK_LIMIT))


def parse_stock_form():
    f = {
        "product_id": request.form.get("product_id", "").strip(),
        "quantity": request.form.get("quantity", "").strip(),
        "note": request.form.get("note", "").strip(),
        "lot_no": request.form.get("lot_no", "").strip(),
        "unit_cost": request.form.get("unit_cost", "").strip(),
        "purpose": request.form.get("purpose", "").strip(),
        "expiry_date": request.form.get("expiry_date", "").strip(),
        "qty_unit": request.form.get("qty_unit", "stock").strip(),
    }
    qty = safe_int(f["quantity"])
    if qty is None or qty <= 0:
        return f, "數量必須為正整數"
    if qty > MAX_QUANTITY:
        return f, f"數量過大,單筆不可超過 {MAX_QUANTITY:,},請確認輸入"
    if safe_int(f["product_id"]) is None:
        return f, "請選擇商品"
    if f["unit_cost"]:
        try:
            f["unit_cost_val"] = float(f["unit_cost"])
            if not math.isfinite(f["unit_cost_val"]) or f["unit_cost_val"] < 0:
                raise ValueError
            if f["unit_cost_val"] > MAX_QUANTITY:
                return f, "成本單價過大,請確認輸入"
        except (ValueError, TypeError, OverflowError):
            return f, "成本單價必須為有效的非負數字"
    else:
        f["unit_cost_val"] = None
    return f, None


EMPTY_STOCK_FORM = {"product_id": "", "quantity": "", "note": "", "lot_no": "",
                    "unit_cost": "", "purpose": "", "expiry_date": "", "qty_unit": "stock"}


def stock_done_msg(action):
    # 連續登記:成功後 302 回登記頁,從 done/qty 參數組出成功訊息(含最新庫存)
    done = safe_int(request.args.get("done", ""))
    qty = safe_int(request.args.get("qty", ""))
    if done is None or qty is None:
        return None
    row = get_db().execute(
        "SELECT name, quantity, unit FROM products WHERE id = ?", (done,)).fetchone()
    if row is None:
        return None
    return f"{action}成功:{row['name']} ×{qty},目前庫存 {row['quantity']} {row['unit']},可繼續登記下一筆"


@app.route("/stock/in", methods=["GET", "POST"])
@login_required
def stock_in():
    # 第一步(沒帶 product_id)一律導到共用選料頁,而不是塞一個兩千項的下拉選單。
    # 舊網址只帶 done(登記成功的商品),仍當作 product_id 用,舊書籤才不會斷。
    if request.method == "GET" and not request.args.get("product_id"):
        if request.args.get("done"):
            return redirect(url_for("stock_in", product_id=request.args.get("done"),
                                    done=request.args.get("done"), qty=request.args.get("qty", "")))
        return redirect(url_for("pick", to="in"))
    f = dict(EMPTY_STOCK_FORM)
    f["product_id"] = request.form.get("product_id") or request.args.get("product_id", "")
    error = None
    if request.method == "POST":
        f, error = parse_stock_form()
        if not error:
            db = get_db()
            pid, qty = int(f["product_id"]), int(f["quantity"])
            prow = db.execute(
                "SELECT purchase_unit, units_per_purchase FROM products WHERE id = ?", (pid,)).fetchone()
            if prow is None:
                error = "找不到指定的商品"
            elif f["qty_unit"] == "purchase" and not (prow["purchase_unit"] or "").strip():
                error = "此商品尚未設定採購單位,請先在商品編輯頁設定,或改用庫存單位輸入"
            else:
                if f["qty_unit"] == "purchase":
                    qty = qty * max(1, prow["units_per_purchase"] or 1)
                    if qty > MAX_QUANTITY:
                        error = "換算後數量過大,請確認輸入"
            if not error:
                ts = now_str()
                db.execute("UPDATE products SET quantity = quantity + ? WHERE id = ?", (qty, pid))
                cur = db.execute("""
                    INSERT INTO transactions (product_id, user_id, type, quantity, note, purpose, created_at)
                    VALUES (?, ?, 'in', ?, ?, ?, ?)
                """, (pid, session["user_id"], qty, f["note"], f["purpose"], ts))
                tx_id = cur.lastrowid
                lot_no = f["lot_no"] or f"L{datetime.now(timezone.utc).strftime('%Y%m%d')}-{tx_id}"
                try:
                    db.execute("""
                        INSERT INTO lots (product_id, transaction_id, lot_no, qty_received,
                                          qty_remaining, unit_cost, note, received_at, expiry_date)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (pid, tx_id, lot_no, qty, qty, f["unit_cost_val"], f["note"], ts,
                          f["expiry_date"]))
                except sqlite3.IntegrityError:
                    db.rollback()
                    error = "此商品已有相同批號,請改用其他批號"
                else:
                    db.commit()
                    # 一定要把 product_id 帶回去。舊版只帶 done/qty,GET 時 product_id 是空字串,
                    # 瀏覽器就會自動選中排序第一筆的商品——成功訊息還停在上一筆,
                    # 使用者打完數量送出,貨就入到完全不相干的料號上了。
                    return redirect(url_for("stock_in", product_id=pid, done=pid, qty=qty))
    return render_stock_form(True, f, error)


@app.route("/stock/out", methods=["GET", "POST"])
@login_required
def stock_out():
    # 舊網址只帶 done(登記成功的商品),仍當作 product_id 用,舊書籤才不會斷
    if request.method == "GET" and not request.args.get("product_id"):
        if request.args.get("done"):
            return redirect(url_for("stock_out", product_id=request.args.get("done"),
                                    done=request.args.get("done"), qty=request.args.get("qty", "")))
        return redirect(url_for("pick", to="out"))
    f = dict(EMPTY_STOCK_FORM)
    f["product_id"] = request.form.get("product_id") or request.args.get("product_id", "")
    error = None
    if request.method == "POST":
        f, error = parse_stock_form()
        if not error:
            db = get_db()
            pid, qty = int(f["product_id"]), int(f["quantity"])
            row = db.execute("SELECT quantity FROM products WHERE id = ?", (pid,)).fetchone()
            if row is None:
                error = "找不到指定的商品"
            else:
                # 原子更新的條件是「可用量」而不是現貨:已被工單預留的量不該被領走。
                # 舊版寫 quantity >= ?,預留形同虛設。
                updated = db.execute("""
                    UPDATE products SET quantity = quantity - ?
                    WHERE id = ? AND quantity - COALESCE(
                        (SELECT SUM(r.quantity) FROM reservations r
                         WHERE r.product_id = products.id AND r.status = 'active'), 0) >= ?
                """, (qty, pid, qty)).rowcount
                if updated == 0:
                    db.rollback()
                    resv = reserved_map().get(pid, 0)
                    # 訊息要說實話:沒有預留時擋下來的就是庫存不足;
                    # 有預留時才是可用量不足,而且要把三個數字都寫出來,
                    # 否則使用者看到「庫存 100 卻領不到 70」會以為系統壞了
                    if resv:
                        error = (f"可用量不足,無法出庫(現貨 {row['quantity']}、"
                                 f"已預留 {resv}、可用 {row['quantity'] - resv},要求出庫 {qty})")
                    else:
                        error = (f"庫存不足,無法出庫(目前庫存:{row['quantity']},"
                                 f"要求出庫:{qty})")
                else:
                    ts = now_str()
                    cur = db.execute("""
                        INSERT INTO transactions (product_id, user_id, type, quantity, note, purpose, created_at)
                        VALUES (?, ?, 'out', ?, ?, ?, ?)
                    """, (pid, session["user_id"], qty, f["note"], f["purpose"], ts))
                    tx_id = cur.lastrowid
                    consume_lots(db, pid, qty, tx_id, ts)
                    db.commit()
                    return redirect(url_for("stock_out", product_id=pid, done=pid, qty=qty))
    return render_stock_form(False, f, error)


def recent_moves(pid, limit=10):
    """登記頁下半的「本次已登記」:最近 N 筆,每筆可直接沖銷。
    全站原本沒有任何更正入口,打錯只能反向出庫——那會讓 FIFO 扣到錯的批,
    把批號、成本與效期一起改壞。"""
    db = get_db()
    rows = db.execute("""
        SELECT t.*, u.username,
               CASE WHEN t.type = 'in'
                    THEN (SELECT l.lot_no FROM lots l WHERE l.transaction_id = t.id)
                    ELSE (SELECT GROUP_CONCAT(l2.lot_no || '×' || c.quantity, '、')
                          FROM lot_consumptions c JOIN lots l2 ON c.lot_id = l2.id
                          WHERE c.transaction_id = t.id)
               END AS lot_info,
               (SELECT 1 FROM transactions r WHERE r.reverses = t.id) AS reversed
        FROM transactions t JOIN users u ON t.user_id = u.id
        WHERE t.product_id = ? ORDER BY t.id DESC LIMIT ?
    """, (pid, limit)).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d["local_time"] = fmt_local(r["created_at"])
        d["is_reversal"] = bool(r["reverses"])
        out.append(d)
    return out


def render_stock_form(is_in, f, error):
    prod = product_or_none(f.get("product_id"))
    if prod is None:
        return redirect(url_for("pick", to="in" if is_in else "out"))
    title = "入庫登記" if is_in else "出庫登記"
    return render_page(PAGE_STOCK_FORM, title=title, page_title=f"{title}:{prod['name']}",
                       narrow=True, f=f, is_in=is_in, prod=prod, error=error,
                       recent=recent_moves(prod["id"]),
                       msg=stock_done_msg(title[:2]),
                       back_url=url_for("pick", to="in" if is_in else "out"),
                       back_label="換一項料")


@app.route("/transactions/<int:tid>/reverse", methods=["POST"])
@login_required
def transaction_reverse(tid):
    """沖銷一筆異動:把庫存與批次帳都還原到登記前,而不是再開一筆反向異動。
    反向出庫會讓 FIFO 去扣真正最早的那一批,把批號、成本與效期一起改壞。"""
    db = get_db()
    t = db.execute("SELECT * FROM transactions WHERE id = ?", (tid,)).fetchone()
    back = url_for("stock_in" if (t and t["type"] == "in") else "stock_out",
                   product_id=(t["product_id"] if t else 0))
    if t is None:
        return render_page(PAGE_PICK, title="沖銷", narrow=True,
                           to="in", q="", rows=[], extra={}, post_to=None,
                           remember=False, link_base="", truncated=False,
                           error="找不到這筆異動")

    def fail(msg):
        f = dict(EMPTY_STOCK_FORM); f["product_id"] = str(t["product_id"])
        return render_stock_form(t["type"] == "in", f, msg)

    if t["reverses"]:
        return fail("這筆本身就是沖銷紀錄,不能再沖銷")
    if db.execute("SELECT 1 FROM transactions WHERE reverses = ?", (tid,)).fetchone():
        return fail("這筆異動已經沖銷過了")
    if t["user_id"] != session.get("user_id") and not session.get("is_admin"):
        return render_page(
            "<h1>沒有權限</h1><p>只有登記者本人或管理員可以沖銷這筆異動。</p>"
            "<p><a class=\"btn ghost\" href=\"{{ url_for('index') }}\">回庫存總覽</a></p>",
            title="沒有權限", page_title="沒有權限", narrow=True), 403
    ts = now_str()
    if t["type"] == "in":
        lot = db.execute("SELECT * FROM lots WHERE transaction_id = ?", (tid,)).fetchone()
        if lot is None:
            return fail("找不到這筆入庫建立的批次,無法安全沖銷")
        if lot["qty_remaining"] != lot["qty_received"]:
            used = lot["qty_received"] - lot["qty_remaining"]
            return fail(f"這批已經被領用 {used},不能整筆沖銷。請改以出庫更正,或先沖銷後續的出庫")
        upd = db.execute("UPDATE products SET quantity = quantity - ? WHERE id = ? AND quantity >= ?",
                         (t["quantity"], t["product_id"], t["quantity"])).rowcount
        if upd == 0:
            db.rollback()
            return fail("目前庫存已低於這筆入庫量,無法沖銷")
        db.execute("DELETE FROM lots WHERE id = ?", (lot["id"],))
        rtype = "out"
    else:
        for c in db.execute("SELECT * FROM lot_consumptions WHERE transaction_id = ?", (tid,)).fetchall():
            db.execute("UPDATE lots SET qty_remaining = qty_remaining + ? WHERE id = ?",
                       (c["quantity"], c["lot_id"]))
        db.execute("DELETE FROM lot_consumptions WHERE transaction_id = ?", (tid,))
        db.execute("UPDATE products SET quantity = quantity + ? WHERE id = ?",
                   (t["quantity"], t["product_id"]))
        rtype = "in"
    db.execute("""
        INSERT INTO transactions (product_id, user_id, type, quantity, note, purpose, created_at, reverses)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (t["product_id"], session["user_id"], rtype, t["quantity"],
          f"沖銷異動 #{tid}", t["purpose"] or "", ts, tid))
    audit("沖銷異動", "transaction", tid,
          f"{'入庫' if t['type'] == 'in' else '出庫'} {t['quantity']}")
    db.commit()
    return redirect(back + f"&done_reverse={tid}" if "?" in back else back)


# ---------------------------------------------------------------------------
# 異動歷史
# ---------------------------------------------------------------------------

def history_filters():
    return {
        "product_id": request.args.get("product_id", "").strip(),
        "type": request.args.get("type", "").strip(),
        "start": request.args.get("start", "").strip(),
        "end": request.args.get("end", "").strip(),
        "purpose": request.args.get("purpose", "").strip(),
        "pq": request.args.get("pq", "").strip(),
    }


def query_transactions(filters, limit=None, offset=0):
    # lot_info:入庫顯示建立的批號;出庫顯示 FIFO 消耗明細(批號×數量)
    # limit=None 代表不分頁(CSV 匯出用);有值時多取一筆以判斷還有沒有下一頁
    sql = """
        SELECT t.*, p.name AS product_name, p.sku, p.unit, u.username,
               CASE WHEN t.type = 'in'
                    THEN (SELECT l.lot_no FROM lots l WHERE l.transaction_id = t.id)
                    ELSE (SELECT GROUP_CONCAT(l2.lot_no || '×' || c.quantity, '、')
                          FROM lot_consumptions c JOIN lots l2 ON c.lot_id = l2.id
                          WHERE c.transaction_id = t.id)
               END AS lot_info
        FROM transactions t
        JOIN products p ON t.product_id = p.id
        JOIN users u ON t.user_id = u.id
        WHERE 1=1
    """
    params = []
    pid = safe_int(filters["product_id"])
    if pid is not None:
        sql += " AND t.product_id = ?"
        params.append(pid)
    elif filters.get("pq"):
        # 文字篩選商品:比對我方料號、品名、儲位(取代舊的兩千項下拉選單)
        like = f"%{filters['pq']}%"
        sql += " AND (p.sku LIKE ? OR p.name LIKE ? OR p.location LIKE ?)"
        params += [like, like, like]
    if filters["type"] in ("in", "out"):
        sql += " AND t.type = ?"
        params.append(filters["type"])
    if filters["purpose"]:
        sql += " AND t.purpose LIKE ?"
        params.append(f"%{filters['purpose']}%")
    # 使用者填的是台灣日期,DB 存 UTC,需換算區間才不會整體偏移 8 小時
    start_utc, end_utc = local_date_to_utc_range(filters["start"], filters["end"])
    if start_utc:
        sql += " AND t.created_at >= ?"
        params.append(start_utc)
    if end_utc:
        sql += " AND t.created_at <= ?"
        params.append(end_utc)
    sql += " ORDER BY t.id DESC"
    if limit is not None:      # 匯出端點傳 limit=None 取得完整資料
        sql += " LIMIT ? OFFSET ?"
        params += [limit + 1, offset]   # 多取一筆用來判斷是否還有下一頁
    return get_db().execute(sql, params).fetchall()


HISTORY_PER_PAGE = 200


@app.route("/history")
@login_required
def history():
    filters = history_filters()
    page = max(1, safe_int(request.args.get("page", "1"), 1) or 1)
    fetched = query_transactions(filters, limit=HISTORY_PER_PAGE,
                                 offset=(page - 1) * HISTORY_PER_PAGE)
    has_next = len(fetched) > HISTORY_PER_PAGE
    rows = []
    for r in fetched[:HISTORY_PER_PAGE]:
        d = dict(r)
        d["created_local"] = fmt_local(r["created_at"])
        rows.append(d)
    return render_page(PAGE_HISTORY, rows=rows, filters=filters, title="異動歷史",
                       picked_product=product_or_none(filters["product_id"]),
                       pager=build_pager("history", page, has_next, **filters))


# ---------------------------------------------------------------------------
# 供應商管理
# ---------------------------------------------------------------------------

EMPTY_SUPPLIER_FORM = {"name": "", "contact": "", "phone": "", "note": ""}


def parse_supplier_form():
    f = {
        "name": request.form.get("name", "").strip(),
        "contact": request.form.get("contact", "").strip(),
        "phone": request.form.get("phone", "").strip(),
        "note": request.form.get("note", "").strip(),
    }
    if not f["name"]:
        return f, "供應商名稱不可為空"
    return f, None


@app.route("/suppliers")
@login_required
def suppliers():
    rows = get_db().execute("""
        SELECT s.*, (SELECT COUNT(*) FROM products p WHERE p.supplier_id = s.id) AS product_count
        FROM suppliers s ORDER BY s.id
    """).fetchall()
    return render_page(PAGE_SUPPLIERS, rows=rows, page_title="供應商")


@app.route("/suppliers/new", methods=["GET", "POST"])
@login_required
def supplier_new():
    f = dict(EMPTY_SUPPLIER_FORM)
    error = None
    if request.method == "POST":
        f, error = parse_supplier_form()
        if not error:
            db = get_db()
            cur = db.execute(
                "INSERT INTO suppliers (name, contact, phone, note, created_at) VALUES (?, ?, ?, ?, ?)",
                (f["name"], f["contact"], f["phone"], f["note"], now_str()),
            )
            audit("新增供應商", "供應商", cur.lastrowid, f["name"])
            db.commit()
            return redirect(url_for("suppliers"))
    return render_page(PAGE_SUPPLIER_FORM, title="新增供應商", f=f, error=error)


@app.route("/suppliers/<int:sid>/edit", methods=["GET", "POST"])
@login_required
def supplier_edit(sid):
    db = get_db()
    row = db.execute("SELECT * FROM suppliers WHERE id = ?", (sid,)).fetchone()
    if row is None:
        return redirect(url_for("suppliers"))
    error = None
    if request.method == "POST":
        f, error = parse_supplier_form()
        if not error:
            db.execute(
                "UPDATE suppliers SET name=?, contact=?, phone=?, note=? WHERE id=?",
                (f["name"], f["contact"], f["phone"], f["note"], sid),
            )
            audit("編輯供應商", "供應商", sid, f["name"])
            db.commit()
            return redirect(url_for("suppliers"))
    else:
        f = {"name": row["name"], "contact": row["contact"],
             "phone": row["phone"], "note": row["note"]}
    return render_page(PAGE_SUPPLIER_FORM, title="編輯供應商", f=f, error=error)


@app.route("/suppliers/<int:sid>/delete", methods=["POST"])
@admin_required
def supplier_delete(sid):
    db = get_db()
    # FK ON DELETE SET NULL:商品的供應商欄位自動清空
    srow = db.execute("SELECT name FROM suppliers WHERE id = ?", (sid,)).fetchone()
    db.execute("DELETE FROM suppliers WHERE id = ?", (sid,))
    if srow:
        audit("刪除供應商", "供應商", sid, srow["name"])
    db.commit()
    return redirect(url_for("suppliers"))


# ---------------------------------------------------------------------------
# 報表
# ---------------------------------------------------------------------------


def report_verdicts(rows, aging, accuracy_info, total_value):
    """報表最上面的四張判讀卡:數字必須附一句白話結論。
    算不出來就說算不出來並解釋原因——顯示一個 0 會被當成真的。"""
    db = get_db()
    total = len(rows)
    priced = sum(1 for r in rows if (r["unit_price"] or 0) > 0)
    stale = next((a["count"] for a in aging if a["key"] == "90+"), 0)
    lots_ok = not db.execute("""
        SELECT 1 FROM products p WHERE p.quantity <> COALESCE(
            (SELECT SUM(l.qty_remaining) FROM lots l WHERE l.product_id = p.id), 0) LIMIT 1
    """).fetchone()
    out = []
    if accuracy_info:
        acc = float(accuracy_info["accuracy_str"] or 0)
        tone = "good" if acc >= 95 else "warn" if acc >= 90 else "bad"
        say = ("達到業界標準(95~99%)。" if acc >= 95 else
               "低於業界標準 95%,建議提高盤點頻率。" if acc >= 90 else
               "帳實差距偏大,建議先盤 A 類與異動頻繁的料。")
        out.append({"label": "帳實相符率", "value": accuracy_info["accuracy_str"] + "%",
                    "say": say + f"依據:{accuracy_info['name']}", "tone": tone,
                    "id": "accuracy-value", "small": False})
    else:
        out.append({"label": "帳實相符率", "value": "尚無盤點紀錄",
                    "say": "做完第一次循環盤點後,這裡會出現準確率(業界標準 95~99%)。",
                    "tone": "", "id": "", "small": True})
    if total:
        pct = round(stale * 100 / total)
        out.append({"label": "呆滯庫存(90 天以上未動)", "value": f"{stale:,} 項",
                    "say": (f"占全部 {total:,} 項的 {pct}%。" +
                            ("多數是搬遷進來的期初庫存,尚未產生新異動。" if pct >= 80 else
                             "建議檢視是否可以退料、改用或報廢。")),
                    "tone": "bad" if pct >= 50 else "warn" if pct >= 20 else "good",
                    "id": "", "small": False})
    if priced == 0 and total:
        out.append({"label": "庫存總價值", "value": "無法計算",
                    "say": f"{total:,} 項的單價都是 0。要看價值必須先補單價,ABC 分析同理。",
                    "tone": "", "id": "", "small": True})
    else:
        out.append({"label": "庫存總價值", "value": total_value,
                    "say": (f"依現時庫存與單價計算;{total - priced:,} 項未設單價未計入。"
                            if priced < total else "全部品項都有單價,價值可信。"),
                    "tone": "good", "id": "", "small": False})
    out.append({"label": "批次帳一致性", "value": "100%" if lots_ok else "有帳差",
                "say": ("每項商品的批次剩餘量加總與現貨完全相符,沒有帳差。" if lots_ok else
                        "有商品的批次剩餘量與現貨不符,請檢查稽核紀錄。"),
                "tone": "good" if lots_ok else "bad", "id": "", "small": False})
    return out


@app.route("/report")
@login_required
def report():
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()
    cond, params = "", []
    start_utc, end_utc = local_date_to_utc_range(start, end)
    if start_utc:
        cond += " AND t.created_at >= ?"
        params.append(start_utc)
    if end_utc:
        cond += " AND t.created_at <= ?"
        params.append(end_utc)
    db = get_db()
    rows = [dict(r) for r in db.execute(f"""
        SELECT p.id, p.sku, p.name, p.quantity, p.unit_price,
               COALESCE(SUM(CASE WHEN t.type = 'in'  THEN t.quantity END), 0) AS total_in,
               COALESCE(SUM(CASE WHEN t.type = 'out' THEN t.quantity END), 0) AS total_out
        FROM products p
        LEFT JOIN transactions t ON t.product_id = p.id {cond}
        GROUP BY p.id ORDER BY p.id
    """, params).fetchall()]
    # 加權平均成本(存貨計價):僅以尚有剩餘且有登記成本的批次計算
    avg_costs = {r["product_id"]: r for r in db.execute("""
        SELECT product_id,
               SUM(qty_remaining * unit_cost) * 1.0 / SUM(qty_remaining) AS avg_cost
        FROM lots WHERE qty_remaining > 0 AND unit_cost IS NOT NULL
        GROUP BY product_id
    """).fetchall()}
    for r in rows:
        r["net"] = r["total_in"] - r["total_out"]
        r["value"] = r["quantity"] * r["unit_price"]
        r["unit_price_str"] = fmt_money(r["unit_price"])
        r["value_str"] = fmt_money(r["value"])
        ac = avg_costs.get(r["id"])
        r["avg_cost_str"] = fmt_money(ac["avg_cost"]) if ac else "—"
    # ABC 分析(柏拉圖法則):依庫存價值累積占比分級,A ≤80%、B ≤95%、其餘 C
    total_value = sum(r["value"] for r in rows)
    abc_rows = sorted(rows, key=lambda r: r["value"], reverse=True)
    cum = 0.0
    for r in abc_rows:
        # 分級依據是「本項之前的累積占比」:跨越 80% 門檻的那一項本身仍屬 A
        # (柏拉圖法則的重點是找出構成多數價值的少數關鍵料號;若改用本項之後的
        #  累積量判斷,單一高價品占比就會超過 80% 而使 A 級從缺,分級失去意義)
        prev_pct = (cum / total_value * 100) if total_value > 0 else 100.0
        cum += r["value"]
        cum_pct = (cum / total_value * 100) if total_value > 0 else 100.0
        r["pct_str"] = fmt_num(r["value"] / total_value * 100) if total_value > 0 else "0"
        r["cum_pct_str"] = fmt_num(cum_pct)
        r["abc"] = "A" if prev_pct < 80 else ("B" if prev_pct < 95 else "C")
    # 庫齡分析(Inventory Aging):在庫批次依庫齡分桶
    now_utc = datetime.now(timezone.utc)
    buckets = [["0-30 天", 0, 30, 0, 0, "0-30"], ["31-60 天", 31, 60, 0, 0, "31-60"],
               ["61-90 天", 61, 90, 0, 0, "61-90"], ["90 天以上", 91, None, 0, 0, "90+"]]
    for l in db.execute("SELECT qty_remaining, received_at FROM lots WHERE qty_remaining > 0").fetchall():
        age = (now_utc - parse_utc(l["received_at"])).days
        for b in buckets:
            if age >= b[1] and (b[2] is None or age <= b[2]):
                b[3] += l["qty_remaining"]
                b[4] += 1              # 批數:長條與判讀卡用「幾項」而不是「幾件」
                break
    aging_total = sum(b[3] for b in buckets)
    count_total = sum(b[4] for b in buckets) or 1
    # 顏色由「意思」決定:越舊越接近警示色,不是依位置輪流換色
    tone = {"0-30": "var(--onhand)", "31-60": "var(--onhand)",
            "61-90": "var(--amber)", "90+": "var(--fault)"}
    aging = [{"label": b[0], "qty": b[3], "count": b[4], "key": b[5],
              "pct": round(b[4] * 100 / count_total, 1),
              "color": tone[b[5]],
              "pct_str": fmt_num(b[3] / aging_total * 100) if aging_total > 0 else "0"}
             for b in buckets]
    # XYZ 分級(依需求變異係數)與 ABC-XYZ 組合:價值之外再看穩定度
    for r in abc_rows:
        stats = usage_stats(r["id"])
        xyz, cv = xyz_class(stats)
        r["xyz"] = xyz
        r["cv_str"] = fmt_num(cv) if cv is not None else "—"
        r["abc_xyz"] = f"{r['abc']}{xyz}" if xyz != "—" else r["abc"]
    # 庫存準確率:最近一次已過帳盤點的相符比例(業界基準 95–99%)
    last_count = db.execute("""
        SELECT name, accuracy, posted_at FROM stock_counts
        WHERE status = 'posted' AND accuracy IS NOT NULL ORDER BY id DESC LIMIT 1
    """).fetchone()
    accuracy_info = None
    if last_count:
        accuracy_info = {"name": last_count["name"],
                         "accuracy_str": fmt_num(last_count["accuracy"]),
                         "posted_local": fmt_local(last_count["posted_at"])}
    return render_page(
        PAGE_REPORT, rows=rows, start=start, end=end,
        abc_rows=abc_rows, aging=aging, accuracy_info=accuracy_info,
        total_in=sum(r["total_in"] for r in rows),
        total_out=sum(r["total_out"] for r in rows),
        total_net=sum(r["net"] for r in rows),
        total_qty=sum(r["quantity"] for r in rows),
        total_value_str=fmt_money(total_value), page_title="庫存報表",
        verdicts=report_verdicts(rows, aging, accuracy_info, fmt_money(total_value)),
    )


# ---------------------------------------------------------------------------
# 以圖搜圖:上傳照片 → dHash 比對庫內所有照片 → 相似度排名
# ---------------------------------------------------------------------------

@app.route("/search/image", methods=["GET", "POST"])
@login_required
def image_search():
    results = None
    error = None
    if request.method == "POST":
        if not HAS_PIL:
            error = "此功能需要安裝 Pillow 套件"
        else:
            file = request.files.get("photo")
            if not file or not file.filename:
                error = "請先選擇要搜尋的照片"
            else:
                try:
                    query_hash = compute_dhash(io.BytesIO(file.read()))
                except Exception:
                    error = "無法讀取圖片內容,請確認檔案為有效的圖片"
                    query_hash = None
                if query_hash:
                    rows = get_db().execute("""
                        SELECT i.id, i.filename, i.phash, i.product_id,
                               p.sku, p.name, p.quantity, p.unit
                        FROM product_images i JOIN products p ON i.product_id = p.id
                        WHERE i.phash != ''
                    """).fetchall()
                    scored = []
                    for r in rows:
                        dist = hamming_distance(query_hash, r["phash"])
                        scored.append({**dict(r), "distance": dist,
                                       "similarity": round((64 - dist) / 64 * 100)})
                    scored.sort(key=lambda x: x["distance"])
                    # 同一商品多張照片只留最相似的一張,列出前 10 名
                    seen, results = set(), []
                    for s in scored:
                        if s["product_id"] in seen:
                            continue
                        seen.add(s["product_id"])
                        results.append(s)
                        if len(results) >= 10:
                            break
    return render_page(PAGE_IMAGE_SEARCH, results=results, has_pil=HAS_PIL, error=error, page_title="以圖找料")


# ---------------------------------------------------------------------------
# 循環盤點:建立盤點單 → 逐項輸入實盤數 → 過帳修正庫存
# ---------------------------------------------------------------------------

SCOPE_LABELS = {"all": "全部商品", "A": "A 類", "B": "B 類", "C": "C 類", "category": "指定分類"}


def abc_class_map():
    """依庫存價值累積占比計算每個商品的 ABC 分級(與報表同一套規則)。"""
    rows = get_db().execute(
        "SELECT id, quantity * unit_price AS value FROM products").fetchall()
    total = sum(r["value"] for r in rows) or 0
    ranked = sorted(rows, key=lambda r: r["value"], reverse=True)
    result, cum = {}, 0.0
    for r in ranked:
        prev_pct = (cum / total * 100) if total > 0 else 100.0
        cum += r["value"]
        result[r["id"]] = "A" if prev_pct < 80 else ("B" if prev_pct < 95 else "C")
    return result


@app.route("/counts")
@login_required
def counts_page(error=None, msg=None):
    db = get_db()
    rows = []
    for r in db.execute("SELECT * FROM stock_counts ORDER BY id DESC LIMIT 100").fetchall():
        d = dict(r)
        d["created_local"] = fmt_local(r["created_at"])
        d["accuracy_str"] = fmt_num(r["accuracy"]) if r["accuracy"] is not None else "—"
        rows.append(d)
    categories = [r["category"] for r in db.execute(
        "SELECT DISTINCT category FROM products WHERE category != '' ORDER BY category").fetchall()]
    return render_page(PAGE_COUNTS, rows=rows, categories=categories,
                       today=today_local().isoformat(), error=error, msg=msg, page_title="循環盤點")


@app.route("/counts/new", methods=["POST"])
@admin_required
def count_new():
    name = request.form.get("name", "").strip()
    scope = request.form.get("scope", "all").strip()
    category = request.form.get("category", "").strip()
    if not name:
        return counts_page(error="盤點單名稱不可為空")
    db = get_db()
    if scope == "category":
        if not category:
            return counts_page(error="範圍選「指定分類」時必須選擇分類")
        products = db.execute(
            "SELECT id, quantity FROM products WHERE category = ? ORDER BY id", (category,)).fetchall()
        scope_label = f"分類:{category}"
    elif scope == "location":
        # 人是站在一排貨架前盤點的:依儲位前綴建單,清單才跟走位一致
        loc = request.form.get("location", "").strip()
        if not loc:
            return counts_page(error="範圍選「指定儲位」時必須填儲位前綴(例:CE、D-15)")
        products = db.execute(
            "SELECT id, quantity FROM products WHERE location LIKE ? ORDER BY location, sku",
            (f"{loc}%",)).fetchall()
        scope_label = f"儲位:{loc}"
    elif scope in ("A", "B", "C"):
        abc = abc_class_map()
        products = [p for p in db.execute("SELECT id, quantity FROM products ORDER BY id").fetchall()
                    if abc.get(p["id"]) == scope]
        scope_label = f"{scope} 類"
    else:
        products = db.execute("SELECT id, quantity FROM products ORDER BY id").fetchall()
        scope_label = "全部商品"
    if not products:
        return counts_page(error="此範圍內沒有商品,請改選其他範圍")
    ts = now_str()
    cur = db.execute("""
        INSERT INTO stock_counts (name, scope, status, username, created_at)
        VALUES (?, ?, 'open', ?, ?)
    """, (name, scope_label, session.get("username", ""), ts))
    cid = cur.lastrowid
    for p in products:
        # 建立當下就固定系統帳,之後的進出不影響本次盤點的比較基準
        db.execute("""
            INSERT INTO stock_count_items (count_id, product_id, system_qty)
            VALUES (?, ?, ?)
        """, (cid, p["id"], p["quantity"]))
    audit("建立盤點單", "盤點", cid, f"{name}({scope_label},{len(products)} 項)")
    db.commit()
    return redirect(url_for("count_detail", cid=cid))


COUNT_PER_PAGE = 50


def render_count_detail(cid, error=None, msg=None):
    db = get_db()
    c = db.execute("SELECT * FROM stock_counts WHERE id = ?", (cid,)).fetchone()
    if c is None:
        return counts_page(error="找不到指定的盤點單")
    cd = dict(c)
    cd["created_local"] = fmt_local(c["created_at"])
    cd["posted_local"] = fmt_local(c["posted_at"]) if c["posted_at"] else ""
    cd["accuracy_str"] = fmt_num(c["accuracy"]) if c["accuracy"] is not None else "—"
    sort = request.args.get("sort", "")
    filt = request.args.get("filter", "")
    page = max(1, safe_int(request.args.get("page", "1"), 1) or 1)

    def rows(limit=None, offset=0):
        # 人是站在一排貨架前盤點的,依儲位排序時清單順序才跟走位一致
        order = "p.location, p.sku" if sort == "loc" else "p.id"
        where = ""
        if filt == "todo":
            where = " AND i.counted_qty IS NULL"
        elif filt == "diff":
            where = " AND i.counted_qty IS NOT NULL AND i.counted_qty <> i.system_qty"
        sql = f"""SELECT i.*, p.sku, p.name, p.location FROM stock_count_items i
                  JOIN products p ON i.product_id = p.id
                  WHERE i.count_id = ?{where} ORDER BY {order}"""
        params = [cid]
        if limit is not None:
            sql += " LIMIT ? OFFSET ?"
            params += [limit + 1, offset]
        return db.execute(sql, params).fetchall()

    # 列印模式:紙本盤點單只印一頁 50 列是沒有用的——人拿著半張單走進倉庫,
    # 剩下的料就沒人盤。列印時一次帶出全部明細,不分頁。
    printing = request.args.get("print") == "1"
    if printing:
        fetched = rows()
        has_next = False
        chunk = fetched
    else:
        fetched = rows(COUNT_PER_PAGE, (page - 1) * COUNT_PER_PAGE)
        has_next = len(fetched) > COUNT_PER_PAGE
        chunk = fetched[:COUNT_PER_PAGE]
    items = []
    for r in chunk:
        d = dict(r)
        d["diff"] = (r["counted_qty"] - r["system_qty"]) if r["counted_qty"] is not None else None
        items.append(d)
    tot = db.execute("SELECT COUNT(*) AS c FROM stock_count_items WHERE count_id = ?", (cid,)).fetchone()["c"]
    counted = db.execute("SELECT COUNT(*) AS c FROM stock_count_items WHERE count_id = ? AND counted_qty IS NOT NULL",
                         (cid,)).fetchone()["c"]
    diff_count = db.execute("""SELECT COUNT(*) AS c FROM stock_count_items
                               WHERE count_id = ? AND counted_qty IS NOT NULL AND counted_qty <> system_qty""",
                            (cid,)).fetchone()["c"]
    return render_page(PAGE_COUNT_DETAIL, c=cd, items=items, total=tot,
                       counted=counted, diff_count=diff_count,
                       page_filled=sum(1 for i in items if i["counted_qty"] is not None),
                       page=page, sort=sort, filter=filt, printing=printing,
                       pager="" if printing else build_pager("count_detail", page, has_next, cid=cid,
                                         **({"sort": sort} if sort else {}),
                                         **({"filter": filt} if filt else {})),
                       error=error, msg=msg, page_title="盤點單",
                       back_url=url_for('counts_page'), back_label="回盤點單清單")


@app.route("/counts/<int:cid>")
@login_required
def count_detail(cid):
    return render_count_detail(cid)


def count_back(cid, anchor=None):
    """儲存後回到原頁原頁碼,並定位到剛存的那一列——
    舊版一列一送出後跳回最頂端,現場在第 40 列存完要自己捲回去確認。"""
    url = url_for("count_detail", cid=cid,
                  page=request.form.get("page") or None,
                  sort=request.form.get("sort") or None,
                  filter=request.form.get("filter") or None)
    return redirect(url + (f"#i{anchor}" if anchor else ""))


@app.route("/counts/<int:cid>/count", methods=["POST"])
@login_required
def count_record(cid):
    """一次儲存整張表(每列 qty_<item_id> / note_<item_id>)。
    舊版每列各自是一張表單,打完十列按第一列的儲存,其餘九列靜默消失。
    only 參數保留單列儲存當安全網。"""
    db = get_db()
    c = db.execute("SELECT status FROM stock_counts WHERE id = ?", (cid,)).fetchone()
    if c is None:
        return counts_page(error="找不到指定的盤點單")
    if c["status"] == "posted":
        return render_count_detail(cid, error="此盤點單已過帳,不可再修改")
    ts = now_str()
    # 相容層:舊契約是 product_id + counted_qty 單列送出。批次端點本來就吃得下單列,
    # 保留它讓舊書籤、舊腳本與既有驗收條目不會斷。
    legacy_pid = safe_int(request.form.get("product_id", ""))
    if legacy_pid is not None:
        qty = safe_int(request.form.get("counted_qty", ""))
        if qty is None or qty < 0:
            return render_count_detail(cid, error="實盤數必須為 0 或正整數")
        if qty > MAX_QUANTITY:
            return render_count_detail(cid, error="實盤數過大,請確認輸入")
        db.execute("""UPDATE stock_count_items SET counted_qty = ?, note = ?, counted_at = ?
                      WHERE count_id = ? AND product_id = ?""",
                   (qty, request.form.get("note", "").strip(), ts, cid, legacy_pid))
        db.commit()
        return redirect(url_for("count_detail", cid=cid))
    only = safe_int(request.args.get("only", ""))
    saved, bad = 0, []
    for key, raw in request.form.items():
        if not key.startswith("qty_"):
            continue
        item_id = safe_int(key[4:])
        if item_id is None or (only is not None and item_id != only):
            continue
        raw = raw.strip()
        note = request.form.get(f"note_{item_id}", "").strip()
        if raw == "":
            # 留白代表「這一列還沒盤」,不是盤到 0。表單一定會把既有值帶回來,
            # 所以送出空值是使用者刻意清掉,應還原成未盤狀態。
            db.execute("""UPDATE stock_count_items SET counted_qty = NULL, counted_at = NULL, note = ?
                          WHERE id = ? AND count_id = ?""", (note, item_id, cid))
            continue
        qty = safe_int(raw)
        if qty is None or qty < 0 or qty > MAX_QUANTITY:
            bad.append(item_id)
            continue
        db.execute("""UPDATE stock_count_items SET counted_qty = ?, note = ?, counted_at = ?
                      WHERE id = ? AND count_id = ?""", (qty, note, ts, item_id, cid))
        saved += 1
    db.commit()
    if bad:
        return render_count_detail(cid, error=f"有 {len(bad)} 列的實盤數不是 0 或正整數,那幾列沒有存進去")
    return count_back(cid, only or (request.form.get("page") and None))


@app.route("/counts/<int:cid>/fill", methods=["POST"])
@login_required
def count_fill(cid):
    """無差異全部確認:把還沒盤的列一次填成系統帳。帳面本來就正確時最常用。"""
    db = get_db()
    c = db.execute("SELECT status FROM stock_counts WHERE id = ?", (cid,)).fetchone()
    if c is None:
        return counts_page(error="找不到指定的盤點單")
    if c["status"] == "posted":
        return render_count_detail(cid, error="此盤點單已過帳,不可再修改")
    n = db.execute("""UPDATE stock_count_items SET counted_qty = system_qty, counted_at = ?
                      WHERE count_id = ? AND counted_qty IS NULL""", (now_str(), cid)).rowcount
    db.commit()
    return render_count_detail(cid, msg=f"已把 {n} 列未盤的品項填成系統帳(視為無差異)")


@app.route("/counts/<int:cid>/post", methods=["POST"])
@admin_required
def count_post(cid):
    db = get_db()
    c = db.execute("SELECT * FROM stock_counts WHERE id = ?", (cid,)).fetchone()
    if c is None:
        return counts_page(error="找不到指定的盤點單")
    if c["status"] == "posted":
        return render_count_detail(cid, error="此盤點單已過帳,不可重複過帳")
    items = db.execute("""
        SELECT i.*, p.quantity AS current_qty FROM stock_count_items i
        JOIN products p ON i.product_id = p.id
        WHERE i.count_id = ? AND i.counted_qty IS NOT NULL
    """, (cid,)).fetchall()
    if not items:
        return render_count_detail(cid, error="尚未輸入任何實盤數,無法過帳")
    ts = now_str()
    adjusted = matched = 0
    for it in items:
        diff = it["counted_qty"] - it["current_qty"]   # 以過帳當下的庫存為準才不會蓋掉期間的進出
        if diff == 0:
            matched += 1
            continue
        adjusted += 1
        pid = it["product_id"]
        note = f"盤點調整({c['name']})" + (f":{it['note']}" if it["note"] else "")
        if diff > 0:
            # 盤盈:補一筆入庫並建立調整批,批次帳才會跟著對上
            db.execute("UPDATE products SET quantity = quantity + ? WHERE id = ?", (diff, pid))
            cur = db.execute("""
                INSERT INTO transactions (product_id, user_id, type, quantity, note, purpose, created_at)
                VALUES (?, ?, 'in', ?, ?, '盤點調整', ?)
            """, (pid, session["user_id"], diff, note, ts))
            tx_id = cur.lastrowid
            db.execute("""
                INSERT INTO lots (product_id, transaction_id, lot_no, qty_received,
                                  qty_remaining, note, received_at)
                VALUES (?, ?, ?, ?, ?, '盤盈調整批', ?)
            """, (pid, tx_id, f"CNT{cid}-{tx_id}", diff, diff, ts))
        else:
            # 盤虧:走與出庫相同的批次消耗路徑,維持恆等式
            loss = -diff
            db.execute("UPDATE products SET quantity = quantity - ? WHERE id = ? AND quantity >= ?",
                       (loss, pid, loss))
            cur = db.execute("""
                INSERT INTO transactions (product_id, user_id, type, quantity, note, purpose, created_at)
                VALUES (?, ?, 'out', ?, ?, '盤點調整', ?)
            """, (pid, session["user_id"], loss, note, ts))
            consume_lots(db, pid, loss, cur.lastrowid, ts)
    total_counted = matched + adjusted
    accuracy = (matched / total_counted * 100) if total_counted else 0.0
    db.execute("""
        UPDATE stock_counts SET status = 'posted', posted_at = ?, accuracy = ? WHERE id = ?
    """, (ts, accuracy, cid))
    audit("盤點過帳", "盤點", cid,
          f"{c['name']}:盤點 {total_counted} 項,相符 {matched} 項,調整 {adjusted} 項,準確率 {accuracy:.1f}%")
    db.commit()
    return render_count_detail(cid, msg=f"過帳完成:調整 {adjusted} 項,庫存準確率 {accuracy:.1f}%")


# ---------------------------------------------------------------------------
# 採購訂單:下訂 → 出貨 → 到貨(自動生成收貨單)→ 核對放行 → 結案
# ---------------------------------------------------------------------------

ORDER_HEADER = ["料號", "品名", "數量", "單價", "預計到貨日", "備註"]
PO_STEPS = ["ordered", "shipped", "arrived", "closed"]


def parse_order_rows(db, rows, default_eta=""):
    """把資料列轉成採購明細;逐列比對料號(與收貨單共用 match_part)。"""
    items = []
    for line_no, row in rows:
        row = list(row) + [""] * (6 - len(row))
        raw_sku, raw_name, qty_s, cost_s, eta, note = [str(c).strip() for c in row[:6]]
        if not raw_sku and not raw_name:
            continue
        qty = safe_int(qty_s) or 0
        if qty < 0 or qty > MAX_QUANTITY:
            qty = 0
        try:
            cost = float(cost_s) if cost_s else None
            if cost is not None and (not math.isfinite(cost) or cost < 0 or cost > MAX_QUANTITY):
                cost = None
        except (ValueError, TypeError, OverflowError):
            cost = None
        pid, mtype, mnote = match_part(db, raw_sku, raw_name)
        items.append({"line_no": line_no, "raw_sku": raw_sku, "raw_name": raw_name,
                      "product_id": pid, "match_type": mtype, "match_note": mnote,
                      "ordered_qty": qty, "unit_cost": cost,
                      "eta": eta or default_eta, "note": note})
    return items


@app.route("/orders")
@login_required
def orders_page(error=None, msg=None):
    db = get_db()
    status = request.args.get("status", "").strip()
    where, params = "", []
    if status in PO_STATUS_LABEL:
        where, params = "WHERE o.status = ?", [status]
    rows = []
    for r in db.execute(f"""
            SELECT o.*,
                   (SELECT COUNT(*) FROM purchase_order_items i WHERE i.po_id = o.id) AS item_count,
                   (SELECT COALESCE(SUM(i.ordered_qty), 0) FROM purchase_order_items i WHERE i.po_id = o.id) AS ordered_total,
                   (SELECT COALESCE(SUM(i.received_qty), 0) FROM purchase_order_items i WHERE i.po_id = o.id) AS received_total
            FROM purchase_orders o {where} ORDER BY o.id DESC LIMIT 100
        """, params).fetchall():
        d = dict(r)
        d["created_local"] = fmt_local(r["created_at"])
        rows.append(d)
    return render_page(PAGE_ORDERS, rows=rows,
                       suppliers=db.execute("SELECT id, name FROM suppliers ORDER BY name").fetchall(),
                       summary=po_status_summary(), labels=PO_STATUS_LABEL,
                       status_filter=status if status in PO_STATUS_LABEL else "",
                       error=error, msg=msg, page_title="採購單")


@app.route("/orders/template.csv")
@login_required
def order_template():
    return csv_response(ORDER_HEADER,
                        [("ABC-001", "範例品名", "100", "12.5", "2026-09-30", "此列為範例,請刪除")],
                        "order_template.csv")


@app.route("/orders/upload", methods=["POST"])
@login_required
def order_upload():
    file = request.files.get("file")
    if not file or not file.filename:
        return orders_page(error="請先選擇採購明細檔案(Excel 或 CSV)")
    rows, error = read_table_file(file.filename, file.read())
    if error:
        return orders_page(error=error)
    db = get_db()
    eta = request.form.get("eta", "").strip()
    items = parse_order_rows(db, rows[1:], eta)
    if not items:
        return orders_page(error="檔案中沒有可讀取的明細,請確認格式:料號,品名,數量,單價,預計到貨日,備註")
    supplier_id = safe_int(request.form.get("supplier_id", ""))
    supplier_name = ""
    if supplier_id is not None:
        srow = db.execute("SELECT name FROM suppliers WHERE id = ?", (supplier_id,)).fetchone()
        supplier_name = srow["name"] if srow else ""
        if srow is None:
            supplier_id = None
    ts = now_str()
    cur = db.execute("""
        INSERT INTO purchase_orders (po_no, supplier_id, supplier_name, status, eta,
                                     note, username, created_at)
        VALUES (?, ?, ?, 'ordered', ?, ?, ?, ?)
    """, (request.form.get("po_no", "").strip(), supplier_id, supplier_name, eta,
          request.form.get("note", "").strip(), session.get("username", ""), ts))
    oid = cur.lastrowid
    for it in items:
        db.execute("""
            INSERT INTO purchase_order_items (po_id, line_no, raw_sku, raw_name, product_id,
                                              match_type, match_note, ordered_qty, unit_cost, eta, note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (oid, it["line_no"], it["raw_sku"], it["raw_name"], it["product_id"],
              it["match_type"], it["match_note"], it["ordered_qty"], it["unit_cost"],
              it["eta"], it["note"]))
    unmatched = sum(1 for i in items if i["product_id"] is None)
    audit("建立採購單", "採購", oid, f"{file.filename}:{len(items)} 列,未對應 {unmatched} 列")
    db.commit()
    return redirect(url_for("order_detail", oid=oid))


def render_order_detail(oid, error=None, msg=None):
    db = get_db()
    o = db.execute("SELECT * FROM purchase_orders WHERE id = ?", (oid,)).fetchone()
    if o is None:
        return orders_page(error="找不到指定的採購單")
    od = dict(o)
    od["created_local"] = fmt_local(o["created_at"])
    od["arrived_local"] = fmt_local(o["arrived_at"]) if o["arrived_at"] else ""
    items, ordered_total, received_total, onorder_total = [], 0, 0, 0
    for row in db.execute("""
            SELECT i.*, p.sku AS psku, p.name AS pname FROM purchase_order_items i
            LEFT JOIN products p ON i.product_id = p.id
            WHERE i.po_id = ? ORDER BY i.line_no, i.id
        """, (oid,)).fetchall():
        d = dict(row)
        label = MATCH_LABELS.get(row["match_type"], row["match_type"])
        if row["match_type"] == "alias" and row["match_note"]:
            label = f"別名:{row['match_note']}"
        d["match_label"] = label
        d["onorder"] = max(0, row["ordered_qty"] - row["received_qty"])
        ordered_total += row["ordered_qty"]
        received_total += row["received_qty"]
        onorder_total += d["onorder"]
        items.append(d)
    receipts = []
    for r in db.execute("SELECT * FROM receipts WHERE po_id = ? ORDER BY id DESC", (oid,)).fetchall():
        rd = dict(r)
        rd["created_local"] = fmt_local(r["created_at"])
        receipts.append(rd)
    step = PO_STEPS.index(o["status"]) if o["status"] in PO_STEPS else 0
    return render_page(PAGE_ORDER_DETAIL, o=od, items=items, total=len(items),
                       ordered_total=ordered_total, received_total=received_total,
                       onorder_total=onorder_total, receipts=receipts,
                       labels=PO_STATUS_LABEL, step_index=step,
                       today=today_local().isoformat(), error=error, msg=msg, page_title="採購單明細", back_url=url_for('orders_page'), back_label="回採購單清單")



@app.route("/orders/new", methods=["GET", "POST"])
@login_required
def order_new():
    """手動建立採購單(單一品項)。把系統最後一條死路接通:
    以前在短缺佇列看到要補貨,只能自己記下料號、切到採購單、用 Excel 生一個檔再上傳。"""
    db = get_db()
    prod = product_or_none(request.form.get("product_id") or request.args.get("product_id", ""))
    qty = request.form.get("ordered_qty") or request.args.get("qty", "")
    eta = request.args.get("eta", "")
    error = None
    if request.method == "POST":
        n = safe_int(str(qty))
        if prod is None:
            error = "找不到指定的商品"
        elif n is None or n <= 0:
            error = "訂購數量必須是大於 0 的整數"
        elif n > MAX_QUANTITY:
            error = "訂購數量過大,請確認輸入"
        else:
            ts = now_str()
            po_no = request.form.get("po_no", "").strip()
            supplier_name = request.form.get("supplier_name", "").strip()
            sup = db.execute("SELECT id FROM suppliers WHERE name = ?", (supplier_name,)).fetchone()
            cur = db.execute("""
                INSERT INTO purchase_orders (po_no, supplier_id, supplier_name, status, eta,
                                             note, username, created_at)
                VALUES (?, ?, ?, 'ordered', ?, ?, ?, ?)
            """, (po_no, sup["id"] if sup else None, supplier_name,
                  request.form.get("eta", "").strip(), request.form.get("note", "").strip(),
                  session.get("username", ""), ts))
            oid = cur.lastrowid
            if not po_no:
                po_no = f"PO-{datetime.now(timezone.utc).strftime('%Y%m%d')}-{oid}"
                db.execute("UPDATE purchase_orders SET po_no = ? WHERE id = ?", (po_no, oid))
            cost = request.form.get("unit_cost", "").strip()
            try:
                unit_cost = float(cost) if cost else None
                if unit_cost is not None and not math.isfinite(unit_cost):
                    unit_cost = None
            except (ValueError, OverflowError):
                unit_cost = None
            db.execute("""
                INSERT INTO purchase_order_items (po_id, line_no, raw_sku, raw_name, product_id,
                                                  match_type, match_note, ordered_qty, received_qty,
                                                  unit_cost, eta, note)
                VALUES (?, 1, ?, ?, ?, 'sku', '手動建立', ?, 0, ?, ?, '')
            """, (oid, prod["sku"], prod["name"], prod["id"], n, unit_cost,
                  request.form.get("eta", "").strip()))
            audit("建立採購單", "purchase_order", oid, f"{po_no} · {prod['sku']} × {n}")
            db.commit()
            return redirect(url_for("order_detail", oid=oid))
    return render_page(PAGE_ORDER_NEW, prod=prod, qty=qty, eta=eta, error=error,
                       title="建立採購單", page_title="建立採購單", narrow=True,
                       back_url=url_for("orders_page"), back_label="回採購單清單")


@app.route("/orders/<int:oid>")
@login_required
def order_detail(oid):
    return render_order_detail(oid)


def open_order_or_error(oid):
    """取回仍可推進的採購單;不可動時回 (None, 錯誤訊息)。"""
    o = get_db().execute("SELECT * FROM purchase_orders WHERE id = ?", (oid,)).fetchone()
    if o is None:
        return None, "找不到指定的採購單"
    if o["status"] == "closed":
        return None, "此採購單已入庫結案,不可再變更"
    if o["status"] == "cancelled":
        return None, "此採購單已作廢,不可再變更"
    return o, None


@app.route("/orders/<int:oid>/items/<int:item_id>/map", methods=["POST"])
@login_required
def order_item_map(oid, item_id):
    o, err = open_order_or_error(oid)
    if err:
        return orders_page(error=err) if "找不到" in err else render_order_detail(oid, error=err)
    db = get_db()
    pid = safe_int(request.form.get("product_id", ""))
    if pid is None:
        return render_order_detail(oid, error="請選擇要對應的我方商品")
    item = db.execute("SELECT * FROM purchase_order_items WHERE id = ? AND po_id = ?",
                      (item_id, oid)).fetchone()
    if item is None:
        return render_order_detail(oid, error="找不到指定的明細列")
    if db.execute("SELECT 1 FROM products WHERE id = ?", (pid,)).fetchone() is None:
        return render_order_detail(oid, error="找不到指定的商品")
    db.execute("UPDATE purchase_order_items SET product_id = ?, match_type = 'manual', match_note = '' WHERE id = ?",
               (pid, item_id))
    remembered = ""
    if request.form.get("remember") and o["supplier_name"] and item["raw_sku"]:
        try:
            db.execute("""
                INSERT INTO part_aliases (product_id, company, alias_sku, note, created_at)
                VALUES (?, ?, ?, ?, ?)
            """, (pid, o["supplier_name"], item["raw_sku"], f"採購單 {o['po_no'] or oid} 建立", now_str()))
            remembered = ",並記住為別名"
            audit("新增別名", "商品", pid, f"{o['supplier_name']}:{item['raw_sku']}(採購單對應)")
        except sqlite3.IntegrityError:
            pass
    db.commit()
    return render_order_detail(oid, msg=f"已對應第 {item['line_no']} 列{remembered}")


@app.route("/orders/<int:oid>/ship", methods=["POST"])
@login_required
def order_ship(oid):
    o, err = open_order_or_error(oid)
    if err:
        return orders_page(error=err) if "找不到" in err else render_order_detail(oid, error=err)
    if o["status"] != "ordered":
        return render_order_detail(oid, error="此採購單已標記過出貨,不可重複標記")
    db = get_db()
    db.execute("UPDATE purchase_orders SET status = 'shipped', shipped_at = ?, tracking_no = ? WHERE id = ?",
               (request.form.get("shipped_at", "").strip() or today_local().isoformat(),
                request.form.get("tracking_no", "").strip(), oid))
    audit("採購單出貨", "採購", oid, o["po_no"] or f"#{oid}")
    db.commit()
    return render_order_detail(oid, msg="已標記為出貨中")


@app.route("/orders/<int:oid>/arrive", methods=["POST"])
@login_required
def order_arrive(oid):
    """登記到貨,並把訂單明細自動轉成收貨單——這是「登記自動化」的核心:
    同一份明細只登記一次,現場只需核對數量,不必重新上傳或重打料號。"""
    o, err = open_order_or_error(oid)
    if err:
        return orders_page(error=err) if "找不到" in err else render_order_detail(oid, error=err)
    db = get_db()
    # 分批到貨是常態:只要還有未收足的品項就能再登記一次,但不允許同時存在
    # 兩張待核對的收貨單——那會讓現場不知道該核對哪一張,也會重複入庫
    waiting = db.execute(
        "SELECT id FROM receipts WHERE po_id = ? AND status = 'open' ORDER BY id DESC",
        (oid,)).fetchone()
    if waiting:
        return render_order_detail(
            oid, error=f"此採購單已登記過到貨,且收貨單 #{waiting['id']} 還在待核對——"
                       "請先完成該收貨單的核對與放行,再登記下一批到貨")
    items = db.execute("""
        SELECT * FROM purchase_order_items WHERE po_id = ? ORDER BY line_no, id
    """, (oid,)).fetchall()
    pending = [i for i in items if i["ordered_qty"] > i["received_qty"]]
    if not pending:
        return render_order_detail(oid, error="此採購單的品項都已收足,沒有待到貨的明細")
    ts = now_str()
    # 分批到貨時單號要能分辨是第幾批,否則兩張收貨單長得一模一樣
    batch = db.execute("SELECT COUNT(*) AS c FROM receipts WHERE po_id = ?", (oid,)).fetchone()["c"] + 1
    base = o["po_no"] or f"PO#{oid}"
    ref = f"{base}-到貨" if batch == 1 else f"{base}-到貨第{batch}批"
    cur = db.execute("""
        INSERT INTO receipts (ref_no, supplier_id, supplier_name, source, status,
                              note, username, created_at, po_id)
        VALUES (?, ?, ?, ?, 'open', ?, ?, ?, ?)
    """, (ref, o["supplier_id"], o["supplier_name"], f"採購單到貨自動建立(#{oid},第 {batch} 批)",
          o["note"], session.get("username", ""), ts, oid))
    rid = cur.lastrowid
    for it in pending:
        db.execute("""
            INSERT INTO receipt_items (receipt_id, line_no, raw_sku, raw_name, product_id,
                                       match_type, match_note, expected_qty, lot_no,
                                       expiry_date, unit_cost, note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, '', '', ?, ?)
        """, (rid, it["line_no"], it["raw_sku"], it["raw_name"], it["product_id"],
              it["match_type"], it["match_note"], it["ordered_qty"] - it["received_qty"],
              it["unit_cost"], it["note"]))
    db.execute("UPDATE purchase_orders SET status = 'arrived', arrived_at = ? WHERE id = ?", (ts, oid))
    audit("採購單到貨", "採購", oid,
          f"{o['po_no'] or '#' + str(oid)}:自動建立收貨單 #{rid},{len(pending)} 個品項")
    db.commit()
    return redirect(url_for("receipt_detail", rid=rid))


@app.route("/orders/<int:oid>/cancel", methods=["POST"])
@admin_required
def order_cancel(oid):
    db = get_db()
    o = db.execute("SELECT * FROM purchase_orders WHERE id = ?", (oid,)).fetchone()
    if o is None:
        return orders_page(error="找不到指定的採購單")
    if o["status"] == "closed":
        return render_order_detail(oid, error="此採購單已結案,不可作廢")
    db.execute("UPDATE purchase_orders SET status = 'cancelled' WHERE id = ?", (oid,))
    audit("採購單作廢", "採購", oid, o["po_no"] or f"#{oid}")
    db.commit()
    return redirect(url_for("order_detail", oid=oid))


def writeback_po_receipt(db, receipt_row, posted):
    """收貨單放行後回寫來源採購單的已收量;全數收齊即自動結案。
    posted 為 {product_id: 實收量} 的累計。"""
    po_id = receipt_row["po_id"] if "po_id" in receipt_row.keys() else None
    if not po_id or not posted:
        return
    items = db.execute("SELECT * FROM purchase_order_items WHERE po_id = ? ORDER BY line_no, id",
                       (po_id,)).fetchall()
    remaining = dict(posted)
    for it in items:
        pid = it["product_id"]
        if pid is None or remaining.get(pid, 0) <= 0:
            continue
        need = it["ordered_qty"] - it["received_qty"]
        if need <= 0:
            continue
        take = min(need, remaining[pid])
        db.execute("UPDATE purchase_order_items SET received_qty = received_qty + ? WHERE id = ?",
                   (take, it["id"]))
        remaining[pid] -= take
    still = db.execute("""
        SELECT COALESCE(SUM(CASE WHEN ordered_qty > received_qty
                                 THEN ordered_qty - received_qty ELSE 0 END), 0) AS left_qty
        FROM purchase_order_items WHERE po_id = ?
    """, (po_id,)).fetchone()["left_qty"]
    if still == 0:
        db.execute("UPDATE purchase_orders SET status = 'closed', closed_at = ? WHERE id = ?",
                   (now_str(), po_id))
        audit("採購單結案", "採購", po_id, "全部品項已入庫")


# ---------------------------------------------------------------------------
# 收貨單(ASN):供應商檔案 → 預先登記 → 逐項核對 → 放行才入庫
# ---------------------------------------------------------------------------

RECEIPT_HEADER = ["料號", "品名", "數量", "批號", "效期", "單價", "備註"]
MATCH_LABELS = {"sku": "我方料號", "alias": "別名對應", "manual": "人工指定", "none": "未對應"}


def match_part(db, raw_sku, raw_name=""):
    """把檔案上的料號對應到我方商品:先比我方 SKU,再比跨公司別名,最後比品名。
    回傳 (product_id, match_type, match_note)。"""
    if raw_sku:
        row = db.execute("SELECT id FROM products WHERE sku = ? COLLATE NOCASE", (raw_sku,)).fetchone()
        if row:
            return row["id"], "sku", ""
        # 供應商用自己的料號時,靠既有的跨公司料號對照找回我方商品
        row = db.execute(
            "SELECT product_id, company FROM part_aliases WHERE alias_sku = ? COLLATE NOCASE",
            (raw_sku,)).fetchone()
        if row:
            return row["product_id"], "alias", row["company"]
    if raw_name:
        row = db.execute("SELECT id FROM products WHERE name = ? COLLATE NOCASE", (raw_name,)).fetchone()
        if row:
            return row["id"], "sku", "品名相符"
    return None, "none", ""


def parse_receipt_rows(db, rows):
    """把資料列轉成收貨明細;逐列比對料號。回傳 [(line_no, dict)]。"""
    items = []
    for line_no, row in rows:
        row = list(row) + [""] * (7 - len(row))
        raw_sku, raw_name, qty_s, lot_no, expiry, cost_s, note = [str(c).strip() for c in row[:7]]
        if not raw_sku and not raw_name:
            continue
        qty = safe_int(qty_s) or 0
        if qty < 0 or qty > MAX_QUANTITY:
            qty = 0
        try:
            cost = float(cost_s) if cost_s else None
            if cost is not None and (not math.isfinite(cost) or cost < 0 or cost > MAX_QUANTITY):
                cost = None
        except (ValueError, TypeError, OverflowError):
            cost = None
        pid, mtype, mnote = match_part(db, raw_sku, raw_name)
        items.append({
            "line_no": line_no, "raw_sku": raw_sku, "raw_name": raw_name,
            "product_id": pid, "match_type": mtype, "match_note": mnote,
            "expected_qty": qty, "lot_no": lot_no, "expiry_date": expiry,
            "unit_cost": cost, "note": note,
        })
    return items


@app.route("/receipts")
@login_required
def receipts_page(error=None, msg=None):
    db = get_db()
    rows = []
    for r in db.execute("""
            SELECT r.*,
                   (SELECT COUNT(*) FROM receipt_items i WHERE i.receipt_id = r.id) AS item_count,
                   (SELECT COUNT(*) FROM receipt_items i
                     WHERE i.receipt_id = r.id AND i.received_qty IS NOT NULL) AS checked_count
            FROM receipts r ORDER BY r.id DESC LIMIT 100
        """).fetchall():
        d = dict(r)
        d["created_local"] = fmt_local(r["created_at"])
        rows.append(d)
    suppliers = db.execute("SELECT id, name FROM suppliers ORDER BY name").fetchall()
    return render_page(PAGE_RECEIPTS, rows=rows, suppliers=suppliers, error=error, msg=msg, page_title="收貨單")


@app.route("/receipts/template.csv")
@login_required
def receipt_template():
    # 空白範例檔:使用者可直接把這個格式轉給供應商填
    return csv_response(RECEIPT_HEADER,
                        [("ABC-001", "範例品名", "100", "LOT-2026A", "2027-12-31", "12.5", "此列為範例,請刪除")],
                        "receipt_template.csv")


@app.route("/receipts/upload", methods=["POST"])
@login_required
def receipt_upload():
    file = request.files.get("file")
    if not file or not file.filename:
        return receipts_page(error="請先選擇到貨明細檔案(Excel 或 CSV)")
    rows, error = read_table_file(file.filename, file.read())
    if error:
        return receipts_page(error=error)
    db = get_db()
    items = parse_receipt_rows(db, rows[1:])   # 略過標題列
    if not items:
        return receipts_page(error="檔案中沒有可讀取的明細資料,請確認格式:料號,品名,數量,批號,效期,單價,備註")
    supplier_id = safe_int(request.form.get("supplier_id", ""))
    supplier_name = ""
    if supplier_id is not None:
        srow = db.execute("SELECT name FROM suppliers WHERE id = ?", (supplier_id,)).fetchone()
        supplier_name = srow["name"] if srow else ""
        if srow is None:
            supplier_id = None
    ts = now_str()
    cur = db.execute("""
        INSERT INTO receipts (ref_no, supplier_id, supplier_name, source, status,
                              note, username, created_at)
        VALUES (?, ?, ?, ?, 'open', ?, ?, ?)
    """, (request.form.get("ref_no", "").strip(), supplier_id, supplier_name,
          f"檔案上傳:{file.filename}", request.form.get("note", "").strip(),
          session.get("username", ""), ts))
    rid = cur.lastrowid
    for it in items:
        db.execute("""
            INSERT INTO receipt_items (receipt_id, line_no, raw_sku, raw_name, product_id,
                                       match_type, match_note, expected_qty, lot_no,
                                       expiry_date, unit_cost, note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (rid, it["line_no"], it["raw_sku"], it["raw_name"], it["product_id"],
              it["match_type"], it["match_note"], it["expected_qty"], it["lot_no"],
              it["expiry_date"], it["unit_cost"], it["note"]))
    unmatched = sum(1 for i in items if i["product_id"] is None)
    audit("建立收貨單", "收貨", rid,
          f"{file.filename}:{len(items)} 列,未對應 {unmatched} 列")
    db.commit()
    return redirect(url_for("receipt_detail", rid=rid))


def render_receipt_detail(rid, error=None, msg=None):
    db = get_db()
    r = db.execute("SELECT * FROM receipts WHERE id = ?", (rid,)).fetchone()
    if r is None:
        return receipts_page(error="找不到指定的收貨單")
    rd = dict(r)
    rd["created_local"] = fmt_local(r["created_at"])
    rd["posted_local"] = fmt_local(r["posted_at"]) if r["posted_at"] else ""
    items = []
    for row in db.execute("""
            SELECT i.*, p.sku AS psku, p.name AS pname FROM receipt_items i
            LEFT JOIN products p ON i.product_id = p.id
            WHERE i.receipt_id = ? ORDER BY i.line_no, i.id
        """, (rid,)).fetchall():
        d = dict(row)
        label = MATCH_LABELS.get(row["match_type"], row["match_type"])
        if row["match_type"] == "alias" and row["match_note"]:
            label = f"別名:{row['match_note']}"
        d["match_label"] = label
        items.append(d)
    checked = sum(1 for i in items if i["received_qty"] is not None)
    unmatched = sum(1 for i in items if i["product_id"] is None)
    total_qty = sum(i["expected_qty"] for i in items)
    skipped = "、".join(str(i["line_no"]) for i in items
                        if i["product_id"] is None or not (i["received_qty"] or 0))
    # 收貨單本來就不分頁,列印模式只是換一句說明並收掉操作類連結
    return render_page(PAGE_RECEIPT_DETAIL, r=rd, items=items, total=len(items),
                       checked=checked, unmatched=unmatched, total_qty=total_qty,
                       printing=request.args.get("print") == "1",
                       skipped=skipped, error=error, msg=msg, page_title="收貨單明細", back_url=url_for('receipts_page'), back_label="回收貨單清單")


@app.route("/receipts/<int:rid>")
@login_required
def receipt_detail(rid):
    return render_receipt_detail(rid)


def open_receipt_or_error(rid):
    """取回可編輯的收貨單;不可編輯時回 (None, 錯誤訊息)。"""
    r = get_db().execute("SELECT * FROM receipts WHERE id = ?", (rid,)).fetchone()
    if r is None:
        return None, "找不到指定的收貨單"
    if r["status"] == "posted":
        return None, "此收貨單已放行,不可再修改"
    if r["status"] == "cancelled":
        return None, "此收貨單已作廢,不可再修改"
    return r, None


@app.route("/receipts/<int:rid>/items/<int:item_id>/map", methods=["POST"])
@login_required
def receipt_item_map(rid, item_id):
    r, err = open_receipt_or_error(rid)
    if err:
        return receipts_page(error=err) if r is None and "找不到" in err else render_receipt_detail(rid, error=err)
    db = get_db()
    pid = safe_int(request.form.get("product_id", ""))
    if pid is None:
        return render_receipt_detail(rid, error="請選擇要對應的我方商品")
    item = db.execute("SELECT * FROM receipt_items WHERE id = ? AND receipt_id = ?",
                      (item_id, rid)).fetchone()
    if item is None:
        return render_receipt_detail(rid, error="找不到指定的明細列")
    if db.execute("SELECT 1 FROM products WHERE id = ?", (pid,)).fetchone() is None:
        return render_receipt_detail(rid, error="找不到指定的商品")
    db.execute("UPDATE receipt_items SET product_id = ?, match_type = 'manual', match_note = '' WHERE id = ?",
               (pid, item_id))
    remembered = ""
    # 記住對應:把供應商的料號寫成跨公司別名,下次同一家送同樣的料就自動對上
    if request.form.get("remember") and r["supplier_name"] and item["raw_sku"]:
        try:
            db.execute("""
                INSERT INTO part_aliases (product_id, company, alias_sku, note, created_at)
                VALUES (?, ?, ?, ?, ?)
            """, (pid, r["supplier_name"], item["raw_sku"], f"收貨單 {r['ref_no'] or rid} 建立", now_str()))
            remembered = ",並記住為別名"
            audit("新增別名", "商品", pid, f"{r['supplier_name']}:{item['raw_sku']}(收貨單對應)")
        except sqlite3.IntegrityError:
            pass  # 已有相同對應,不覆蓋
    db.commit()
    return render_receipt_detail(rid, msg=f"已對應第 {item['line_no']} 列{remembered}")


@app.route("/receipts/<int:rid>/items/<int:item_id>/check", methods=["POST"])
@login_required
def receipt_item_check(rid, item_id):
    r, err = open_receipt_or_error(rid)
    if err:
        return receipts_page(error=err) if r is None and "找不到" in err else render_receipt_detail(rid, error=err)
    qty = safe_int(request.form.get("received_qty", ""))
    if qty is None or qty < 0:
        return render_receipt_detail(rid, error="實收數必須為 0 或正整數")
    if qty > MAX_QUANTITY:
        return render_receipt_detail(rid, error="實收數過大,請確認輸入")
    db = get_db()
    cur = db.execute("""
        UPDATE receipt_items SET received_qty = ?, note = ?, checked_at = ?
        WHERE id = ? AND receipt_id = ?
    """, (qty, request.form.get("note", "").strip(), now_str(), item_id, rid))
    if cur.rowcount == 0:
        return render_receipt_detail(rid, error="找不到指定的明細列")
    db.commit()
    return redirect(url_for("receipt_detail", rid=rid))



@app.route("/receipts/<int:rid>/check", methods=["POST"])
@login_required
def receipt_check_all(rid):
    """一次儲存整張收貨明細(每列 qty_<item_id> / note_<item_id>)。
    舊版每列各自是一張表單,現場照著送貨單一路打十幾列再回頭存,
    只有按下去的那一列會被存起來,其餘輸入靜默消失。
    only 參數保留單列儲存當安全網。"""
    r, err = open_receipt_or_error(rid)
    if err:
        return receipts_page(error=err) if r is None and "找不到" in err else render_receipt_detail(rid, error=err)
    db = get_db()
    ts = now_str()
    # 相容層:舊契約是 product_id + counted_qty 單列送出。批次端點本來就吃得下單列,
    # 保留它讓舊書籤、舊腳本與既有驗收條目不會斷。
    legacy_pid = safe_int(request.form.get("product_id", ""))
    if legacy_pid is not None:
        qty = safe_int(request.form.get("counted_qty", ""))
        if qty is None or qty < 0:
            return render_count_detail(cid, error="實盤數必須為 0 或正整數")
        if qty > MAX_QUANTITY:
            return render_count_detail(cid, error="實盤數過大,請確認輸入")
        db.execute("""UPDATE stock_count_items SET counted_qty = ?, note = ?, counted_at = ?
                      WHERE count_id = ? AND product_id = ?""",
                   (qty, request.form.get("note", "").strip(), ts, cid, legacy_pid))
        db.commit()
        return redirect(url_for("count_detail", cid=cid))
    only = safe_int(request.args.get("only", ""))
    saved, bad = 0, []
    for key, raw in request.form.items():
        if not key.startswith("qty_"):
            continue
        item_id = safe_int(key[4:])
        if item_id is None or (only is not None and item_id != only):
            continue
        note = request.form.get(f"note_{item_id}", "").strip()
        raw = raw.strip()
        if raw == "":
            # 留白代表「這一列還沒核對」,不是收到 0。同上,送出空值是刻意清掉。
            db.execute("""UPDATE receipt_items SET received_qty = NULL, checked_at = NULL, note = ?
                          WHERE id = ? AND receipt_id = ?""", (note, item_id, rid))
            continue
        qty = safe_int(raw)
        if qty is None or qty < 0 or qty > MAX_QUANTITY:
            bad.append(item_id)
            continue
        db.execute("""UPDATE receipt_items SET received_qty = ?, note = ?, checked_at = ?
                      WHERE id = ? AND receipt_id = ?""", (qty, note, ts, item_id, rid))
        saved += 1
    db.commit()
    if bad:
        return render_receipt_detail(rid, error=f"有 {len(bad)} 列的實收數不是 0 或正整數,那幾列沒有存進去")
    return redirect(url_for("receipt_detail", rid=rid) + (f"#i{only}" if only else ""))


@app.route("/receipts/<int:rid>/fill", methods=["POST"])
@login_required
def receipt_fill(rid):
    """全部照通知量核對:照單全收是最常見的情況,不該逐列重打一次。"""
    r, err = open_receipt_or_error(rid)
    if err:
        return receipts_page(error=err) if r is None and "找不到" in err else render_receipt_detail(rid, error=err)
    db = get_db()
    n = db.execute("""UPDATE receipt_items SET received_qty = expected_qty, checked_at = ?
                      WHERE receipt_id = ? AND product_id IS NOT NULL""", (now_str(), rid)).rowcount
    db.commit()
    return render_receipt_detail(rid, msg=f"已把 {n} 列已對應的品項填成通知量,請確認數量無誤後放行")


@app.route("/receipts/<int:rid>/post", methods=["POST"])
@login_required
def receipt_post(rid):
    db = get_db()
    r = db.execute("SELECT * FROM receipts WHERE id = ?", (rid,)).fetchone()
    if r is None:
        return receipts_page(error="找不到指定的收貨單")
    if r["status"] == "posted":
        return render_receipt_detail(rid, error="此收貨單已放行,不可重複放行")
    if r["status"] == "cancelled":
        return render_receipt_detail(rid, error="此收貨單已作廢,無法放行")
    items = db.execute("SELECT * FROM receipt_items WHERE receipt_id = ? ORDER BY line_no, id",
                       (rid,)).fetchall()
    # 有料進來卻對不到帳,是庫存失準最常見的起點,寧可擋下也不要放行
    blocked = [i for i in items if i["product_id"] is None and (i["received_qty"] or 0) > 0]
    if blocked:
        lines = "、".join(str(i["line_no"]) for i in blocked)
        return render_receipt_detail(
            rid, error=f"尚有未對應的料號(第 {lines} 列)有實收數量,請先指定我方商品再放行")
    postable = [i for i in items if i["product_id"] is not None and (i["received_qty"] or 0) > 0]
    if not postable:
        return render_receipt_detail(rid, error="沒有可入庫的明細:請先核對實收數量(實收為 0 的列會被略過)")
    ts = now_str()
    ref = r["ref_no"] or f"#{rid}"
    total_qty = 0
    posted_by_product = {}     # 供回寫來源採購單的已收量
    for it in items:
        if it["product_id"] is None or not (it["received_qty"] or 0) > 0:
            continue
        pid, qty = it["product_id"], it["received_qty"]
        posted_by_product[pid] = posted_by_product.get(pid, 0) + qty
        note = f"收貨單 {ref}" + (f":{it['note']}" if it["note"] else "")
        db.execute("UPDATE products SET quantity = quantity + ? WHERE id = ?", (qty, pid))
        cur = db.execute("""
            INSERT INTO transactions (product_id, user_id, type, quantity, note, purpose, created_at)
            VALUES (?, ?, 'in', ?, ?, ?, ?)
        """, (pid, session["user_id"], qty, note, f"收貨 {ref}", ts))
        tx_id = cur.lastrowid
        lot_no = it["lot_no"] or f"R{rid}-{tx_id}"
        try:
            db.execute("""
                INSERT INTO lots (product_id, transaction_id, lot_no, qty_received,
                                  qty_remaining, unit_cost, note, received_at, expiry_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (pid, tx_id, lot_no, qty, qty, it["unit_cost"], note, ts, it["expiry_date"] or ""))
        except sqlite3.IntegrityError:
            # 同商品批號重複:改用不會撞號的自動批號,維持批次帳恆等式
            db.execute("""
                INSERT INTO lots (product_id, transaction_id, lot_no, qty_received,
                                  qty_remaining, unit_cost, note, received_at, expiry_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (pid, tx_id, f"R{rid}-{tx_id}", qty, qty, it["unit_cost"],
                  f"{note}(原批號 {lot_no} 重複)", ts, it["expiry_date"] or ""))
        total_qty += qty
    db.execute("UPDATE receipts SET status = 'posted', posted_at = ? WHERE id = ?", (ts, rid))
    audit("收貨單放行", "收貨", rid, f"{ref}:入庫 {len(postable)} 項,合計 {total_qty}")
    writeback_po_receipt(db, r, posted_by_product)   # 有來源採購單時回寫已收量並判斷結案
    db.commit()
    return render_receipt_detail(rid, msg=f"放行完成:{len(postable)} 項已入庫,合計 {total_qty}")


@app.route("/receipts/<int:rid>/cancel", methods=["POST"])
@admin_required
def receipt_cancel(rid):
    db = get_db()
    r = db.execute("SELECT * FROM receipts WHERE id = ?", (rid,)).fetchone()
    if r is None:
        return receipts_page(error="找不到指定的收貨單")
    if r["status"] == "posted":
        return render_receipt_detail(rid, error="此收貨單已放行,不可作廢")
    db.execute("UPDATE receipts SET status = 'cancelled' WHERE id = ?", (rid,))
    audit("收貨單作廢", "收貨", rid, r["ref_no"] or f"#{rid}")
    db.commit()
    return redirect(url_for("receipt_detail", rid=rid))


# ---------------------------------------------------------------------------
# 預留 / 可用量
# ---------------------------------------------------------------------------

@app.route("/reservations")
@login_required
def reservations_page(error=None, msg=None):
    db = get_db()
    rows = []
    for r in db.execute("""
            SELECT r.*, p.name, p.sku FROM reservations r JOIN products p ON r.product_id = p.id
            WHERE r.status = 'active' ORDER BY r.id DESC
        """).fetchall():
        d = dict(r)
        d["created_local"] = fmt_local(r["created_at"])
        rows.append(d)
    return render_page(PAGE_RESERVATIONS, rows=rows, title="預留",
                       picked_product=product_or_none(request.args.get("product_id", "")),
                       error=error, msg=msg)


@app.route("/reservations/new", methods=["POST"])
@login_required
def reservation_new():
    pid = safe_int(request.form.get("product_id", ""))
    qty = safe_int(request.form.get("quantity", ""))
    purpose = request.form.get("purpose", "").strip()
    if pid is None:
        return reservations_page(error="請選擇商品")
    if qty is None or qty <= 0:
        return reservations_page(error="預留數量必須為正整數")
    if qty > MAX_QUANTITY:
        return reservations_page(error="預留數量過大,請確認輸入")
    db = get_db()
    row = db.execute("SELECT name, quantity FROM products WHERE id = ?", (pid,)).fetchone()
    if row is None:
        return reservations_page(error="找不到指定的商品")
    available = row["quantity"] - reserved_qty(pid)
    if qty > available:
        return reservations_page(
            error=f"可用量不足,無法預留(目前可用 {available},要求預留 {qty})")
    db.execute("""
        INSERT INTO reservations (product_id, quantity, purpose, username, status, created_at)
        VALUES (?, ?, ?, ?, 'active', ?)
    """, (pid, qty, purpose, session.get("username", ""), now_str()))
    audit("建立預留", "商品", pid, f"{row['name']} ×{qty}" + (f"({purpose})" if purpose else ""))
    db.commit()
    return redirect(url_for("reservations_page"))


@app.route("/reservations/<int:rid>/release", methods=["POST"])
@login_required
def reservation_release(rid):
    db = get_db()
    row = db.execute("SELECT * FROM reservations WHERE id = ?", (rid,)).fetchone()
    if row is None or row["status"] != "active":
        return reservations_page(error="找不到有效的預留紀錄")
    db.execute("UPDATE reservations SET status = 'released', released_at = ? WHERE id = ?",
               (now_str(), rid))
    audit("釋放預留", "商品", row["product_id"], f"預留 #{rid} ×{row['quantity']}")
    db.commit()
    return redirect(url_for("reservations_page"))


# ---------------------------------------------------------------------------
# 存貨規劃:安全庫存建議(依用量變異推導,取代憑印象填的固定門檻)
# ---------------------------------------------------------------------------

def planning_rows():
    rows = []
    for p in get_db().execute("SELECT * FROM products ORDER BY id").fetchall():
        d = dict(p)
        stats = usage_stats(p["id"])
        ss, rop = suggest_safety_stock(p, stats)
        xyz, cv = xyz_class(stats)
        d["stats"], d["ss"], d["rop"], d["xyz"], d["cv"] = stats, ss, rop, xyz, cv
        d["mean_str"] = fmt_num(stats["mean"]) if stats else "—"
        d["sd_str"] = fmt_num(stats["sd"]) if stats else "—"
        d["ss_str"] = str(ss) if ss is not None else "—"
        d["rop_str"] = str(rop) if rop is not None else "—"
        d["service_level_str"] = fmt_num(p["service_level"] or 95)
        rows.append(d)
    return rows


@app.route("/planning")
@login_required
def planning_page(error=None, msg=None):
    rows = planning_rows()
    return render_page(PAGE_PLANNING, rows=rows, window=USAGE_WINDOW_DAYS,
                       has_suggestion=any(r["ss"] is not None for r in rows),
                       error=error, msg=msg, page_title="補貨規劃")


@app.route("/planning/apply", methods=["POST"])
@admin_required
def planning_apply():
    db = get_db()
    applied = 0
    for r in planning_rows():
        if r["ss"] is None:
            continue
        db.execute("UPDATE products SET low_stock_threshold = ? WHERE id = ?", (r["ss"], r["id"]))
        applied += 1
    audit("套用安全庫存建議", "規劃", "", f"更新 {applied} 項商品的低庫存門檻")
    db.commit()
    return planning_page(msg=f"已將 {applied} 項商品的低庫存門檻更新為建議安全庫存")


# ---------------------------------------------------------------------------
# QR 料架標籤
# ---------------------------------------------------------------------------

@app.route("/products/<int:pid>/qr.png")
@login_required
def product_qr(pid):
    if not HAS_QRCODE:
        return Response("需要安裝 qrcode 套件", status=503, mimetype="text/plain")
    # QR 內容是該商品詳細頁的完整網址:手機掃了就直接開頁面登記進出
    target = public_base_url() + url_for("product_detail", pid=pid)
    img = qrcode.make(target)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return Response(buf.getvalue(), mimetype="image/png",
                    headers={"Cache-Control": "max-age=3600"})


# A4 直式、10mm 邊界下實測一頁約放 24 張標籤(3 欄 × 8 列)
LABELS_PER_SHEET = 24


def location_prefixes(db):
    """把儲位收斂成可以按的群組。貨架是 A-03-2 這種三段式,取第一段就是一排貨架;
    「防潮箱3層」「桌上」「Joyce保管」沒有分隔號,整串就是自己的群組。"""
    seen = []
    for r in db.execute("SELECT DISTINCT location FROM products WHERE location <> '' ORDER BY location"):
        loc = r["location"]
        pre = loc.split("-")[0] if "-" in loc else loc
        if pre not in seen:
            seen.append(pre)
    return seen


@app.route("/labels")
@login_required
def labels_page():
    """整廠 2,279 項料印出來約 98 頁。補印一個櫃子的標籤不該被迫重印整廠,
    所以這裡收儲位/分類/單一料號三個篩選,並在印之前先說清楚會印幾張幾頁。"""
    db = get_db()
    loc = request.args.get("loc", "").strip()
    category = request.args.get("category", "").strip()
    sku = request.args.get("sku", "").strip()
    where, params = [], []
    if loc:
        where.append("location LIKE ?")
        params.append(loc + "%")
    if category:
        where.append("category = ?")
        params.append(category)
    if sku:
        where.append("sku = ?")
        params.append(sku)
    sql = "SELECT id, sku, name, location FROM products"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY location, sku"
    products = db.execute(sql, params).fetchall()
    cats = [r["category"] for r in db.execute(
        "SELECT DISTINCT category FROM products WHERE category <> '' ORDER BY category")]
    sheets = (len(products) + LABELS_PER_SHEET - 1) // LABELS_PER_SHEET
    return render_page(PAGE_LABELS, products=products, has_qrcode=HAS_QRCODE,
                       loc=loc, category=category, sku=sku,
                       prefixes=location_prefixes(db), cats=cats,
                       count=len(products), sheets=sheets,
                       qr_base=public_base_url(), page_title="料架標籤")


# ---------------------------------------------------------------------------
# CSV 大量匯入(商品 / 別名),支援 UTF-8 與 Big5(cp950)
# ---------------------------------------------------------------------------

def decode_csv_bytes(raw):
    # 台灣 Excel 另存 CSV 預設 cp950;先試 utf-8-sig(含 BOM 也能吃)再試 cp950
    for enc in ("utf-8-sig", "cp950"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return None


EXCEL_EXTS = (".xlsx", ".xlsm")


def cell_to_text(value):
    # Excel 儲存格轉乾淨字串:數字欄讀回來是 float(100.0),日期是 datetime,
    # 直接 str() 會讓「數量 100」變成「100.0」而使整列匯入失敗
    if value is None:
        return ""
    if hasattr(value, "strftime"):          # date / datetime
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def read_table_file(filename, raw):
    """把上傳的表格檔讀成 [(行號, [欄位字串...])];回傳 (rows, error)。
    支援 Excel(.xlsx/.xlsm)、CSV、Tab 分隔(.tsv/.txt 或內容自動偵測)。"""
    name = (filename or "").lower()
    if name.endswith(EXCEL_EXTS):
        if not HAS_OPENPYXL:
            return None, ("此主機未安裝 openpyxl,無法讀取 Excel。"
                          "請改用 CSV,或執行 pip install -r requirements.txt 後重新啟動。")
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
                cells = [cell_to_text(c) for c in r]
                if any(cells):              # Excel 常有殘留空列,整列空白就略過
                    rows.append((i, cells))
            wb.close()
            return rows, None
        except Exception:
            return None, "Excel 檔案無法讀取,請確認檔案未損毀,且為 .xlsx 格式"
    if name.endswith(".xls"):
        return None, "不支援舊版 .xls 格式,請用 Excel 另存為 .xlsx 後再上傳"
    text = decode_csv_bytes(raw)
    if text is None:
        return None, "無法辨識檔案編碼,請使用 UTF-8 或 Big5(cp950)編碼"
    first_line = text.split("\n", 1)[0]
    # 從副檔名或首列內容判斷分隔符,讓「從系統匯出的 Tab 檔」也能直接上傳
    delimiter = "\t" if (name.endswith((".tsv", ".txt"))
                         or first_line.count("\t") > first_line.count(",")) else ","
    rows = []
    for i, r in enumerate(csv.reader(io.StringIO(text), delimiter=delimiter), start=1):
        cells = [c.strip() for c in r]
        if any(cells):
            rows.append((i, cells))
    return rows, None


def import_products_rows(rows):
    # 欄位:SKU,名稱,分類,單位,單價,低庫存門檻,供應商,初始庫存[,儲位](第一列為標題)
    # 第 9 欄儲位為選填,供舊系統搬遷時一次帶入,免得事後逐筆補
    db = get_db()
    report = []
    ok = 0
    for line_no, row in rows:
        row = list(row) + [""] * (9 - len(row))
        sku, name, category, unit, price_s, threshold_s, supplier_name, init_qty_s, location = \
            [c.strip() for c in row[:9]]
        label = f"{sku} {name}".strip()
        if not sku or not name:
            report.append({"line": line_no, "label": label or "(空列)", "status": "跳過:SKU 與名稱必填"})
            continue
        if db.execute("SELECT 1 FROM products WHERE sku = ?", (sku,)).fetchone():
            report.append({"line": line_no, "label": label, "status": "跳過:SKU 已存在"})
            continue
        try:
            price = float(price_s) if price_s else 0.0
            threshold = int(threshold_s) if threshold_s else 0
            init_qty = int(init_qty_s) if init_qty_s else 0
            if price < 0 or threshold < 0 or init_qty < 0:
                raise ValueError
            if not math.isfinite(price):
                raise ValueError
            # 超大整數會在 INSERT 時丟 OverflowError(非 ValueError),
            # 若不在此攔下,整批匯入會中斷且已成功的列全部消失
            if price > MAX_QUANTITY or threshold > MAX_QUANTITY or init_qty > MAX_QUANTITY:
                raise ValueError
        except (ValueError, TypeError, OverflowError):
            report.append({"line": line_no, "label": label, "status": "跳過:數字欄位格式錯誤或數值過大"})
            continue
        supplier_id = None
        if supplier_name:
            srow = db.execute("SELECT id FROM suppliers WHERE name = ?", (supplier_name,)).fetchone()
            if srow:
                supplier_id = srow["id"]
            else:  # 供應商不存在時自動建立,方便整批倒資料
                cur = db.execute(
                    "INSERT INTO suppliers (name, created_at) VALUES (?, ?)",
                    (supplier_name, now_str()))
                supplier_id = cur.lastrowid
        ts = now_str()
        cur = db.execute("""
            INSERT INTO products (name, sku, category, unit, unit_price,
                                  low_stock_threshold, quantity, supplier_id, location, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (name, sku, category, unit or "個", price, threshold, init_qty,
              supplier_id, location, ts))
        new_pid = cur.lastrowid
        if init_qty > 0:  # 初始庫存寫成入庫異動 + 建立批次,維持歷史與批次帳一致
            tcur = db.execute("""
                INSERT INTO transactions (product_id, user_id, type, quantity, note, created_at)
                VALUES (?, ?, 'in', ?, 'CSV匯入', ?)
            """, (new_pid, session["user_id"], init_qty, ts))
            db.execute("""
                INSERT INTO lots (product_id, transaction_id, lot_no, qty_received,
                                  qty_remaining, note, received_at)
                VALUES (?, ?, ?, ?, ?, 'CSV匯入', ?)
            """, (new_pid, tcur.lastrowid, f"IMP-{tcur.lastrowid}", init_qty, init_qty, ts))
        ok += 1
        report.append({"line": line_no, "label": label, "status": "成功"})
    db.commit()
    return ok, report


def import_aliases_rows(rows):
    # 欄位:我方SKU,公司,別名料號,備註(第一列為標題)
    db = get_db()
    report = []
    ok = 0
    for line_no, row in rows:
        row = list(row) + [""] * (4 - len(row))
        sku, company, alias_sku, note = [c.strip() for c in row[:4]]
        label = f"{sku} → {company}:{alias_sku}"
        if not sku or not company or not alias_sku:
            report.append({"line": line_no, "label": label, "status": "跳過:我方SKU、公司、別名料號必填"})
            continue
        prow = db.execute("SELECT id FROM products WHERE sku = ?", (sku,)).fetchone()
        if prow is None:
            report.append({"line": line_no, "label": label, "status": "跳過:找不到我方SKU"})
            continue
        try:
            db.execute("""
                INSERT INTO part_aliases (product_id, company, alias_sku, note, created_at)
                VALUES (?, ?, ?, ?, ?)
            """, (prow["id"], company, alias_sku, note, now_str()))
            ok += 1
            report.append({"line": line_no, "label": label, "status": "成功"})
        except sqlite3.IntegrityError:
            report.append({"line": line_no, "label": label, "status": "跳過:此公司+料號組合已存在"})
    db.commit()
    return ok, report


@app.route("/import", methods=["GET", "POST"])
@admin_required
def csv_import():
    report_rows = None
    ok_count = skip_count = 0
    error = None
    if request.method == "POST":
        mode = request.form.get("mode", "products")
        file = request.files.get("csv_file")
        if not file or not file.filename:
            error = "請先選擇檔案(Excel 或 CSV)"
        else:
            all_rows, error = read_table_file(file.filename, file.read())
            if not error:
                data_rows = all_rows[1:]  # 略過標題列
                if mode == "aliases":
                    ok_count, report_rows = import_aliases_rows(data_rows)
                else:
                    ok_count, report_rows = import_products_rows(data_rows)
                skip_count = len(report_rows) - ok_count
                db = get_db()
                audit("CSV 匯入", "匯入", mode,
                      f"檔名 {file.filename},成功 {ok_count} 筆、跳過 {skip_count} 筆")
                db.commit()
    return render_page(PAGE_IMPORT, report_rows=report_rows, ok_count=ok_count,
                       skip_count=skip_count, error=error, has_openpyxl=HAS_OPENPYXL,
                       msg=f"成功匯入 {ok_count} 筆,跳過 {skip_count} 筆" if report_rows is not None and not error else None, page_title="CSV 匯入")


# ---------------------------------------------------------------------------
# CSV 匯出(UTF-8 加 BOM,Excel 開啟中文不亂碼)
# ---------------------------------------------------------------------------

def csv_safe(value):
    # 防 CSV/DDE 公式注入:Excel 會把 = + - @ 開頭的儲存格當公式執行,
    # 前置單引號讓它保持純文字(使用者可控欄位如品名、備註都會經過這裡)
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value
    return value


def csv_response(header, data_rows, filename):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    writer.writerows([tuple(csv_safe(c) for c in row) for row in data_rows])
    return Response(
        "\ufeff" + buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}",
                 "Content-Type": "text/csv; charset=utf-8"},
    )


@app.route("/export/inventory.csv")
@login_required
def export_inventory():
    rows = query_products(limit=None)
    data = [(r["sku"], r["name"], r["location"] or "", r["category"], r["unit"],
             fmt_num(r["unit_price"]), r["quantity"], r["reserved"], r["available"],
             r["low_stock_threshold"], r["supplier_name"] or "",
             fmt_num(r["quantity"] * r["unit_price"])) for r in rows]
    filename = f"inventory_{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv"
    return csv_response(
        ["SKU", "名稱", "儲位", "分類", "單位", "單價", "現貨量", "預留量", "可用量",
         "低庫存門檻", "供應商", "庫存價值"],
        data, filename)


@app.route("/export/transactions.csv")
@login_required
def export_transactions():
    rows = query_transactions(history_filters(), limit=None)
    data = [(fmt_local(r["created_at"]), "入庫" if r["type"] == "in" else "出庫", r["sku"],
             r["product_name"], r["quantity"], r["unit"], r["lot_info"] or "",
             r["purpose"] or "", r["note"], r["username"])
            for r in rows]
    filename = f"transactions_{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv"
    return csv_response(
        ["時間(台灣)", "類型", "SKU", "商品名稱", "數量", "單位", "批次", "用途／工單", "備註", "操作人員"],
        data, filename)


# ---------------------------------------------------------------------------
# 啟動
# ---------------------------------------------------------------------------

init_db()

def cli_setup_admin(username, password):
    """建立第一個管理員(僅在完全沒有帳號時允許),供安裝腳本免開瀏覽器完成。"""
    with app.test_request_context():
        if user_count() > 0:
            row = get_db().execute("SELECT username FROM users ORDER BY id LIMIT 1").fetchone()
            print(f"  已有帳號({row['username']}…),略過建立管理員")
            return True
        err = check_password_policy(password)
        if err:
            print(f"  [錯誤] {err}")
            return False
        try:
            create_user(username, password, True)   # 不回傳值,由呼叫端 commit
        except sqlite3.IntegrityError:
            print("  [錯誤] 帳號已存在")
            return False
        audit("建立管理員", "帳號", username, "安裝腳本建立第一個管理員")
        get_db().commit()
        print(f"  已建立管理員帳號:{username}")
        return True


def cli_list_users():
    """列出所有帳號(不含密碼——密碼是單向雜湊,存的不是原文,救不回來)。
    忘記密碼的人往往連當初取的帳號名稱也忘了,所以先讓他看得到有哪些帳號。"""
    with app.test_request_context():
        rows = get_db().execute(
            "SELECT username, is_admin, created_at FROM users ORDER BY id").fetchall()
        if not rows:
            print("  這個資料庫還沒有任何帳號。直接用瀏覽器開系統註冊,第一個註冊的人就是管理員。")
            return True
        print(f"  共 {len(rows)} 個帳號:")
        for r in rows:
            role = "管理員" if r["is_admin"] else "一般同事"
            print(f"    {r['username']}    ({role},建立於 {fmt_local(r['created_at'])})")
        print("")
        print("  忘記密碼請用:python inventory_app.py --reset-password <帳號> <新密碼>")
        return True


def cli_reset_password(username, password):
    """在伺服器本機重設某個帳號的密碼。

    這是唯一的自救路徑。密碼以 werkzeug 的單向雜湊儲存,原文沒有存下來,
    任何人都不可能「查出」舊密碼——包含系統作者。所以只能改成新的。

    安全邊界是「碰得到這台電腦的檔案」:能執行這個指令的人本來就能直接
    讀寫 inventory.db,再多加一道密碼保護沒有意義。這跟系統本來的定位
    一致(公司內網、機器放在公司裡)。重設會寫進稽核紀錄。
    """
    with app.test_request_context():
        db = get_db()
        row = db.execute("SELECT id, username, is_admin FROM users WHERE username = ?",
                         (username,)).fetchone()
        if row is None:
            print(f"  [錯誤] 找不到帳號「{username}」。目前有這些帳號:")
            for r in db.execute("SELECT username FROM users ORDER BY id").fetchall():
                print(f"    {r['username']}")
            return False
        err = check_password_policy(password)
        if err:
            print(f"  [錯誤] {err}")
            return False
        db.execute("UPDATE users SET password_hash = ? WHERE id = ?",
                   (generate_password_hash(password), row["id"]))
        audit("重設密碼", "帳號", row["username"], "由伺服器本機的命令列重設")
        db.commit()
        role = "管理員" if row["is_admin"] else "一般同事"
        print(f"  已重設「{row['username']}」({role})的密碼,現在可以用新密碼登入了。")
        return True


def cli_import(path):
    """從命令列匯入商品檔,以第一位管理員名義記錄異動。
    讓安裝腳本一次跑完轉檔→匯入,使用者不必在瀏覽器裡挑檔案挑類型。"""
    if not os.path.exists(path):
        print(f"  [錯誤] 找不到檔案:{path}")
        return False
    with app.test_request_context():
        db = get_db()
        row = db.execute(
            "SELECT id, username FROM users WHERE is_admin = 1 ORDER BY id LIMIT 1").fetchone()
        if row is None:
            print("  [錯誤] 系統還沒有管理員帳號,無法匯入")
            return False
        session["user_id"], session["username"], session["is_admin"] = row["id"], row["username"], 1
        with open(path, "rb") as fh:
            rows, err = read_table_file(os.path.basename(path), fh.read())
        if err:
            print(f"  [錯誤] {err}")
            return False
        ok, report = import_products_rows(rows[1:])   # 略過標題列
        skipped = len(report) - ok
        audit("CSV 匯入", "匯入", "products", f"命令列匯入 {os.path.basename(path)},成功 {ok} 筆、跳過 {skipped} 筆")
        db.commit()
        print(f"  匯入完成:成功 {ok} 筆,跳過 {skipped} 筆")
        for r in report:
            if r["status"] != "成功":
                print(f"    第 {r['line']} 列 {r['label']}:{r['status']}")
        return True


if __name__ == "__main__":
    # 安裝腳本用的兩個子命令;都不啟動伺服器,跑完即結束
    if len(sys.argv) > 1 and sys.argv[1] == "--setup-admin":
        if len(sys.argv) < 4:
            sys.exit("用法:python inventory_app.py --setup-admin <帳號> <密碼>")
        sys.exit(0 if cli_setup_admin(sys.argv[2], sys.argv[3]) else 1)
    if len(sys.argv) > 1 and sys.argv[1] == "--import":
        if len(sys.argv) < 3:
            sys.exit("用法:python inventory_app.py --import <商品匯入檔>")
        sys.exit(0 if cli_import(sys.argv[2]) else 1)
    if len(sys.argv) > 1 and sys.argv[1] == "--list-users":
        sys.exit(0 if cli_list_users() else 1)
    if len(sys.argv) > 1 and sys.argv[1] == "--reset-password":
        if len(sys.argv) < 4:
            sys.exit("用法:python inventory_app.py --reset-password <帳號> <新密碼>\n"
                     "     不知道有哪些帳號請先跑:python inventory_app.py --list-users")
        sys.exit(0 if cli_reset_password(sys.argv[2], sys.argv[3]) else 1)
    if len(sys.argv) > 1 and sys.argv[1] == "--lan-url":
        # 啟動腳本用這個把「同事要輸入的那一行」抓進變數,再寫成文字檔。
        # 直接叫使用者看 ipconfig 會列出四五個位址(還包含虛擬網卡),挑不出來。
        u = lan_url()
        if not u:
            sys.exit("找不到內網位址,請確認這台電腦已接上公司網路")
        print(u)
        sys.exit(0)

    port = int(os.environ.get("PORT", 5000))
    start_backup_thread()   # 啟動先備份一份,之後每 24 小時自動備份
    print(f"庫存管理系統啟動中…(版本 {APP_VERSION})")
    print(f"  資料庫:{os.path.abspath(DB_PATH)}")
    print(f"  照片:{IMAGE_DIR}")
    print(f"  備份:{BACKUP_DIR}" + (f"(另同步到 {EXTRA_BACKUP_DIR})" if EXTRA_BACKUP_DIR else ""))
    # 印「可以直接貼進瀏覽器」的網址。曾經印 http://0.0.0.0:PORT 導致使用者照著
    # 輸入而得到 ERR_ADDRESS_INVALID —— 0.0.0.0 是「綁定所有網卡」的意思,不是網址。
    lan = local_ip()
    print(f"  這台電腦請開:  http://localhost:{port}")
    if lan:
        print(f"  同事請開:      http://{lan}:{port}")
    else:
        print(f"  同事請開:      http://(本機內網IP):{port}  ← 用 ipconfig 查 IPv4 位址")
    print(f"  (伺服器綁定 0.0.0.0 代表接受所有網卡連線,這串不是可輸入的網址)")
    # 空資料庫時 /register 是開著的,而且第一個註冊的人就是管理員。
    # 在公司內網,「開著」等於誰先打開網頁誰就拿到管理員權限——包含還不該有權限的人。
    if not has_any_user():
        print("")
        print("  ============================================================")
        print("  【還沒有任何帳號】第一個註冊的人會成為管理員。")
        print("  請你現在就先自己註冊完,再把網址給同事,否則誰先開誰就是管理員。")
        print("  註冊完之後系統會自動關閉自助註冊,之後的帳號由你在「帳號管理」新增。")
        print("  ============================================================")
        print("")
    try:
        # waitress:正式營運級伺服器(純 Python、Windows 友善),取代開發用伺服器
        from waitress import serve
        print(f"  伺服器:waitress,連接埠 {port}")
        serve(app, host="0.0.0.0", port=port, threads=8)
    except ImportError:
        # 未安裝 waitress 時仍可啟動,確保公司電腦離線安裝失敗也不會卡住
        print("  注意:未安裝 waitress,改用內建伺服器(建議 pip install waitress)")
        app.run(host="0.0.0.0", port=port)
